"""
GW Construction Management - Backend API
- FastAPI + SQLite (단일 파일 DB, 외부 DB 불필요)
- 직원, 현장, 배치, 출퇴근(GPS), 법인 관리
- 핵심 정책:
    * 배치는 'plan' / 'actual' / 'reported' 3종으로 분리 저장
    * 출퇴근은 GPS 좌표를 받아 현장 지오펜스 안인지 검증
"""
import os
import re
import math
import json
import sqlite3
from pathlib import Path
import secrets
from datetime import date, datetime
from contextlib import contextmanager
from typing import Optional, List

from fastapi import FastAPI, HTTPException, Query, Body, Depends, Request, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel
from starlette.middleware.sessions import SessionMiddleware

try:
    import bcrypt
except ImportError:
    bcrypt = None  # 패키지 미설치 시 임시 평문 모드 (배포 시엔 항상 설치됨)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# DB 경로 — 환경변수 DB_PATH 로 영구 디스크로 옮길 수 있음 (Render 유료 디스크 등)
DB_PATH = os.environ.get("DB_PATH") or os.path.join(os.path.dirname(os.path.abspath(__file__)), "construction.db")
FRONTEND_DIR = os.path.join(BASE_DIR, "frontend")
SECRET_KEY = os.environ.get("SECRET_KEY") or secrets.token_hex(32)

# ========================================================================
# DB
# ========================================================================
SCHEMA = """
CREATE TABLE IF NOT EXISTS companies (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL,
  business_no TEXT,
  ceo TEXT,
  license_info TEXT
);
CREATE TABLE IF NOT EXISTS sites (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  company_id INTEGER REFERENCES companies(id),
  name TEXT NOT NULL,
  address TEXT,
  latitude REAL,
  longitude REAL,
  geofence_meters INTEGER DEFAULT 200,
  contract_amount INTEGER DEFAULT 0,
  paid_amount INTEGER DEFAULT 0,
  start_date TEXT,
  end_date TEXT,
  status TEXT DEFAULT 'active',
  manager TEXT
);
CREATE TABLE IF NOT EXISTS workers (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  company_id INTEGER REFERENCES companies(id),
  name TEXT NOT NULL,
  phone TEXT,
  worker_type TEXT NOT NULL DEFAULT 'daily',  -- 'daily' | 'office'
  daily_wage INTEGER DEFAULT 0,
  job_role TEXT,
  hired_date TEXT,
  rrn_last TEXT,
  bank_account TEXT,
  note TEXT
);
CREATE TABLE IF NOT EXISTS deployments (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  worker_id INTEGER NOT NULL REFERENCES workers(id) ON DELETE CASCADE,
  site_id INTEGER REFERENCES sites(id) ON DELETE SET NULL,  -- nullable (급여 그리드는 현장 무관)
  date TEXT NOT NULL,
  kind TEXT NOT NULL DEFAULT 'plan',  -- 'plan' | 'actual' | 'reported'
  note TEXT,
  UNIQUE(worker_id, date, kind)
);
-- 나라장터 입찰 분석 (Phase 6)
CREATE TABLE IF NOT EXISTS tenders (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  tender_no TEXT UNIQUE,                  -- 공고번호 (나라장터)
  title TEXT NOT NULL,
  org_name TEXT,                          -- 발주기관
  category TEXT,                          -- 공사 종류 (전기·토목·건축·기계설비 등)
  license_required TEXT,                  -- 필요 면허 (텍스트)
  budget INTEGER DEFAULT 0,               -- 추정가격
  region TEXT,                            -- 지역
  site_address TEXT,
  posted_at TEXT,                         -- 공고일
  deadline TEXT,                          -- 입찰 마감
  bid_open_at TEXT,                       -- 개찰일
  contact TEXT,
  status TEXT DEFAULT 'open',             -- 'open'|'closed'|'awarded'|'cancelled'
  award_company TEXT,                     -- 낙찰사
  award_amount INTEGER,
  source TEXT DEFAULT 'g2b',              -- 'g2b'|'manual'|'mock'
  raw_url TEXT,                           -- 원본 공고 URL
  raw_data TEXT,                          -- 원본 JSON
  synced_at TEXT DEFAULT (datetime('now')),
  review_status TEXT DEFAULT 'new',       -- 'new'|'interested'|'bidding'|'skipped'|'won'|'lost'
  review_note TEXT,
  reviewed_by INTEGER REFERENCES users(id),
  reviewed_at TEXT
);
CREATE INDEX IF NOT EXISTS tender_status ON tenders(status, deadline);
CREATE INDEX IF NOT EXISTS tender_review ON tenders(review_status);

CREATE TABLE IF NOT EXISTS my_bids (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  tender_id INTEGER REFERENCES tenders(id),
  company_id INTEGER REFERENCES companies(id),
  bid_amount INTEGER,
  bid_at TEXT DEFAULT (datetime('now')),
  result TEXT DEFAULT 'pending',          -- 'pending'|'won'|'lost'|'cancelled'
  result_at TEXT,
  note TEXT,
  submitted_by INTEGER REFERENCES users(id)
);

CREATE TABLE IF NOT EXISTS competitors (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL,
  business_no TEXT UNIQUE,
  note TEXT,
  watched INTEGER DEFAULT 1,
  added_at TEXT DEFAULT (datetime('now')),
  added_by INTEGER REFERENCES users(id)
);

CREATE TABLE IF NOT EXISTS competitor_bids (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  competitor_id INTEGER REFERENCES competitors(id),
  tender_id INTEGER REFERENCES tenders(id),
  bid_amount INTEGER,
  result TEXT,                            -- 'won'|'lost'|'unknown'
  detected_at TEXT DEFAULT (datetime('now')),
  UNIQUE(competitor_id, tender_id)
);
CREATE INDEX IF NOT EXISTS comp_bid_detect ON competitor_bids(detected_at);

-- 차량 (Vehicle) 마스터 + 배정
CREATE TABLE IF NOT EXISTS vehicles (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL,                     -- 예: 5톤 트럭 1호
  plate_no TEXT,                          -- 차량번호 12가1234
  vehicle_type TEXT,                      -- 덤프트럭|포클레인|지게차|승합|기타
  capacity TEXT,                          -- 5톤, 25톤, 0.7㎥ 등
  company_id INTEGER REFERENCES companies(id),
  status TEXT DEFAULT 'available',        -- 'available'|'in_use'|'maintenance'|'retired'
  purchased_at TEXT,
  note TEXT,
  created_at TEXT DEFAULT (datetime('now'))
);
CREATE TABLE IF NOT EXISTS vehicle_assignments (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  vehicle_id INTEGER NOT NULL REFERENCES vehicles(id) ON DELETE CASCADE,
  driver_id  INTEGER REFERENCES workers(id),
  site_id    INTEGER REFERENCES sites(id),
  assigned_at TEXT DEFAULT (datetime('now')),
  returned_at TEXT,                       -- NULL = 현재 진행 중
  note TEXT
);
CREATE INDEX IF NOT EXISTS va_vehicle_active ON vehicle_assignments(vehicle_id, returned_at);
CREATE INDEX IF NOT EXISTS va_driver_active  ON vehicle_assignments(driver_id, returned_at);
CREATE INDEX IF NOT EXISTS va_site_active    ON vehicle_assignments(site_id, returned_at);

-- 면허 (정식) — 회사별 면허 마스터
CREATE TABLE IF NOT EXISTS licenses (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
  license_type TEXT NOT NULL,             -- 토목공사업, 건축공사업, 기계설비공사업 등
  license_no TEXT,                        -- 등록번호
  issued_at TEXT,
  expires_at TEXT,                        -- 갱신 만료일
  capacity_amount INTEGER DEFAULT 0,      -- 시평액 (원)
  status TEXT DEFAULT 'active',           -- 'active'|'expired'|'suspended'
  note TEXT,
  created_at TEXT DEFAULT (datetime('now'))
);
CREATE INDEX IF NOT EXISTS lic_company ON licenses(company_id);
CREATE INDEX IF NOT EXISTS lic_expiry  ON licenses(expires_at);

-- 면허 ↔ 직원 등재 (한 면허에 어느 정규직이 기술자로 등재됐는지)
CREATE TABLE IF NOT EXISTS license_workers (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  license_id INTEGER NOT NULL REFERENCES licenses(id) ON DELETE CASCADE,
  worker_id  INTEGER NOT NULL REFERENCES workers(id) ON DELETE CASCADE,
  role TEXT,                              -- 등재 역할 (현장대리인·기술인 등)
  registered_at TEXT DEFAULT (datetime('now')),
  note TEXT,
  UNIQUE(license_id, worker_id)
);
CREATE INDEX IF NOT EXISTS lw_license ON license_workers(license_id);
CREATE INDEX IF NOT EXISTS lw_worker  ON license_workers(worker_id);

-- 직원이 보유한 자격증 (worker가 가진 — '면허 등재'와 다름)
CREATE TABLE IF NOT EXISTS worker_certifications (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  worker_id INTEGER NOT NULL REFERENCES workers(id) ON DELETE CASCADE,
  cert_name TEXT NOT NULL,                -- '토목기사', '석면해체감리원', '건축초급'
  cert_no TEXT,                           -- 등록번호
  cert_level TEXT,                        -- '초급'|'중급'|'고급'|'특급'|'기능사'|'기사'|'기술사'
  acquired_at TEXT,
  expires_at TEXT,                        -- 일부만 갱신 필요
  related_business TEXT,                  -- '토목', '건축', '안전' 등 관련 업종
  note TEXT
);
CREATE INDEX IF NOT EXISTS wc_worker ON worker_certifications(worker_id);

-- 주주·임원 명부
CREATE TABLE IF NOT EXISTS shareholders (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
  name TEXT NOT NULL,
  role TEXT,                              -- '대표','이사','감사','주주'
  rrn TEXT,                               -- 주민번호
  address TEXT,
  shares_pct REAL,                        -- 지분율
  contribution INTEGER,                   -- 출자금
  registered_at TEXT,                     -- 등기일
  worker_id INTEGER REFERENCES workers(id),  -- 직원과 동일인이면 연결
  note TEXT
);
CREATE INDEX IF NOT EXISTS sh_company ON shareholders(company_id);

-- 급여 기록 (월별 정규화)
CREATE TABLE IF NOT EXISTS payroll_records (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  worker_id INTEGER NOT NULL REFERENCES workers(id) ON DELETE CASCADE,
  company_id INTEGER REFERENCES companies(id),
  year_month TEXT NOT NULL,               -- '2025-10'
  base_salary INTEGER DEFAULT 0,
  allowances TEXT,                        -- JSON {meal, vehicle, position, tenure, site, skill, ...}
  deductions TEXT,                        -- JSON {income_tax, pension, health, emp_ins, ...}
  gross_pay INTEGER DEFAULT 0,
  total_deductions INTEGER DEFAULT 0,
  net_pay INTEGER DEFAULT 0,
  pay_date TEXT,
  note TEXT,
  created_at TEXT DEFAULT (datetime('now')),
  UNIQUE(worker_id, year_month)
);
CREATE INDEX IF NOT EXISTS pr_worker_month ON payroll_records(worker_id, year_month);

-- 직종별 일당 단가표
CREATE TABLE IF NOT EXISTS wage_rates (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  job_role TEXT NOT NULL,                 -- '보통인부', '직영인부', '철근공' 등
  wage_per_day INTEGER NOT NULL,
  effective_from TEXT,
  effective_to TEXT,
  note TEXT
);

CREATE TABLE IF NOT EXISTS process_instances (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  workflow TEXT NOT NULL,                  -- 'sales'|'hr_onboarding'|'daily_ops'|...
  subject_type TEXT NOT NULL,              -- 'Place'|'Person'|'Document'
  subject_id INTEGER NOT NULL,
  scope_key TEXT,                          -- 일일운영 같은 일자별 인스턴스용 (예: '2026-04-25')
  current_state TEXT NOT NULL,
  meta TEXT,                               -- JSON
  started_at TEXT DEFAULT (datetime('now')),
  updated_at TEXT,
  completed_at TEXT,
  UNIQUE(workflow, subject_type, subject_id, scope_key)
);
CREATE INDEX IF NOT EXISTS proc_workflow ON process_instances(workflow, current_state);

CREATE TABLE IF NOT EXISTS notifications (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  unique_key TEXT UNIQUE,                  -- 중복 방지
  rule_type TEXT,                          -- 'pending_approval'|'no_gps'|'expiring'|...
  severity TEXT DEFAULT 'info',            -- 'info'|'warning'|'urgent'
  title TEXT NOT NULL,
  message TEXT,
  link TEXT,
  related_entity_type TEXT,
  related_entity_id INTEGER,
  is_read INTEGER DEFAULT 0,
  resolved INTEGER DEFAULT 0,
  created_at TEXT DEFAULT (datetime('now')),
  read_at TEXT,
  resolved_at TEXT
);
CREATE INDEX IF NOT EXISTS notif_status ON notifications(resolved, severity, created_at);

CREATE TABLE IF NOT EXISTS relations (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  subject_type TEXT NOT NULL,             -- 'Person'|'Place'|'Org'|'Resource'|'Document'
  subject_id   INTEGER NOT NULL,
  predicate    TEXT NOT NULL,             -- 'employed_by'|'owns'|'manages'|'has_role_in'|...
  object_type  TEXT NOT NULL,
  object_id    INTEGER NOT NULL,
  metadata     TEXT,                       -- JSON, optional
  valid_from   TEXT,                       -- 시간적 관계 (옵션)
  valid_to     TEXT,
  created_at   TEXT DEFAULT (datetime('now')),
  UNIQUE(subject_type, subject_id, predicate, object_type, object_id)
);
CREATE INDEX IF NOT EXISTS rel_subject ON relations(subject_type, subject_id);
CREATE INDEX IF NOT EXISTS rel_object  ON relations(object_type, object_id);
CREATE INDEX IF NOT EXISTS rel_pred    ON relations(predicate);

CREATE TABLE IF NOT EXISTS events (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  type TEXT NOT NULL,                     -- 'ClockIn'|'Deploy'|'WorkerCreated'|...
  occurred_at TEXT NOT NULL,              -- ISO timestamp (KST naive)
  actors TEXT,                            -- JSON {worker_id, user_id, ...}
  place TEXT,                             -- JSON {site_id, lat, lng, ...}
  payload TEXT,                           -- JSON 이벤트별 상세
  financial TEXT,                         -- JSON {amount, account, kind: expense|revenue}
  created_by INTEGER REFERENCES users(id),-- 어떤 관리자가 트리거했는지 (있으면)
  source TEXT DEFAULT 'api'               -- 'admin_ui'|'mobile'|'public'|'system'
);
CREATE INDEX IF NOT EXISTS events_type_time ON events(type, occurred_at);
CREATE INDEX IF NOT EXISTS events_time ON events(occurred_at);

CREATE TABLE IF NOT EXISTS users (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  username TEXT UNIQUE NOT NULL,
  password_hash TEXT NOT NULL,
  name TEXT,
  role TEXT DEFAULT 'admin',          -- 'admin' | 'manager'
  company_id INTEGER REFERENCES companies(id),
  created_at TEXT DEFAULT (datetime('now'))
);
-- ====== 급여·노무 모듈 (Phase 7) ======
-- 외주 협력사 (형틀반·철근반·방수반 등 — 노무비 직접 신고 X, 사업자 거래)
CREATE TABLE IF NOT EXISTS subcontractors (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL,
  business_no TEXT,
  work_type TEXT,                          -- 형틀/철근/방수/조경/...
  bank_name TEXT,
  account_holder TEXT,
  account_no TEXT,
  leader_name TEXT,                        -- 외주반장 이름
  phone TEXT,
  address TEXT,
  note TEXT,
  created_at TEXT DEFAULT (datetime('now'))
);

-- 특고직 (특수형태근로종사자) — 장비기사. 사업자등록번호 가진 사람들
CREATE TABLE IF NOT EXISTS equipment_operators (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL,
  rrn TEXT,                                -- 주민번호
  business_no TEXT,                        -- 사업자등록번호
  equipment_type TEXT,                     -- B/H, D/T15TON, 5T살수, 추레라 등
  vendor_name TEXT,                        -- 소속 회사 (글로벌운수·통일중기 등)
  daily_rate INTEGER DEFAULT 0,            -- 일 단가 (원)
  phone TEXT,
  bank_name TEXT,
  account_holder TEXT,
  account_no TEXT,
  note TEXT,
  created_at TEXT DEFAULT (datetime('now'))
);

-- 직종별 표준 일당 (worker.daily_wage 의 default 또는 신규 입력 시 자동 채움)
CREATE TABLE IF NOT EXISTS wage_rates (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  job_role TEXT NOT NULL,                  -- 목공/철근공/보통인부/작업반장/특별인부 등
  daily_wage INTEGER NOT NULL,
  effective_from TEXT,                     -- 적용 시작일 (NULL = 항상)
  effective_to TEXT,
  note TEXT,
  UNIQUE(job_role, effective_from)
);

-- 공제율 마스터 (수정 가능 — 매년 변경됨)
CREATE TABLE IF NOT EXISTS tax_rates (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  year INTEGER NOT NULL,                   -- 적용 연도
  rate_type TEXT NOT NULL,                 -- pension/health/ltc/employment/income_tax/local_tax/income_tax_exempt/retirement_fund_per_day
  rate REAL NOT NULL,                      -- 비율 (%) 또는 금액 (원)
  is_amount INTEGER DEFAULT 0,             -- 1=금액(원), 0=비율
  note TEXT,
  UNIQUE(year, rate_type)
);

-- 월 단위 급여 마감
CREATE TABLE IF NOT EXISTS payroll_periods (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  year_month TEXT NOT NULL,                -- '2025-10'
  company_id INTEGER REFERENCES companies(id),
  status TEXT DEFAULT 'draft',             -- draft|closed|reported
  closed_at TEXT,
  closed_by INTEGER REFERENCES users(id),
  note TEXT,
  UNIQUE(year_month, company_id)
);

-- 명세서 한 줄 (사람 × 월 × 회사). 직영 일용직만. 외주/특고직은 별도 처리.
CREATE TABLE IF NOT EXISTS payroll_lines (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  period_id INTEGER NOT NULL REFERENCES payroll_periods(id) ON DELETE CASCADE,
  worker_id INTEGER NOT NULL REFERENCES workers(id) ON DELETE CASCADE,
  site_id INTEGER REFERENCES sites(id),    -- 주 현장 (선택)
  days_worked INTEGER DEFAULT 0,
  daily_wage INTEGER DEFAULT 0,
  gross_pay INTEGER DEFAULT 0,
  -- 공제 항목 (각각 원 단위)
  national_pension INTEGER DEFAULT 0,
  health_insurance INTEGER DEFAULT 0,
  ltc_insurance INTEGER DEFAULT 0,
  employment_insurance INTEGER DEFAULT 0,
  income_tax INTEGER DEFAULT 0,
  local_tax INTEGER DEFAULT 0,
  retirement_fund INTEGER DEFAULT 0,        -- 사업주 부담 (참고용)
  -- 합계
  total_deductions INTEGER DEFAULT 0,
  net_pay INTEGER DEFAULT 0,
  -- 신고 분류
  is_subject_4ins INTEGER DEFAULT 0,        -- 8일 이상 또는 220만원 이상
  reported_at TEXT,                         -- 근로내용확인신고 완료 시각
  -- 생성 시점 스냅샷
  bank_name TEXT,
  account_holder TEXT,
  account_no TEXT,
  note TEXT,
  calculated_at TEXT DEFAULT (datetime('now')),
  UNIQUE(period_id, worker_id)
);
CREATE INDEX IF NOT EXISTS pl_period ON payroll_lines(period_id);
CREATE INDEX IF NOT EXISTS pl_worker ON payroll_lines(worker_id);

-- 외주 노무비 입금 내역 (협력사 단위 — 노무비 신고 X)
CREATE TABLE IF NOT EXISTS subcontractor_payments (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  period_id INTEGER REFERENCES payroll_periods(id) ON DELETE CASCADE,
  subcontractor_id INTEGER NOT NULL REFERENCES subcontractors(id),
  site_id INTEGER REFERENCES sites(id),
  work_period TEXT,                        -- "2025-10-01 ~ 10-31"
  amount INTEGER NOT NULL,
  paid_at TEXT,
  invoice_no TEXT,
  note TEXT,
  created_at TEXT DEFAULT (datetime('now'))
);

-- 특고직 노무비 입금 내역
CREATE TABLE IF NOT EXISTS equipment_payments (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  period_id INTEGER REFERENCES payroll_periods(id) ON DELETE CASCADE,
  operator_id INTEGER NOT NULL REFERENCES equipment_operators(id),
  site_id INTEGER REFERENCES sites(id),
  days_worked INTEGER DEFAULT 0,
  daily_rate INTEGER DEFAULT 0,
  amount INTEGER NOT NULL,
  paid_at TEXT,
  note TEXT,
  created_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS clock_records (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  worker_id INTEGER NOT NULL REFERENCES workers(id) ON DELETE CASCADE,
  site_id INTEGER NOT NULL REFERENCES sites(id) ON DELETE CASCADE,
  date TEXT NOT NULL,
  clock_in TEXT,
  clock_out TEXT,
  in_lat REAL,
  in_lng REAL,
  in_distance_m REAL,
  in_verified INTEGER DEFAULT 0,
  out_lat REAL,
  out_lng REAL,
  out_distance_m REAL,
  out_verified INTEGER DEFAULT 0,
  UNIQUE(worker_id, date)
);
"""

@contextmanager
def conn():
    # autocommit (isolation_level=None) — 각 statement 즉시 commit, lock 해제
    # → 중첩된 with conn() 호출에서 'database is locked' 방지
    # timeout=10s + busy_timeout 으로 만에 하나 동시 write 시 대기
    c = sqlite3.connect(DB_PATH, timeout=10.0, isolation_level=None)
    c.row_factory = sqlite3.Row
    c.execute("PRAGMA foreign_keys = ON")
    c.execute("PRAGMA busy_timeout = 10000")
    try:
        yield c
        # autocommit 모드라 명시 commit 불필요
    finally:
        c.close()

SCHEMA_MIGRATIONS = [
    # 기존 테이블에 컬럼 추가 (있으면 IGNORE)
    "ALTER TABLE sites ADD COLUMN site_category TEXT",
    "ALTER TABLE sites ADD COLUMN required_license TEXT",
    # workers 확장 — 엑셀 임포트용
    "ALTER TABLE workers ADD COLUMN rrn TEXT",
    "ALTER TABLE workers ADD COLUMN address TEXT",
    "ALTER TABLE workers ADD COLUMN bank_name TEXT",
    "ALTER TABLE workers ADD COLUMN account_holder TEXT",
    "ALTER TABLE workers ADD COLUMN job_specialty TEXT",
    "ALTER TABLE workers ADD COLUMN asbestos_certified INTEGER DEFAULT 0",
    "ALTER TABLE workers ADD COLUMN resigned_at TEXT",
    "ALTER TABLE workers ADD COLUMN position TEXT",
    # companies 확장
    "ALTER TABLE companies ADD COLUMN incorporation_date TEXT",
    "ALTER TABLE companies ADD COLUMN registration_date TEXT",
    "ALTER TABLE companies ADD COLUMN fiscal_year_end TEXT",
    "ALTER TABLE companies ADD COLUMN address TEXT",
    "ALTER TABLE companies ADD COLUMN corporate_no TEXT",  # 법인등록번호 (110111-XXXXXXX)
    "ALTER TABLE companies ADD COLUMN phone TEXT",
    "ALTER TABLE companies ADD COLUMN email TEXT",
    "ALTER TABLE companies ADD COLUMN representative_phone TEXT",
    # workers 급여 모듈 추가 컬럼
    "ALTER TABLE workers ADD COLUMN birth_date TEXT",                  # 생년월일 (만나이/고용보험 제외 판정용)
    "ALTER TABLE workers ADD COLUMN exempt_employment_ins INTEGER DEFAULT 0",  # 1952년 이전 출생자 = 고용보험 제외
    "ALTER TABLE workers ADD COLUMN subcontractor_id INTEGER",         # 외주 협력사 소속이면 (보통 NULL — 직영)
]

def _migrate_deployments_nullable_site():
    """deployments.site_id 의 NOT NULL 제거 (이미 nullable 이면 noop)."""
    try:
        with conn() as c:
            cols = c.execute("PRAGMA table_info(deployments)").fetchall()
            site_col = next((col for col in cols if col[1] == 'site_id'), None)
            if site_col is None: return  # 테이블 자체 없음
            # PRAGMA table_info: (cid, name, type, notnull, dflt_value, pk)
            if site_col[3] == 0: return   # 이미 nullable
            print("[migrate] deployments.site_id 를 nullable 로 변경 중...")
            c.executescript("""
                CREATE TABLE deployments_new (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  worker_id INTEGER NOT NULL REFERENCES workers(id) ON DELETE CASCADE,
                  site_id INTEGER REFERENCES sites(id) ON DELETE SET NULL,
                  date TEXT NOT NULL,
                  kind TEXT NOT NULL DEFAULT 'plan',
                  note TEXT,
                  UNIQUE(worker_id, date, kind)
                );
                INSERT INTO deployments_new(id, worker_id, site_id, date, kind, note)
                  SELECT id, worker_id, site_id, date, kind, note FROM deployments;
                DROP TABLE deployments;
                ALTER TABLE deployments_new RENAME TO deployments;
            """)
            print("[migrate] deployments.site_id nullable 변환 완료")
    except Exception as e:
        print(f"[migrate deployments] {e}")

def init_db():
    with conn() as c:
        # WAL 모드 — reader/writer 동시성 향상 (database is locked 방지)
        try: c.execute("PRAGMA journal_mode=WAL")
        except sqlite3.OperationalError: pass
        c.executescript(SCHEMA)
        for stmt in SCHEMA_MIGRATIONS:
            try: c.execute(stmt)
            except sqlite3.OperationalError: pass  # column exists
    # 별도 마이그레이션 (CREATE/INSERT/DROP/ALTER 한 번에)
    _migrate_deployments_nullable_site()

# ========================================================================
# Helpers
# ========================================================================
def haversine_m(lat1, lon1, lat2, lon2):
    """두 GPS 좌표 사이의 거리 (m)."""
    if None in (lat1, lon1, lat2, lon2):
        return None
    R = 6371000.0
    p1 = math.radians(lat1); p2 = math.radians(lat2)
    dp = math.radians(lat2 - lat1); dl = math.radians(lon2 - lon1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2*R*math.asin(math.sqrt(a))

def row_to_dict(row):
    return dict(row) if row else None

def rows(rs):
    return [dict(r) for r in rs]

# ---- 프로세스 정의 (Phase 4) ----
PROCESS_DEFS = {
    'sales': {
        'name': '수주 프로세스', 'subject': 'Place',
        'states': ['발주정보', '견적준비', '입찰참여', '낙찰', '계약체결', '착공', '진행', '준공'],
        'terminal': ['준공'],
    },
    'hr_onboarding': {
        'name': '인력 온보딩', 'subject': 'Person',
        'states': ['가입신청', '신원확인', '계좌등록', '안전교육', '4대보험', '활성'],
        'terminal': ['활성'],
    },
    'daily_ops': {
        'name': '일일 운영', 'subject': 'Place',
        'states': ['배치계획', 'TBM', '출역체크', '작업중', '일보작성', '실적확정'],
        'terminal': ['실적확정'],
    },
    'progress_billing': {
        'name': '기성 청구', 'subject': 'Place',
        'states': ['실적누적', '기성산정', '청구서작성', '제출', '검수', '승인', '수금'],
        'terminal': ['수금'],
    },
    'safety': {
        'name': '안전 관리', 'subject': 'Place',
        'states': ['위험성평가', 'TBM', '작업중', '무사고종료', '사고발생'],
        'terminal': ['무사고종료'],
    },
    'compliance': {
        'name': '신고 컴플라이언스', 'subject': 'Place',
        'states': ['실적확정', '월말집계', '신고서생성', '검토', '제출', '접수확인'],
        'terminal': ['접수확인'],
    },
    'close_out': {
        'name': '정산 준공', 'subject': 'Place',
        'states': ['준공검사', '최종기성', '잔금청구', '하자보증', '실적등재', '결산반영'],
        'terminal': ['결산반영'],
    },
}

def create_or_advance_process(workflow, subject_type, subject_id, target_state,
                              scope_key=None, meta=None):
    """프로세스 인스턴스를 만들거나 다음 상태로 진행. 뒤로 안 가고 같거나 앞 상태일 때만 갱신."""
    defn = PROCESS_DEFS.get(workflow)
    if not defn or target_state not in defn['states']:
        return
    try:
        with conn() as c:
            row = c.execute(
                "SELECT id, current_state FROM process_instances "
                "WHERE workflow=? AND subject_type=? AND subject_id=? AND IFNULL(scope_key,'')=IFNULL(?,'')",
                (workflow, subject_type, subject_id, scope_key)
            ).fetchone()
            if row:
                states = defn['states']
                try:
                    cur_idx = states.index(row['current_state'])
                    tgt_idx = states.index(target_state)
                    if tgt_idx > cur_idx:
                        completed = (target_state in defn.get('terminal', []))
                        c.execute(
                            "UPDATE process_instances SET current_state=?, meta=?, updated_at=datetime('now'), "
                            "completed_at=CASE WHEN ?=1 THEN datetime('now') ELSE completed_at END WHERE id=?",
                            (target_state, json.dumps(meta or {}, ensure_ascii=False),
                             1 if completed else 0, row['id'])
                        )
                        emit_event("ProcessAdvanced", payload={
                            "workflow": workflow, "subject_type": subject_type,
                            "subject_id": subject_id, "scope_key": scope_key,
                            "from": row['current_state'], "to": target_state
                        }, source='system')
                except ValueError:
                    pass
            else:
                c.execute(
                    "INSERT INTO process_instances(workflow,subject_type,subject_id,scope_key,current_state,meta) "
                    "VALUES(?,?,?,?,?,?)",
                    (workflow, subject_type, subject_id, scope_key, target_state,
                     json.dumps(meta or {}, ensure_ascii=False))
                )
                emit_event("ProcessStarted", payload={
                    "workflow": workflow, "subject_type": subject_type,
                    "subject_id": subject_id, "scope_key": scope_key, "state": target_state
                }, source='system')
    except Exception as e:
        print(f"[create_or_advance_process] {e}")

def _process_react_to_event(event_type, actors, place, payload, financial):
    """이벤트가 발생하면 자동으로 관련 프로세스 진행/생성."""
    today_str = date.today().isoformat()
    actors = actors or {}; place = place or {}; payload = payload or {}
    try:
        if event_type == 'WorkerSelfRegistered':
            wid = actors.get('worker_id')
            if wid: create_or_advance_process('hr_onboarding', 'Person', wid, '가입신청')
        elif event_type == 'WorkerCreated':
            wid = actors.get('worker_id')
            if wid: create_or_advance_process('hr_onboarding', 'Person', wid, '활성')
        elif event_type == 'WorkerUpdated':
            wid = actors.get('worker_id')
            if wid and payload.get('daily_wage') and payload.get('daily_wage') > 0:
                create_or_advance_process('hr_onboarding', 'Person', wid, '계좌등록')
        elif event_type == 'SiteCreated':
            sid = place.get('site_id')
            if sid:
                create_or_advance_process('sales', 'Place', sid, '계약체결')
                create_or_advance_process('safety', 'Place', sid, '위험성평가')
        elif event_type == 'Deploy':
            sid = place.get('site_id'); d = payload.get('date') or today_str
            if sid: create_or_advance_process('daily_ops', 'Place', sid, '배치계획', scope_key=d)
        elif event_type == 'ClockIn':
            sid = place.get('site_id'); d = payload.get('date') or today_str
            if sid:
                create_or_advance_process('daily_ops', 'Place', sid, '출역체크', scope_key=d)
                create_or_advance_process('safety', 'Place', sid, '작업중')
                create_or_advance_process('sales', 'Place', sid, '진행')  # 첫 출근 = 진행 단계로
        elif event_type == 'SiteUpdated':
            sid = place.get('site_id')
            if sid and payload.get('status') == 'closed':
                create_or_advance_process('sales', 'Place', sid, '준공')
                create_or_advance_process('close_out', 'Place', sid, '준공검사')
    except Exception as e:
        print(f"[process_react] {e}")

# ---- 온톨로지 / 관계 그래프 (Phase 2) ----
# 6대 엔티티: Person, Place, Organization, Resource, Document, Money
# 기존 테이블 매핑: workers↔Person, sites↔Place, companies↔Organization, users↔Person(role)
# 관계는 명시적으로 relations 테이블에 저장 — FK는 그대로 두고 위에 layered

ENTITY_TABLES = {
    'Person':  ('workers',   'name'),
    'Place':   ('sites',     'name'),
    'Org':     ('companies', 'name'),
    'User':    ('users',     'name'),   # admin/manager 계정도 Person 그래프에
}

def add_relation(subject_type, subject_id, predicate, object_type, object_id,
                 metadata=None, valid_from=None, valid_to=None):
    if not (subject_type and subject_id and predicate and object_type and object_id):
        return
    try:
        with conn() as c:
            c.execute(
                """INSERT OR IGNORE INTO relations
                   (subject_type,subject_id,predicate,object_type,object_id,metadata,valid_from,valid_to)
                   VALUES(?,?,?,?,?,?,?,?)""",
                (subject_type, subject_id, predicate, object_type, object_id,
                 json.dumps(metadata or {}, ensure_ascii=False) if metadata else None,
                 valid_from, valid_to))
    except Exception as e:
        print(f"[add_relation] failed: {e}")

def remove_relations(subject_type=None, subject_id=None, predicate=None,
                     object_type=None, object_id=None):
    sql = "DELETE FROM relations WHERE 1=1"
    args = []
    if subject_type: sql += " AND subject_type=?"; args.append(subject_type)
    if subject_id:   sql += " AND subject_id=?";   args.append(subject_id)
    if predicate:    sql += " AND predicate=?";    args.append(predicate)
    if object_type:  sql += " AND object_type=?";  args.append(object_type)
    if object_id:    sql += " AND object_id=?";    args.append(object_id)
    try:
        with conn() as c:
            c.execute(sql, args)
    except Exception as e:
        print(f"[remove_relations] failed: {e}")

def _backfill_relations():
    """기존 FK 관계를 relations 테이블에 1회 채워넣음 (이미 있는 행은 skip)."""
    try:
        with conn() as c:
            n = c.execute("SELECT COUNT(*) FROM relations").fetchone()[0]
            if n > 0:
                return
            # workers.company_id → Person employed_by Org
            for r in c.execute("SELECT id, company_id FROM workers WHERE company_id IS NOT NULL").fetchall():
                add_relation('Person', r['id'], 'employed_by', 'Org', r['company_id'])
            # sites.company_id → Place owned_by Org
            for r in c.execute("SELECT id, company_id FROM sites WHERE company_id IS NOT NULL").fetchall():
                add_relation('Place', r['id'], 'owned_by', 'Org', r['company_id'])
            # users.company_id → User has_role_in Org
            for r in c.execute("SELECT id, company_id, role FROM users WHERE company_id IS NOT NULL").fetchall():
                add_relation('User', r['id'], 'has_role_in', 'Org', r['company_id'],
                             metadata={'role': r['role']})
            print("[startup] relations backfill 완료")
    except Exception as e:
        print(f"[backfill_relations] {e}")

# ---- 이벤트 코어 (Phase 1: 디지털 트윈의 시작) ----
def emit_event(event_type, actors=None, place=None, payload=None, financial=None,
               created_by=None, source='api'):
    """모든 도메인 액션은 이걸 호출해서 events 테이블에 기록.
    실패해도 주 동작에 영향 없도록 best-effort. Append-only history.
    추가: 이벤트가 자동으로 관련 프로세스를 진행시킴 (Phase 4)."""
    try:
        with conn() as c:
            c.execute(
                """INSERT INTO events(type,occurred_at,actors,place,payload,financial,created_by,source)
                   VALUES(?,?,?,?,?,?,?,?)""",
                (event_type,
                 datetime.now().isoformat(timespec='seconds'),
                 json.dumps(actors or {}, ensure_ascii=False),
                 json.dumps(place or {}, ensure_ascii=False),
                 json.dumps(payload or {}, ensure_ascii=False),
                 json.dumps(financial or {}, ensure_ascii=False),
                 created_by, source)
            )
    except Exception as e:
        print(f"[emit_event] {event_type} failed: {e}")
    # 프로세스 자동 진행 (Process* 이벤트는 무한루프 방지를 위해 제외)
    if not event_type.startswith('Process'):
        try: _process_react_to_event(event_type, actors, place, payload, financial)
        except Exception as e: print(f"[process_react] {e}")

# ========================================================================
# Pydantic models (입력)
# ========================================================================
class CompanyIn(BaseModel):
    name: str
    business_no: Optional[str] = None      # 사업자등록번호 XXX-XX-XXXXX
    corporate_no: Optional[str] = None     # 법인등록번호 XXXXXX-XXXXXXX
    ceo: Optional[str] = None
    license_info: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    representative_phone: Optional[str] = None
    fiscal_year_end: Optional[str] = None
    incorporation_date: Optional[str] = None
    registration_date: Optional[str] = None

class SubcontractorIn(BaseModel):
    name: str
    business_no: Optional[str] = None
    work_type: Optional[str] = None
    bank_name: Optional[str] = None
    account_holder: Optional[str] = None
    account_no: Optional[str] = None
    leader_name: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    note: Optional[str] = None

class EquipmentOperatorIn(BaseModel):
    name: str
    rrn: Optional[str] = None
    business_no: Optional[str] = None
    equipment_type: Optional[str] = None
    vendor_name: Optional[str] = None
    daily_rate: Optional[int] = 0
    phone: Optional[str] = None
    bank_name: Optional[str] = None
    account_holder: Optional[str] = None
    account_no: Optional[str] = None
    note: Optional[str] = None

class WageRateIn(BaseModel):
    job_role: str
    daily_wage: int
    effective_from: Optional[str] = None
    effective_to: Optional[str] = None
    note: Optional[str] = None

class TaxRateIn(BaseModel):
    year: int
    rate_type: str
    rate: float
    is_amount: Optional[int] = 0
    note: Optional[str] = None

class CertificationIn(BaseModel):
    cert_name: str
    cert_level: Optional[str] = None
    cert_no: Optional[str] = None
    acquired_at: Optional[str] = None
    expires_at: Optional[str] = None
    related_business: Optional[str] = None
    note: Optional[str] = None

class ShareholderIn(BaseModel):
    name: str
    role: Optional[str] = None
    rrn: Optional[str] = None
    address: Optional[str] = None
    shares_pct: Optional[float] = None
    contribution: Optional[int] = None
    registered_at: Optional[str] = None
    worker_id: Optional[int] = None
    note: Optional[str] = None

class SiteIn(BaseModel):
    company_id: Optional[int] = None
    name: str
    address: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    geofence_meters: Optional[int] = 200
    contract_amount: Optional[int] = 0
    paid_amount: Optional[int] = 0
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    status: Optional[str] = "active"
    manager: Optional[str] = None

class WorkerIn(BaseModel):
    company_id: Optional[int] = None
    name: str
    phone: Optional[str] = None
    worker_type: Optional[str] = "daily"
    daily_wage: Optional[int] = 0
    job_role: Optional[str] = None
    hired_date: Optional[str] = None
    rrn_last: Optional[str] = None
    bank_account: Optional[str] = None
    note: Optional[str] = None
    # 확장 필드 (직원,주주명부 / 급여대장 등에서)
    rrn: Optional[str] = None
    address: Optional[str] = None
    position: Optional[str] = None
    resigned_at: Optional[str] = None
    bank_name: Optional[str] = None
    account_holder: Optional[str] = None
    asbestos_certified: Optional[int] = 0
    job_specialty: Optional[str] = None

class RegisterIn(BaseModel):
    name: str
    phone: str
    worker_type: Optional[str] = "daily"
    job_role: Optional[str] = None

class LoginIn(BaseModel):
    username: str
    password: str

class WorkerIdentifyIn(BaseModel):
    phone: str

class SignupIn(BaseModel):
    username: str
    password: str
    name: str
    invite_code: Optional[str] = None

class PasswordChangeIn(BaseModel):
    current_password: str
    new_password: str

class DeploymentIn(BaseModel):
    worker_id: int
    site_id: int
    date: str
    kind: str = "plan"   # plan | actual | reported
    note: Optional[str] = None

class ClockIn(BaseModel):
    worker_id: int
    site_id: int
    lat: Optional[float] = None
    lng: Optional[float] = None
    direction: str = "in"   # 'in' or 'out'

# ========================================================================
# App
# ========================================================================
app = FastAPI(title="GW Construction Management API")
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY, session_cookie="gwc_sess",
                   max_age=60*60*24*14, https_only=False, same_site="lax")

from fastapi.responses import JSONResponse
from fastapi.requests import Request as _StarReq

@app.exception_handler(Exception)
async def _global_error_handler(request: _StarReq, exc: Exception):
    """모든 예상치 못한 에러를 JSON 으로 반환 — 클라이언트가 'Internal Server Error' HTML 받고 JSON 파싱 실패하는 문제 방지."""
    import traceback
    if isinstance(exc, HTTPException):
        return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})
    print(f"[ERR] {request.method} {request.url.path}: {exc}")
    print(traceback.format_exc())
    return JSONResponse(
        status_code=500,
        content={"detail": str(exc), "trace": traceback.format_exc()[-1500:]}
    )
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_credentials=False,
    allow_methods=["*"], allow_headers=["*"],
)

# ========================================================================
# Auth helpers
# ========================================================================
def hash_pw(p: str) -> str:
    if bcrypt:
        return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()
    return "plain:" + p   # fallback (개발용)

def verify_pw(p: str, h: str) -> bool:
    if not h: return False
    if h.startswith("plain:"):
        return h == "plain:" + p
    if bcrypt:
        try: return bcrypt.checkpw(p.encode(), h.encode())
        except Exception: return False
    return False

def require_login(request: Request):
    uid = request.session.get("user_id")
    if not uid:
        raise HTTPException(401, "로그인이 필요합니다")
    with conn() as c:
        u = c.execute("SELECT id, username, name, role, company_id FROM users WHERE id=?", (uid,)).fetchone()
    if not u:
        request.session.clear()
        raise HTTPException(401, "세션이 만료되었습니다")
    return dict(u)

def _bootstrap_admin():
    """관리자 계정이 하나도 없으면 기본 admin 계정 생성."""
    try:
        with conn() as c:
            n = c.execute("SELECT COUNT(*) FROM users").fetchone()[0]
            if n == 0:
                pw = os.environ.get("ADMIN_PASSWORD", "admin1234")
                c.execute("INSERT INTO users(username,password_hash,name,role) VALUES(?,?,?,?)",
                          ("admin", hash_pw(pw), "관리자", "admin"))
                used_env = bool(os.environ.get("ADMIN_PASSWORD"))
                print(f"[startup] 기본 관리자 계정 생성 — username=admin, "
                      + ("password=(환경변수 ADMIN_PASSWORD 사용)" if used_env
                         else "password=admin1234  ⚠️ 환경변수 ADMIN_PASSWORD 설정 강력 권장"))
    except Exception as e:
        print(f"[startup] bootstrap admin skipped: {e}")

def _auto_seed_if_empty():
    """빈 DB면 샘플 데이터 자동 입력 (Render 무료 티어 cold start 대비)."""
    try:
        with conn() as c:
            n = c.execute("SELECT COUNT(*) FROM companies").fetchone()[0]
        if n == 0:
            import sys as _sys, os as _os
            _sys.path.insert(0, _os.path.dirname(_os.path.abspath(__file__)))
            from seed import seed as _seed
            _seed()
            print("[startup] 샘플 데이터 자동 입력 완료")
    except Exception as e:
        print(f"[startup] auto-seed skipped: {e}")

def _seed_mock_vehicles_licenses():
    """차량·면허 시연용 데이터."""
    try:
        with conn() as c:
            # Vehicles (회사별 몇 대씩)
            if c.execute("SELECT COUNT(*) FROM vehicles").fetchone()[0] == 0:
                comps = [r[0] for r in c.execute("SELECT id FROM companies LIMIT 3").fetchall()]
                samples = [
                    ("5톤 트럭 1호", "12가1234", "덤프트럭", "5톤"),
                    ("11톤 카고", "34나5678", "카고트럭", "11톤"),
                    ("0.7㎥ 굴삭기", "강원01-1234", "포클레인", "0.7㎥"),
                    ("2.5톤 지게차 A", "광주02-5678", "지게차", "2.5톤"),
                    ("승합차 (사무)", "56다9012", "승합", "12인승"),
                    ("미니 굴삭기", "경기03-7777", "포클레인", "0.3㎥"),
                ]
                for i, (n, p, t, cap) in enumerate(samples):
                    cid = comps[i % len(comps)] if comps else None
                    c.execute("""INSERT INTO vehicles(name,plate_no,vehicle_type,capacity,company_id,status)
                                 VALUES(?,?,?,?,?,?)""",
                              (n, p, t, cap, cid, 'available'))
                    if cid:
                        add_relation('Resource', c.execute("SELECT last_insert_rowid()").fetchone()[0],
                                     'owned_by', 'Org', cid)
                print("[startup] vehicles seed 완료")

            # Licenses (각 회사에 면허 2~3개)
            if c.execute("SELECT COUNT(*) FROM licenses").fetchone()[0] == 0:
                comp_rows = c.execute("SELECT id, license_info FROM companies").fetchall()
                from datetime import timedelta
                today_d_ = date.today()
                license_samples = [
                    [("토목공사업", "서울-12345", today_d_ - timedelta(days=400), today_d_ + timedelta(days=200), 5_000_000_000),
                     ("건축공사업", "서울-12346", today_d_ - timedelta(days=400), today_d_ + timedelta(days=60),  3_000_000_000)],
                    [("철근콘크리트공사업", "서울-22345", today_d_ - timedelta(days=500), today_d_ + timedelta(days=15), 2_000_000_000),
                     ("토공사업", "서울-22346", today_d_ - timedelta(days=300), today_d_ + timedelta(days=350), 1_500_000_000)],
                    [("기계설비공사업", "서울-32345", today_d_ - timedelta(days=200), today_d_ + timedelta(days=400), 2_500_000_000),
                     ("가스시설공사업", "서울-32346", today_d_ - timedelta(days=600), today_d_ + timedelta(days=20),  800_000_000)],
                ]
                for i, comp in enumerate(comp_rows):
                    if i >= len(license_samples): break
                    for lt, ln, issued, expires, cap in license_samples[i]:
                        c.execute(
                            """INSERT INTO licenses(company_id,license_type,license_no,issued_at,expires_at,capacity_amount)
                               VALUES(?,?,?,?,?,?)""",
                            (comp[0], lt, ln, issued.isoformat(), expires.isoformat(), cap)
                        )
                print("[startup] licenses seed 완료")
    except Exception as e:
        print(f"[mock_vehicles_licenses] {e}")

def _seed_payroll_defaults():
    """공제율·표준일당 기본값 시드 — 빈 DB 일 때만."""
    try:
        with conn() as c:
            if c.execute("SELECT COUNT(*) FROM tax_rates").fetchone()[0] == 0:
                year = 2025
                rates = [
                    # rate_type, rate, is_amount, note
                    ("pension", 0.045, 0, "국민연금 4.5% (사업장가입자)"),
                    ("health", 0.03545, 0, "건강보험 3.545%"),
                    ("ltc", 0.06475, 0, "장기요양 6.475% (건강보험액 기준)"),
                    ("employment", 0.009, 0, "고용보험 0.9% (1952년 이전 출생자 제외)"),
                    ("income_tax", 0.06, 0, "갑근세 6%"),
                    ("income_tax_reduction", 0.55, 0, "근로소득세액공제 55% (감면)"),
                    ("local_tax", 0.10, 0, "주민세 = 갑근세 × 10%"),
                    ("income_tax_exempt", 150_000, 1, "갑근세 면세점 일 150,000원"),
                    ("retirement_fund_per_day", 6_500, 1, "건설근로자 퇴직공제 일 6,500원 (사업주 부담)"),
                    ("ins_threshold_days", 8, 1, "4대보험 적용 기준 일수 (월 8일 이상)"),
                    ("ins_threshold_amount", 2_200_000, 1, "4대보험 적용 기준 금액 (월 220만원 이상)"),
                ]
                for rt, val, is_amt, note in rates:
                    c.execute(
                        "INSERT OR IGNORE INTO tax_rates(year, rate_type, rate, is_amount, note) VALUES(?,?,?,?,?)",
                        (year, rt, val, is_amt, note))
            if c.execute("SELECT COUNT(*) FROM wage_rates").fetchone()[0] == 0:
                # 일용직명단 분석 결과 기반
                std = [
                    ("작업반장", 200_000),
                    ("목공",     180_000),
                    ("철근공",    180_000),
                    ("석공",     200_000),
                    ("형틀목수",   200_000),
                    ("미장공",    180_000),
                    ("타일공",    180_000),
                    ("방수공",    170_000),
                    ("도장공",    170_000),
                    ("전기공",    180_000),
                    ("배관공",    180_000),
                    ("울타리공",   170_000),
                    ("보통인부",   160_000),
                    ("단순노무자",  150_000),
                    ("직영인부",   150_000),
                    ("특별인부",   170_000),
                    ("굴삭기기사",  250_000),
                    ("지게차기사",  200_000),
                ]
                for jr, dw in std:
                    c.execute("INSERT OR IGNORE INTO wage_rates(job_role, daily_wage, effective_from) VALUES(?,?,?)",
                              (jr, dw, "2025-01-01"))
            print("[startup] payroll defaults seed 완료 (tax_rates + wage_rates)")
    except Exception as e:
        print(f"[seed_payroll_defaults] {e}")

# ===== 급여 계산 엔진 =====
def _get_tax_rates(year=None):
    """현재 연도의 공제율을 dict 로 반환."""
    if year is None: year = date.today().year
    out = {}
    with conn() as c:
        for r in c.execute("SELECT rate_type, rate, is_amount FROM tax_rates WHERE year=?", (year,)).fetchall():
            out[r['rate_type']] = (r['rate'], bool(r['is_amount']))
        # 없으면 이전 연도 fallback
        if not out:
            for r in c.execute("SELECT rate_type, rate, is_amount FROM tax_rates ORDER BY year DESC").fetchall():
                if r['rate_type'] not in out:
                    out[r['rate_type']] = (r['rate'], bool(r['is_amount']))
    return out

def calculate_payroll(daily_wage: int, days: int, exempt_employment_ins: bool = False, year=None):
    """일당·일수 → 갑근세·주민세·4대보험·실수령 자동 계산.
    리턴: dict — gross / income_tax / local_tax / employment / pension / health / ltc /
                 retirement_fund / total_deductions / net_pay / is_subject_4ins
    """
    if days <= 0 or daily_wage <= 0:
        return {
            "gross_pay": 0, "income_tax": 0, "local_tax": 0,
            "employment_insurance": 0, "national_pension": 0, "health_insurance": 0,
            "ltc_insurance": 0, "retirement_fund": 0,
            "total_deductions": 0, "net_pay": 0, "is_subject_4ins": False,
        }
    rates = _get_tax_rates(year)
    def r(key, default=0):
        v = rates.get(key)
        return v[0] if v else default

    gross = daily_wage * days

    # 갑근세 = (일당 - 면세점) × 6% × 0.55 × 일수
    exempt_amt = r('income_tax_exempt', 150_000)
    if daily_wage <= exempt_amt:
        income_tax = 0
    else:
        per_day_tax = (daily_wage - exempt_amt) * r('income_tax', 0.06) * (1 - r('income_tax_reduction', 0.55))
        # 위 공식: 6% × (1 - 55%) = 2.7%. 하지만 자료에 따라 0.55 가 곱해지는 식으로 표기되기도.
        # 우리 시트 기준: 0.06 × 0.55 = 0.033 이 아니라 0.027 (= 6% × 0.45). 그래서 (1 - 0.55) 로 처리.
        # 실제로 자비스/소득세계산기 사이트 = (일당-15만) × 6% × 0.55 = 직접곱 — 이게 맞음
        # 근로소득세액공제 = 산출세액의 55% 공제 → 잔여 45%만 부담 → (일당-15만) × 6% × 0.45 가 정확
        # 한국 일용직 갑근세 = 산출세액(6%) - 근로소득세액공제(55%) = 6% × (1-0.55) = 2.7% → 우리 시트와 일치
        # 따라서 (1 - reduction) 으로 계산
        income_tax = round(per_day_tax * days)
        # 999원 이하 자동 면세
        if income_tax <= 999:
            income_tax = 0
    local_tax = round(income_tax * r('local_tax', 0.10))

    # 4대보험 적용 여부 (8일 이상 OR 220만원 이상)
    ins_days = int(r('ins_threshold_days', 8))
    ins_amt = int(r('ins_threshold_amount', 2_200_000))
    is_subject = (days >= ins_days) or (gross >= ins_amt)

    # 고용보험 — 1952년 이전 출생자 제외
    employment = 0 if exempt_employment_ins else round(gross * r('employment', 0.009))

    # 국민연금·건강·장기요양 — 4대보험 대상자만
    if is_subject:
        pension = round(gross * r('pension', 0.045))
        health = round(gross * r('health', 0.03545))
        ltc = round(health * r('ltc', 0.06475))
    else:
        pension = health = ltc = 0

    # 퇴직공제 (사업주 부담, 직원 공제는 아님 — 표시용)
    retirement_fund = int(r('retirement_fund_per_day', 6_500)) * days

    total_ded = income_tax + local_tax + employment + pension + health + ltc
    net = gross - total_ded

    return {
        "gross_pay": gross,
        "income_tax": income_tax,
        "local_tax": local_tax,
        "employment_insurance": employment,
        "national_pension": pension,
        "health_insurance": health,
        "ltc_insurance": ltc,
        "retirement_fund": retirement_fund,
        "total_deductions": total_ded,
        "net_pay": net,
        "is_subject_4ins": is_subject,
    }

def _seed_mock_competitor_bids():
    """경쟁사 + 일부 입찰 기록 시연 데이터."""
    try:
        with conn() as c:
            if c.execute("SELECT COUNT(*) FROM competitors").fetchone()[0] > 0:
                return
            samples = [
                ("(주)대형건설", "111-22-33333", "원도급사. 대형 토목 위주."),
                ("진양건설(주)", "222-33-44444", "전문건설업, 콘크리트 강함."),
                ("(주)녹색이엔지", "333-44-55555", "기계설비 전문."),
            ]
            for name, biz, note in samples:
                c.execute("INSERT INTO competitors(name,business_no,note) VALUES(?,?,?)",
                          (name, biz, note))
            # 가능하면 mock tender 와 매핑
            tids = [r[0] for r in c.execute("SELECT id FROM tenders LIMIT 5").fetchall()]
            cids = [r[0] for r in c.execute("SELECT id FROM competitors").fetchall()]
            if tids and cids:
                for ti, t in enumerate(tids[:3]):
                    cid = cids[ti % len(cids)]
                    try:
                        c.execute(
                            "INSERT OR IGNORE INTO competitor_bids(competitor_id,tender_id,bid_amount,result) "
                            "VALUES(?,?,?,?)",
                            (cid, t, 800000000 + ti*100000000, 'unknown'))
                    except Exception: pass
    except Exception as e:
        print(f"[mock_competitors] {e}")

@app.on_event("startup")
def _startup():
    init_db()
    _bootstrap_admin()
    _auto_seed_if_empty()
    _backfill_relations()
    _backfill_processes()
    # 시연용 나라장터 mock — DB 비어있을 때만 (외부 conn 닫고 호출 — nested conn lock 방지)
    try:
        n_tenders = 0
        with conn() as c:
            n_tenders = c.execute("SELECT COUNT(*) FROM tenders").fetchone()[0]
        if n_tenders == 0:
            _seed_mock_tenders()
            _seed_mock_competitor_bids()
    except Exception as e: print(f"[startup mock tenders] {e}")
    _seed_mock_vehicles_licenses()
    _seed_payroll_defaults()
    _evaluate_rules()

# ----- Auth -----
@app.post("/api/login")
def login(payload: LoginIn, request: Request):
    with conn() as c:
        u = c.execute("SELECT * FROM users WHERE username=?", (payload.username,)).fetchone()
    if not u or not verify_pw(payload.password, u["password_hash"]):
        raise HTTPException(401, "아이디 또는 비밀번호가 일치하지 않습니다")
    request.session["user_id"] = u["id"]
    request.session["username"] = u["username"]
    request.session["role"] = u["role"]
    return {"ok": True, "username": u["username"], "name": u["name"], "role": u["role"]}

@app.post("/api/logout")
def logout(request: Request):
    request.session.clear()
    return {"ok": True}

@app.get("/api/me")
def me(request: Request):
    uid = request.session.get("user_id")
    if not uid:
        return {"authenticated": False}
    with conn() as c:
        u = c.execute("SELECT id, username, name, role, company_id FROM users WHERE id=?", (uid,)).fetchone()
    if not u:
        request.session.clear()
        return {"authenticated": False}
    return {"authenticated": True, **dict(u)}

# ----- 관리자/매니저 자가 가입 -----
@app.post("/api/signup")
def signup(payload: SignupIn, request: Request):
    """본사 직원이 직접 계정 생성. 환경변수 ADMIN_INVITE_CODE 가 설정돼 있으면 일치해야 함.
    설정 안 돼 있으면 누구나 가입 가능 (프로토타입 단계)."""
    username = (payload.username or "").strip().lower()
    password = payload.password or ""
    name = (payload.name or "").strip()
    if len(username) < 3:
        raise HTTPException(400, "아이디는 3자 이상이어야 합니다.")
    if not username.replace('_','').replace('.','').isalnum():
        raise HTTPException(400, "아이디는 영문·숫자·_·. 만 사용 가능합니다.")
    if len(password) < 6:
        raise HTTPException(400, "비밀번호는 6자 이상이어야 합니다.")
    if len(name) < 2:
        raise HTTPException(400, "이름을 입력해주세요.")

    invite_code_required = os.environ.get("ADMIN_INVITE_CODE", "")
    if invite_code_required and (payload.invite_code or "") != invite_code_required:
        raise HTTPException(403, "초대 코드가 맞지 않습니다. 본사에 문의해주세요.")

    with conn() as c:
        dup = c.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()
        if dup:
            raise HTTPException(409, "이미 사용 중인 아이디입니다.")
        # 첫 가입자는 admin, 나머지는 manager
        n = c.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        role = "admin" if n == 0 else "manager"
        cur = c.execute("INSERT INTO users(username,password_hash,name,role) VALUES(?,?,?,?)",
                        (username, hash_pw(password), name, role))
        new_id = cur.lastrowid
    # 가입 즉시 로그인 처리
    request.session["user_id"] = new_id
    request.session["username"] = username
    request.session["role"] = role
    emit_event("AdminSignedUp",
               actors={"user_id": new_id},
               payload={"username": username, "name": name, "role": role},
               created_by=new_id, source="public")
    return {"ok": True, "id": new_id, "username": username, "name": name, "role": role}

@app.get("/api/signup-config")
def signup_config():
    """가입 화면이 초대 코드를 보여줄지 결정."""
    return {"invite_required": bool(os.environ.get("ADMIN_INVITE_CODE", ""))}

# ----- 사용자 관리 (admin 전용) -----
@app.get("/api/users")
def list_users(user: dict = Depends(require_login)):
    if user.get("role") != "admin":
        raise HTTPException(403, "관리자 전용입니다.")
    with conn() as c:
        return rows(c.execute(
            "SELECT id, username, name, role, company_id, created_at FROM users ORDER BY id"
        ).fetchall())

class UserUpdateIn(BaseModel):
    role: Optional[str] = None
    name: Optional[str] = None
    new_password: Optional[str] = None

@app.put("/api/users/{uid}")
def update_user(uid: int, payload: UserUpdateIn, user: dict = Depends(require_login)):
    if user.get("role") != "admin":
        raise HTTPException(403, "관리자 전용입니다.")
    fields, vals = [], []
    if payload.role and payload.role in ("admin", "manager"):
        fields.append("role=?"); vals.append(payload.role)
    if payload.name is not None:
        fields.append("name=?"); vals.append(payload.name)
    if payload.new_password:
        if len(payload.new_password) < 6:
            raise HTTPException(400, "비밀번호는 6자 이상이어야 합니다.")
        fields.append("password_hash=?"); vals.append(hash_pw(payload.new_password))
    if not fields:
        return {"ok": True}
    vals.append(uid)
    with conn() as c:
        c.execute(f"UPDATE users SET {','.join(fields)} WHERE id=?", vals)
    return {"ok": True}

@app.delete("/api/users/{uid}")
def delete_user(uid: int, user: dict = Depends(require_login)):
    if user.get("role") != "admin":
        raise HTTPException(403, "관리자 전용입니다.")
    if uid == user["id"]:
        raise HTTPException(400, "본인 계정은 삭제할 수 없습니다.")
    with conn() as c:
        c.execute("DELETE FROM users WHERE id=?", (uid,))
    return {"ok": True}

# ----- 워커(폰) 본인인증 -----
def normalize_phone(phone: str) -> str:
    return ''.join(ch for ch in (phone or "") if ch.isdigit())

@app.post("/api/worker/identify")
def worker_identify(payload: WorkerIdentifyIn, request: Request):
    """폰번호로 본인 확인. 일치하면 세션에 worker_id 저장."""
    digits = normalize_phone(payload.phone)
    if len(digits) < 10:
        raise HTTPException(400, "전화번호를 010-XXXX-XXXX 형식으로 입력해주세요.")
    with conn() as c:
        rows_ = c.execute(
            "SELECT id, name, phone, worker_type, job_role FROM workers").fetchall()
    me = None
    for r in rows_:
        if normalize_phone(r["phone"]) == digits:
            me = dict(r); break
    if not me:
        raise HTTPException(404, "등록된 번호가 아닙니다. 신규라면 가입을 먼저 해주세요.")
    request.session["worker_id"] = me["id"]
    request.session["worker_name"] = me["name"]
    return {"ok": True, **me}

@app.get("/api/worker/me")
def worker_me(request: Request):
    wid = request.session.get("worker_id")
    if not wid:
        return {"identified": False}
    with conn() as c:
        r = c.execute(
            "SELECT id, name, phone, worker_type, job_role FROM workers WHERE id=?",
            (wid,)).fetchone()
    if not r:
        request.session.pop("worker_id", None)
        request.session.pop("worker_name", None)
        return {"identified": False}
    return {"identified": True, **dict(r)}

@app.post("/api/worker/logout")
def worker_logout(request: Request):
    request.session.pop("worker_id", None)
    request.session.pop("worker_name", None)
    return {"ok": True}

@app.post("/api/me/password")
def change_password(payload: PasswordChangeIn, user: dict = Depends(require_login)):
    with conn() as c:
        u = c.execute("SELECT password_hash FROM users WHERE id=?", (user["id"],)).fetchone()
        if not u or not verify_pw(payload.current_password, u["password_hash"]):
            raise HTTPException(401, "현재 비밀번호가 맞지 않습니다")
        if len(payload.new_password) < 6:
            raise HTTPException(400, "새 비밀번호는 6자 이상이어야 합니다")
        c.execute("UPDATE users SET password_hash=? WHERE id=?",
                  (hash_pw(payload.new_password), user["id"]))
    return {"ok": True}

# ----- Companies -----
@app.get("/api/companies")
def list_companies(_: dict = Depends(require_login)):
    with conn() as c:
        return rows(c.execute("SELECT * FROM companies ORDER BY id").fetchall())

@app.post("/api/companies")
def create_company(payload: CompanyIn, user: dict = Depends(require_login)):
    with conn() as c:
        cur = c.execute(
            """INSERT INTO companies(name,business_no,corporate_no,ceo,license_info,address,
                                     phone,email,representative_phone,
                                     fiscal_year_end,incorporation_date,registration_date)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
            (payload.name, payload.business_no, payload.corporate_no, payload.ceo,
             payload.license_info, payload.address,
             payload.phone, payload.email, payload.representative_phone,
             payload.fiscal_year_end,
             payload.incorporation_date, payload.registration_date)
        )
        new_id = cur.lastrowid
    emit_event("CompanyCreated",
               actors={"company_id": new_id},
               payload={"name": payload.name, "business_no": payload.business_no,
                        "corporate_no": payload.corporate_no, "ceo": payload.ceo},
               created_by=user["id"], source="admin_ui")
    return {"id": new_id}

@app.put("/api/companies/{cid}")
def update_company(cid: int, payload: CompanyIn, user: dict = Depends(require_login)):
    with conn() as c:
        existing = c.execute("SELECT * FROM companies WHERE id=?", (cid,)).fetchone()
        if not existing:
            raise HTTPException(404, "company not found")
        c.execute(
            """UPDATE companies SET name=?, business_no=?, corporate_no=?, ceo=?, license_info=?,
                                    address=?, phone=?, email=?, representative_phone=?,
                                    fiscal_year_end=?, incorporation_date=?, registration_date=?
               WHERE id=?""",
            (payload.name, payload.business_no, payload.corporate_no, payload.ceo,
             payload.license_info, payload.address,
             payload.phone, payload.email, payload.representative_phone,
             payload.fiscal_year_end, payload.incorporation_date, payload.registration_date, cid)
        )
    emit_event("CompanyUpdated",
               actors={"company_id": cid},
               payload={"name": payload.name, "business_no": payload.business_no,
                        "ceo": payload.ceo, "address": payload.address},
               created_by=user["id"], source="admin_ui")
    return {"ok": True}

@app.delete("/api/companies/{cid}")
def delete_company(cid: int, user: dict = Depends(require_login)):
    with conn() as c:
        # 의존성 확인 — 워커/현장이 이 법인을 참조 중이면 거부
        worker_cnt = c.execute("SELECT COUNT(*) FROM workers WHERE company_id=?", (cid,)).fetchone()[0]
        site_cnt = c.execute("SELECT COUNT(*) FROM sites WHERE company_id=?", (cid,)).fetchone()[0]
        if worker_cnt or site_cnt:
            raise HTTPException(400, f"이 법인을 참조 중인 직원 {worker_cnt}명, 현장 {site_cnt}건이 있습니다. 먼저 다른 법인으로 옮기거나 정리해주세요.")
        c.execute("DELETE FROM companies WHERE id=?", (cid,))
    emit_event("CompanyDeleted", actors={"company_id": cid},
               created_by=user["id"], source="admin_ui")
    # 관계 정리
    remove_relations(subject_type='Org', subject_id=cid)
    remove_relations(object_type='Org', object_id=cid)
    return {"ok": True}

# ----- 모바일 출퇴근 화면용 공개 API (최소 정보만) -----
@app.get("/api/public/clock-options")
def public_clock_options():
    with conn() as c:
        workers = rows(c.execute(
            "SELECT id, name, job_role, worker_type FROM workers ORDER BY name").fetchall())
        sites = rows(c.execute(
            "SELECT id, name FROM sites WHERE status='active' ORDER BY name").fetchall())
    return {"workers": workers, "sites": sites}

# ----- Sites -----
@app.get("/api/sites")
def list_sites(active_only: bool = False, _: dict = Depends(require_login)):
    sql = "SELECT s.*, c.name AS company_name FROM sites s LEFT JOIN companies c ON s.company_id=c.id"
    if active_only:
        sql += " WHERE s.status='active'"
    sql += " ORDER BY s.id DESC"
    with conn() as c:
        return rows(c.execute(sql).fetchall())

@app.post("/api/sites")
def create_site(payload: SiteIn, user: dict = Depends(require_login)):
    with conn() as c:
        cur = c.execute(
            """INSERT INTO sites(company_id,name,address,latitude,longitude,geofence_meters,
               contract_amount,paid_amount,start_date,end_date,status,manager)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
            (payload.company_id, payload.name, payload.address, payload.latitude, payload.longitude,
             payload.geofence_meters or 200, payload.contract_amount or 0, payload.paid_amount or 0,
             payload.start_date, payload.end_date, payload.status or "active", payload.manager)
        )
        new_id = cur.lastrowid
    emit_event("SiteCreated",
               actors={"company_id": payload.company_id},
               place={"site_id": new_id, "lat": payload.latitude, "lng": payload.longitude,
                      "address": payload.address},
               payload={"name": payload.name, "manager": payload.manager,
                        "start_date": payload.start_date, "end_date": payload.end_date},
               financial={"amount": payload.contract_amount or 0, "account": "계약금액", "kind": "contract"},
               created_by=user["id"], source="admin_ui")
    if payload.company_id:
        add_relation('Place', new_id, 'owned_by', 'Org', payload.company_id)
    return {"id": new_id}

@app.put("/api/sites/{sid}")
def update_site(sid: int, payload: SiteIn, user: dict = Depends(require_login)):
    with conn() as c:
        c.execute(
            """UPDATE sites SET company_id=?, name=?, address=?, latitude=?, longitude=?,
               geofence_meters=?, contract_amount=?, paid_amount=?, start_date=?, end_date=?,
               status=?, manager=? WHERE id=?""",
            (payload.company_id, payload.name, payload.address, payload.latitude, payload.longitude,
             payload.geofence_meters or 200, payload.contract_amount or 0, payload.paid_amount or 0,
             payload.start_date, payload.end_date, payload.status or "active", payload.manager, sid)
        )
    emit_event("SiteUpdated",
               place={"site_id": sid},
               payload={"name": payload.name, "status": payload.status,
                        "contract_amount": payload.contract_amount, "paid_amount": payload.paid_amount},
               created_by=user["id"], source="admin_ui")
    # 관계 동기화 (소속 법인 변경 가능)
    remove_relations(subject_type='Place', subject_id=sid, predicate='owned_by')
    if payload.company_id:
        add_relation('Place', sid, 'owned_by', 'Org', payload.company_id)
    return {"ok": True}

@app.delete("/api/sites/{sid}")
def delete_site(sid: int, user: dict = Depends(require_login)):
    with conn() as c:
        c.execute("DELETE FROM sites WHERE id=?", (sid,))
    emit_event("SiteDeleted", place={"site_id": sid},
               created_by=user["id"], source="admin_ui")
    remove_relations(subject_type='Place', subject_id=sid)
    remove_relations(object_type='Place', object_id=sid)
    return {"ok": True}

# ----- Workers -----
@app.get("/api/workers")
def list_workers(worker_type: Optional[str] = None, q: Optional[str] = None,
                 _: dict = Depends(require_login)):
    sql = "SELECT w.*, c.name AS company_name FROM workers w LEFT JOIN companies c ON w.company_id=c.id WHERE 1=1"
    args = []
    if worker_type:
        sql += " AND w.worker_type=?"; args.append(worker_type)
    if q:
        sql += " AND (w.name LIKE ? OR w.phone LIKE ?)"
        args += [f"%{q}%", f"%{q}%"]
    sql += " ORDER BY w.name"
    with conn() as c:
        return rows(c.execute(sql, args).fetchall())

@app.post("/api/workers")
def create_worker(payload: WorkerIn, user: dict = Depends(require_login)):
    with conn() as c:
        cur = c.execute(
            """INSERT INTO workers(company_id,name,phone,worker_type,daily_wage,job_role,hired_date,
                                    rrn_last,bank_account,note,rrn,address,position,resigned_at,
                                    bank_name,account_holder,asbestos_certified,job_specialty)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (payload.company_id, payload.name, payload.phone, payload.worker_type or "daily",
             payload.daily_wage or 0, payload.job_role, payload.hired_date, payload.rrn_last,
             payload.bank_account, payload.note,
             payload.rrn, payload.address, payload.position, payload.resigned_at,
             payload.bank_name, payload.account_holder, payload.asbestos_certified or 0,
             payload.job_specialty)
        )
        new_id = cur.lastrowid
    emit_event("WorkerCreated",
               actors={"worker_id": new_id, "company_id": payload.company_id},
               payload={"name": payload.name, "worker_type": payload.worker_type,
                        "job_role": payload.job_role, "daily_wage": payload.daily_wage},
               created_by=user["id"], source="admin_ui")
    if payload.company_id:
        add_relation('Person', new_id, 'employed_by', 'Org', payload.company_id)
    return {"id": new_id}

@app.put("/api/workers/{wid}")
def update_worker(wid: int, payload: WorkerIn, user: dict = Depends(require_login)):
    with conn() as c:
        c.execute(
            """UPDATE workers SET company_id=?, name=?, phone=?, worker_type=?, daily_wage=?,
               job_role=?, hired_date=?, rrn_last=?, bank_account=?, note=?,
               rrn=?, address=?, position=?, resigned_at=?,
               bank_name=?, account_holder=?, asbestos_certified=?, job_specialty=?
               WHERE id=?""",
            (payload.company_id, payload.name, payload.phone, payload.worker_type or "daily",
             payload.daily_wage or 0, payload.job_role, payload.hired_date, payload.rrn_last,
             payload.bank_account, payload.note,
             payload.rrn, payload.address, payload.position, payload.resigned_at,
             payload.bank_name, payload.account_holder, payload.asbestos_certified or 0,
             payload.job_specialty, wid)
        )
    emit_event("WorkerUpdated",
               actors={"worker_id": wid, "company_id": payload.company_id},
               payload={"name": payload.name, "daily_wage": payload.daily_wage,
                        "job_role": payload.job_role},
               created_by=user["id"], source="admin_ui")
    remove_relations(subject_type='Person', subject_id=wid, predicate='employed_by')
    if payload.company_id:
        add_relation('Person', wid, 'employed_by', 'Org', payload.company_id)
    return {"ok": True}

@app.delete("/api/workers/{wid}")
def delete_worker(wid: int, user: dict = Depends(require_login)):
    with conn() as c:
        c.execute("DELETE FROM workers WHERE id=?", (wid,))
    emit_event("WorkerDeleted", actors={"worker_id": wid},
               created_by=user["id"], source="admin_ui")
    remove_relations(subject_type='Person', subject_id=wid)
    remove_relations(object_type='Person', object_id=wid)
    return {"ok": True}

# ----- 직원 자가 가입 (공개 엔드포인트) -----
@app.post("/api/register")
def register_worker(payload: RegisterIn):
    """직원이 폰에서 직접 등록. 이름·번호만으로 가입되며, 본사 관리자가 나중에
    소속 법인·일당 등을 보강한다. 같은 번호로 이미 가입된 경우 거부한다."""
    name = (payload.name or "").strip()
    phone = (payload.phone or "").strip()
    if not name or not phone:
        raise HTTPException(400, "이름과 연락처를 모두 입력해주세요.")
    if len(name) < 2:
        raise HTTPException(400, "이름은 2자 이상이어야 합니다.")
    digits = ''.join(ch for ch in phone if ch.isdigit())
    if len(digits) < 10 or len(digits) > 11:
        raise HTTPException(400, "연락처를 010-XXXX-XXXX 형식으로 입력해주세요.")
    with conn() as c:
        dup = c.execute("SELECT id, name FROM workers WHERE REPLACE(REPLACE(phone,'-',''),' ','')=?", (digits,)).fetchone()
        if dup:
            raise HTTPException(409, f"이미 {dup['name']} 님으로 가입된 번호입니다.")
        cur = c.execute(
            """INSERT INTO workers(name,phone,worker_type,job_role,note,hired_date)
               VALUES(?,?,?,?,?,?)""",
            (name, phone, payload.worker_type or "daily", payload.job_role,
             "자가 가입 — 본사 검토 대기", date.today().isoformat())
        )
        new_id = cur.lastrowid
    emit_event("WorkerSelfRegistered",
               actors={"worker_id": new_id},
               payload={"name": name, "phone": phone,
                        "worker_type": payload.worker_type, "job_role": payload.job_role},
               source="public")
    return {"ok": True, "id": new_id, "name": name}

# ----- Deployments -----
@app.get("/api/deployments/range")
def list_deployments_range(
    start: str = Query(...), end: str = Query(...),
    kind: Optional[str] = None,
    _: dict = Depends(require_login)
):
    """기간 범위의 모든 배치 — 주간 보드용."""
    sql = """SELECT d.*, w.name AS worker_name, w.worker_type, w.daily_wage,
             w.job_role, s.name AS site_name FROM deployments d
             JOIN workers w ON d.worker_id=w.id
             JOIN sites s ON d.site_id=s.id
             WHERE d.date BETWEEN ? AND ?"""
    args = [start, end]
    if kind:
        sql += " AND d.kind=?"; args.append(kind)
    sql += " ORDER BY d.date, s.name, w.name"
    with conn() as c:
        return rows(c.execute(sql, args).fetchall())

@app.get("/api/deployments")
def list_deployments(date: str = Query(...), kind: Optional[str] = None,
                     _: dict = Depends(require_login)):
    sql = """SELECT d.*, w.name AS worker_name, w.worker_type, w.daily_wage,
             s.name AS site_name FROM deployments d
             JOIN workers w ON d.worker_id=w.id
             JOIN sites s ON d.site_id=s.id
             WHERE d.date=?"""
    args = [date]
    if kind:
        sql += " AND d.kind=?"; args.append(kind)
    sql += " ORDER BY s.name, w.name"
    with conn() as c:
        return rows(c.execute(sql, args).fetchall())

@app.post("/api/deployments")
def upsert_deployment(payload: DeploymentIn, user: dict = Depends(require_login)):
    with conn() as c:
        c.execute("DELETE FROM deployments WHERE worker_id=? AND date=? AND kind=?",
                  (payload.worker_id, payload.date, payload.kind))
        cur = c.execute(
            "INSERT INTO deployments(worker_id,site_id,date,kind,note) VALUES(?,?,?,?,?)",
            (payload.worker_id, payload.site_id, payload.date, payload.kind, payload.note)
        )
        new_id = cur.lastrowid
    emit_event("Deploy",
               actors={"worker_id": payload.worker_id},
               place={"site_id": payload.site_id},
               payload={"date": payload.date, "kind": payload.kind, "note": payload.note},
               created_by=user["id"], source="admin_ui")
    return {"id": new_id}

@app.delete("/api/deployments/{did}")
def delete_deployment(did: int, user: dict = Depends(require_login)):
    with conn() as c:
        row = c.execute(
            "SELECT worker_id, site_id, date, kind FROM deployments WHERE id=?", (did,)
        ).fetchone()
        c.execute("DELETE FROM deployments WHERE id=?", (did,))
    if row:
        emit_event("DeploymentRemoved",
                   actors={"worker_id": row["worker_id"]},
                   place={"site_id": row["site_id"]},
                   payload={"date": row["date"], "kind": row["kind"]},
                   created_by=user["id"], source="admin_ui")
    return {"ok": True}

@app.post("/api/deployments/copy")
def copy_deployments(src_kind: str = Body(...), dst_kind: str = Body(...), date: str = Body(...),
                     user: dict = Depends(require_login)):
    """계획 → 실적 복사 같은 운영 편의 기능."""
    with conn() as c:
        c.execute("DELETE FROM deployments WHERE date=? AND kind=?", (date, dst_kind))
        c.execute(
            """INSERT INTO deployments(worker_id,site_id,date,kind,note)
               SELECT worker_id,site_id,date,?,note FROM deployments WHERE date=? AND kind=?""",
            (dst_kind, date, src_kind)
        )
        n = c.execute("SELECT COUNT(*) FROM deployments WHERE date=? AND kind=?",
                      (date, dst_kind)).fetchone()[0]
    emit_event("DeploymentsCopied",
               payload={"date": date, "src_kind": src_kind, "dst_kind": dst_kind, "count": n},
               created_by=user["id"], source="admin_ui")
    return {"ok": True}

# ----- Clock in/out (mobile, GPS) -----
@app.post("/api/clock")
def clock(payload: ClockIn, request: Request):
    # 세션의 worker_id 만 신뢰. 본인인증 안 됐으면 거부.
    session_wid = request.session.get("worker_id")
    if not session_wid:
        raise HTTPException(401, "본인인증이 필요합니다. /m 에서 전화번호로 인증해주세요.")
    # body 의 worker_id 는 무시하고 세션값으로 강제
    payload.worker_id = session_wid

    with conn() as c:
        site = c.execute("SELECT * FROM sites WHERE id=?", (payload.site_id,)).fetchone()
        if not site:
            raise HTTPException(404, "site not found")
        worker = c.execute("SELECT * FROM workers WHERE id=?", (payload.worker_id,)).fetchone()
        if not worker:
            raise HTTPException(404, "worker not found")

        today = date.today().isoformat()
        now = datetime.now().isoformat(timespec="seconds")
        dist = haversine_m(payload.lat, payload.lng, site["latitude"], site["longitude"])
        verified = 1 if (dist is not None and dist <= (site["geofence_meters"] or 200)) else 0

        existing = c.execute(
            "SELECT * FROM clock_records WHERE worker_id=? AND date=?",
            (payload.worker_id, today)
        ).fetchone()

        if payload.direction == "in":
            if existing:
                c.execute("""UPDATE clock_records SET clock_in=?, in_lat=?, in_lng=?,
                             in_distance_m=?, in_verified=?, site_id=? WHERE id=?""",
                          (now, payload.lat, payload.lng, dist, verified, payload.site_id, existing["id"]))
            else:
                c.execute("""INSERT INTO clock_records(worker_id,site_id,date,clock_in,in_lat,in_lng,in_distance_m,in_verified)
                             VALUES(?,?,?,?,?,?,?,?)""",
                          (payload.worker_id, payload.site_id, today, now, payload.lat, payload.lng, dist, verified))
            # 자동으로 실적 배치도 기록
            c.execute("DELETE FROM deployments WHERE worker_id=? AND date=? AND kind='actual'",
                      (payload.worker_id, today))
            c.execute("INSERT INTO deployments(worker_id,site_id,date,kind,note) VALUES(?,?,?,'actual','GPS 출근 자동')",
                      (payload.worker_id, payload.site_id, today))
        else:  # out
            if not existing:
                raise HTTPException(400, "출근 기록이 없습니다")
            c.execute("""UPDATE clock_records SET clock_out=?, out_lat=?, out_lng=?,
                         out_distance_m=?, out_verified=? WHERE id=?""",
                      (now, payload.lat, payload.lng, dist, verified, existing["id"]))

    # ── 이벤트 기록 ──
    event_type = "ClockIn" if payload.direction == "in" else "ClockOut"
    financial = None
    if payload.direction == "in":
        financial = {
            "amount": worker["daily_wage"] or 0,
            "account": f"직접노무비/{site['name']}",
            "kind": "expense",
        }
    emit_event(event_type,
               actors={"worker_id": payload.worker_id, "worker_name": worker["name"]},
               place={"site_id": payload.site_id, "site_name": site["name"],
                      "lat": payload.lat, "lng": payload.lng},
               payload={"date": today, "distance_m": round(dist, 1) if dist is not None else None,
                        "verified": bool(verified), "geofence_m": site["geofence_meters"] or 200},
               financial=financial, source="mobile")
    return {
        "ok": True, "verified": bool(verified),
        "distance_m": round(dist, 1) if dist is not None else None,
        "geofence_m": site["geofence_meters"] or 200,
    }

@app.get("/api/clock/today")
def clock_today(_: dict = Depends(require_login)):
    today = date.today().isoformat()
    with conn() as c:
        return rows(c.execute("""
            SELECT cr.*, w.name AS worker_name, s.name AS site_name
            FROM clock_records cr
            JOIN workers w ON cr.worker_id=w.id
            JOIN sites s ON cr.site_id=s.id
            WHERE cr.date=? ORDER BY cr.clock_in DESC
        """, (today,)).fetchall())

# ----- Events (Phase 1: 디지털 트윈 코어) -----
@app.get("/api/events")
def list_events(
    type: Optional[str] = None,
    types: Optional[str] = None,            # 콤마 구분 다중 (e.g. "ClockIn,ClockOut")
    site_id: Optional[int] = None,
    worker_id: Optional[int] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    source: Optional[str] = None,
    limit: int = 200,
    _: dict = Depends(require_login),
):
    """모든 이벤트 조회. 디지털 트윈의 단일 진실 원본(single source of truth)."""
    sql = "SELECT * FROM events WHERE 1=1"
    args = []
    if type:
        sql += " AND type=?"; args.append(type)
    if types:
        ts = [t.strip() for t in types.split(",") if t.strip()]
        if ts:
            sql += " AND type IN (" + ",".join(["?"]*len(ts)) + ")"
            args.extend(ts)
    if site_id is not None:
        sql += " AND json_extract(place,'$.site_id') = ?"; args.append(site_id)
    if worker_id is not None:
        sql += " AND json_extract(actors,'$.worker_id') = ?"; args.append(worker_id)
    if from_date:
        sql += " AND occurred_at >= ?"; args.append(from_date)
    if to_date:
        sql += " AND occurred_at <= ?"; args.append(to_date + "T23:59:59")
    if source:
        sql += " AND source=?"; args.append(source)
    sql += " ORDER BY occurred_at DESC, id DESC LIMIT ?"
    args.append(min(limit, 1000))
    with conn() as c:
        rs = rows(c.execute(sql, args).fetchall())
    # JSON 파싱해서 클라이언트가 쓰기 쉽게
    for r in rs:
        for k in ("actors", "place", "payload", "financial"):
            try: r[k] = json.loads(r.get(k) or "{}")
            except Exception: r[k] = {}
    return rs

@app.get("/api/events/types")
def list_event_types(_: dict = Depends(require_login)):
    """기록된 이벤트 타입 종류 + 각 카운트."""
    with conn() as c:
        return rows(c.execute(
            "SELECT type, COUNT(*) AS cnt FROM events GROUP BY type ORDER BY cnt DESC"
        ).fetchall())

# ----- 그래프 / 온톨로지 (Phase 2) -----
def _entity_lookup(entity_type, entity_id):
    """엔티티 타입+id로 실제 데이터 한 줄 가져오기."""
    if entity_type not in ENTITY_TABLES:
        return None
    table, name_col = ENTITY_TABLES[entity_type]
    with conn() as c:
        row = c.execute(f"SELECT * FROM {table} WHERE id=?", (entity_id,)).fetchone()
    return dict(row) if row else None

@app.get("/api/graph/entities")
def list_entities(entity_type: str, _: dict = Depends(require_login)):
    """엔티티 타입별 목록 (그래프 뷰 진입점)."""
    if entity_type not in ENTITY_TABLES:
        raise HTTPException(400, f"unknown entity type: {entity_type}")
    table, name_col = ENTITY_TABLES[entity_type]
    if entity_type == 'Person':
        sql = "SELECT id, name, worker_type, job_role, phone FROM workers ORDER BY name"
    elif entity_type == 'Place':
        sql = "SELECT id, name, address, status FROM sites ORDER BY status, name"
    elif entity_type == 'Org':
        sql = "SELECT id, name, business_no, ceo FROM companies ORDER BY name"
    elif entity_type == 'User':
        sql = "SELECT id, username AS name, role, company_id FROM users ORDER BY username"
    else:
        sql = f"SELECT id, {name_col} AS name FROM {table} ORDER BY {name_col}"
    with conn() as c:
        return rows(c.execute(sql).fetchall())

@app.get("/api/graph/entity/{entity_type}/{entity_id}")
def graph_entity(entity_type: str, entity_id: int, _: dict = Depends(require_login)):
    """한 엔티티의 전체 그래프 뷰 — 본인 정보 + 모든 관계 + 최근 이벤트."""
    entity = _entity_lookup(entity_type, entity_id)
    if not entity:
        raise HTTPException(404, "entity not found")

    with conn() as c:
        # 1) 이 엔티티가 subject 인 관계
        outgoing = rows(c.execute(
            "SELECT * FROM relations WHERE subject_type=? AND subject_id=?",
            (entity_type, entity_id)).fetchall())
        # 2) 이 엔티티가 object 인 관계
        incoming = rows(c.execute(
            "SELECT * FROM relations WHERE object_type=? AND object_id=?",
            (entity_type, entity_id)).fetchall())

        # 관계 끝의 엔티티 이름 채워주기
        for r in outgoing:
            target = _entity_lookup(r['object_type'], r['object_id'])
            r['object_name'] = (target.get('name') or target.get('username') or '?') if target else '(삭제됨)'
        for r in incoming:
            source = _entity_lookup(r['subject_type'], r['subject_id'])
            r['subject_name'] = (source.get('name') or source.get('username') or '?') if source else '(삭제됨)'

        # 3) 이 엔티티 관련 최근 이벤트 (90일)
        cutoff = (datetime.now().date().toordinal() - 90)
        # 이벤트는 actors/place JSON 안에 id 가 들어감
        if entity_type == 'Person':
            evt_rows = c.execute(
                "SELECT * FROM events WHERE json_extract(actors,'$.worker_id')=? "
                "OR json_extract(actors,'$.user_id')=? "
                "ORDER BY occurred_at DESC LIMIT 100",
                (entity_id, entity_id)).fetchall()
        elif entity_type == 'Place':
            evt_rows = c.execute(
                "SELECT * FROM events WHERE json_extract(place,'$.site_id')=? "
                "ORDER BY occurred_at DESC LIMIT 100",
                (entity_id,)).fetchall()
        elif entity_type == 'Org':
            evt_rows = c.execute(
                "SELECT * FROM events WHERE json_extract(actors,'$.company_id')=? "
                "ORDER BY occurred_at DESC LIMIT 100",
                (entity_id,)).fetchall()
        else:
            evt_rows = []
        events_out = []
        for r in evt_rows:
            d = dict(r)
            for k in ("actors", "place", "payload", "financial"):
                try: d[k] = json.loads(d.get(k) or "{}")
                except Exception: d[k] = {}
            events_out.append(d)

    return {
        "entity_type": entity_type,
        "entity_id": entity_id,
        "entity": entity,
        "outgoing_relations": outgoing,    # 내가 ~한다
        "incoming_relations": incoming,    # ~가 나를 한다
        "recent_events": events_out,
    }

@app.get("/api/graph/predicates")
def list_predicates(_: dict = Depends(require_login)):
    """존재하는 관계 종류 + 카운트."""
    with conn() as c:
        return rows(c.execute(
            "SELECT predicate, subject_type, object_type, COUNT(*) AS cnt "
            "FROM relations GROUP BY predicate, subject_type, object_type ORDER BY cnt DESC"
        ).fetchall())

# ----- Reality Views (Phase 3) — 같은 events 를 3개 시점으로 -----

def _parse_event_json(r):
    d = dict(r)
    for k in ('actors','place','payload','financial'):
        try: d[k] = json.loads(d.get(k) or '{}')
        except Exception: d[k] = {}
    return d

@app.get("/api/views/field")
def view_field(site_id: Optional[int] = None, days: int = 1,
               _: dict = Depends(require_login)):
    """현장 시점 — 오늘/최근 며칠의 실시간 활동 중심."""
    today_d = date.today()
    cutoff = (today_d.replace(day=today_d.day) if days <= 1 else today_d).isoformat()
    if days > 1:
        from datetime import timedelta
        cutoff = (today_d - timedelta(days=days-1)).isoformat()

    with conn() as c:
        # 활성 현장 + 오늘 인원 카운트
        site_q = "SELECT id, name, address, latitude, longitude, geofence_meters FROM sites WHERE status='active'"
        if site_id:
            site_q += " AND id=?"; sites_arg = (site_id,)
        else:
            sites_arg = ()
        active_sites = rows(c.execute(site_q, sites_arg).fetchall())
        for s in active_sites:
            cnt = c.execute(
                "SELECT COUNT(DISTINCT json_extract(actors,'$.worker_id')) FROM events "
                "WHERE type='ClockIn' AND occurred_at >= ? AND json_extract(place,'$.site_id')=?",
                (today_d.isoformat(), s['id'])).fetchone()[0]
            s['clocked_in_today'] = cnt or 0

        # 최근 24h~며칠 events (출퇴근/배치/안전 등 현장에서 일어난 것)
        evt_q = ("SELECT * FROM events WHERE type IN ('ClockIn','ClockOut','Deploy','DeploymentRemoved') "
                 "AND occurred_at >= ?")
        evt_args = [cutoff]
        if site_id:
            evt_q += " AND json_extract(place,'$.site_id')=?"
            evt_args.append(site_id)
        evt_q += " ORDER BY occurred_at DESC LIMIT 200"
        recent = [_parse_event_json(r) for r in c.execute(evt_q, evt_args).fetchall()]

    return {
        "as_of": today_d.isoformat(),
        "days": days,
        "active_sites": active_sites,
        "recent_events": recent,
    }

@app.get("/api/views/admin")
def view_admin(_: dict = Depends(require_login)):
    """행정 시점 — 신고 대상자, 처리 대기 항목, 컴플라이언스 갭."""
    today_d = date.today()
    month_start = today_d.replace(day=1).isoformat()
    week_ago = (datetime.fromisoformat(today_d.isoformat()) -
                __import__('datetime').timedelta(days=7)).date().isoformat()

    with conn() as c:
        # 1) 자가가입 검토 대기
        pending = rows(c.execute(
            "SELECT id, name, phone, hired_date, note FROM workers "
            "WHERE note LIKE '%검토 대기%' OR (daily_wage = 0 AND worker_type='daily') "
            "ORDER BY hired_date DESC LIMIT 50"
        ).fetchall())

        # 2) 이번 달 일용근로내용확인신고 대상 (ClockIn 누적)
        report_targets_raw = rows(c.execute("""
            SELECT json_extract(actors,'$.worker_id') AS worker_id,
                   json_extract(actors,'$.worker_name') AS worker_name,
                   COUNT(*) AS days,
                   GROUP_CONCAT(DISTINCT json_extract(place,'$.site_name')) AS sites
            FROM events
            WHERE type='ClockIn' AND occurred_at >= ?
            GROUP BY worker_id ORDER BY days DESC
        """, (month_start,)).fetchall())

        # 3) GPS 좌표 없는 활성 현장
        sites_no_gps = rows(c.execute(
            "SELECT id, name, address FROM sites WHERE status='active' "
            "AND (latitude IS NULL OR longitude IS NULL OR latitude = 0)"
        ).fetchall())

        # 4) 법인 미배정 워커
        no_company = rows(c.execute(
            "SELECT id, name, phone, worker_type, hired_date FROM workers "
            "WHERE company_id IS NULL ORDER BY hired_date DESC"
        ).fetchall())

        # 5) 일당 미설정 워커 (일용직만)
        no_wage = rows(c.execute(
            "SELECT id, name, phone, hired_date FROM workers "
            "WHERE worker_type='daily' AND (daily_wage IS NULL OR daily_wage = 0) "
            "ORDER BY hired_date DESC"
        ).fetchall())

        # 6) 최근 7일 신규 가입 (자가)
        recent_signups = rows(c.execute(
            "SELECT id, name, phone, hired_date FROM workers "
            "WHERE hired_date >= ? ORDER BY hired_date DESC LIMIT 30",
            (week_ago,)
        ).fetchall())

    return {
        "month_start": month_start,
        "pending_review": pending,
        "report_targets_this_month": report_targets_raw,
        "sites_missing_gps": sites_no_gps,
        "workers_no_company": no_company,
        "workers_no_wage": no_wage,
        "recent_signups_7d": recent_signups,
    }

@app.get("/api/views/finance")
def view_finance(
    site_id: Optional[int] = None,
    company_id: Optional[int] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    _: dict = Depends(require_login),
):
    """재무 시점 — events 의 financial 데이터를 다차원 집계."""
    sql = ("SELECT * FROM events "
           "WHERE json_extract(financial,'$.amount') IS NOT NULL "
           "AND CAST(json_extract(financial,'$.amount') AS INTEGER) > 0")
    args = []
    if site_id:
        sql += " AND json_extract(place,'$.site_id') = ?"; args.append(site_id)
    if company_id:
        sql += " AND (json_extract(actors,'$.company_id') = ? "
        sql += "      OR json_extract(place,'$.site_id') IN (SELECT id FROM sites WHERE company_id=?))"
        args += [company_id, company_id]
    if from_date:
        sql += " AND occurred_at >= ?"; args.append(from_date)
    if to_date:
        sql += " AND occurred_at <= ?"; args.append(to_date + "T23:59:59")
    sql += " ORDER BY occurred_at DESC LIMIT 500"

    with conn() as c:
        ledger = [_parse_event_json(r) for r in c.execute(sql, args).fetchall()]

        # 사이트/법인 이름 매핑
        site_names = {r['id']: r['name'] for r in c.execute("SELECT id, name FROM sites").fetchall()}
        site_company = {r['id']: r['company_id'] for r in c.execute("SELECT id, company_id FROM sites").fetchall()}
        company_names = {r['id']: r['name'] for r in c.execute("SELECT id, name FROM companies").fetchall()}

    by_site = {}      # site_id -> {expense, contract, revenue}
    by_company = {}
    by_account = {}
    by_month = {}     # YYYY-MM -> {expense, contract, revenue}

    for r in ledger:
        fin = r.get('financial', {}) or {}
        amt = int(fin.get('amount') or 0)
        if amt <= 0: continue
        kind = fin.get('kind') or 'expense'   # expense | contract | revenue
        acct = fin.get('account') or '미분류'
        sid = r.get('place', {}).get('site_id')
        cid = (r.get('actors', {}).get('company_id')
               or (site_company.get(sid) if sid else None))
        month = (r.get('occurred_at') or '')[:7]

        # 이름 채워주기
        if sid: r['_site_name'] = site_names.get(sid)
        if cid: r['_company_name'] = company_names.get(cid)

        def _bump(d, key):
            if key not in d: d[key] = {'expense': 0, 'contract': 0, 'revenue': 0, 'count': 0}
            d[key][kind] = d[key].get(kind, 0) + amt
            d[key]['count'] += 1

        if sid: _bump(by_site, sid)
        if cid: _bump(by_company, cid)
        _bump(by_account, acct)
        if month: _bump(by_month, month)

    # dict → list with names for client
    def _list(d, name_map):
        return [{'id': k, 'name': name_map.get(k, f'#{k}'), **v} for k, v in d.items()]

    return {
        "ledger": ledger,
        "by_site": _list(by_site, site_names),
        "by_company": _list(by_company, company_names),
        "by_account": [{'name': k, **v} for k, v in by_account.items()],
        "by_month": [{'name': k, **v} for k, v in sorted(by_month.items())],
        "totals": {
            "expense":  sum(v.get('expense', 0)  for v in by_account.values()),
            "contract": sum(v.get('contract', 0) for v in by_account.values()),
            "revenue":  sum(v.get('revenue', 0)  for v in by_account.values()),
        }
    }

# ----- Processes (Phase 4) -----
@app.get("/api/process-definitions")
def list_process_defs(_: dict = Depends(require_login)):
    return [{"id": k, **v} for k, v in PROCESS_DEFS.items()]

@app.get("/api/processes")
def list_processes(workflow: Optional[str] = None,
                   subject_id: Optional[int] = None,
                   _: dict = Depends(require_login)):
    sql = """SELECT p.*,
             CASE p.subject_type
               WHEN 'Place'  THEN (SELECT name FROM sites    WHERE id=p.subject_id)
               WHEN 'Person' THEN (SELECT name FROM workers  WHERE id=p.subject_id)
               WHEN 'Org'    THEN (SELECT name FROM companies WHERE id=p.subject_id)
             END AS subject_name
             FROM process_instances p WHERE 1=1"""
    args = []
    if workflow:
        sql += " AND p.workflow=?"; args.append(workflow)
    if subject_id is not None:
        sql += " AND p.subject_id=?"; args.append(subject_id)
    sql += " ORDER BY p.updated_at DESC, p.started_at DESC LIMIT 500"
    with conn() as c:
        return rows(c.execute(sql, args).fetchall())

class ProcessAdvanceIn(BaseModel):
    target_state: str
    note: Optional[str] = None

@app.post("/api/processes/{pid}/advance")
def advance_process(pid: int, payload: ProcessAdvanceIn, user: dict = Depends(require_login)):
    with conn() as c:
        row = c.execute("SELECT * FROM process_instances WHERE id=?", (pid,)).fetchone()
        if not row:
            raise HTTPException(404, "process not found")
        defn = PROCESS_DEFS.get(row['workflow'])
        if not defn or payload.target_state not in defn['states']:
            raise HTTPException(400, "invalid target state")
        completed = (payload.target_state in defn.get('terminal', []))
        c.execute(
            "UPDATE process_instances SET current_state=?, updated_at=datetime('now'), "
            "completed_at=CASE WHEN ?=1 THEN datetime('now') ELSE completed_at END WHERE id=?",
            (payload.target_state, 1 if completed else 0, pid)
        )
    emit_event("ProcessAdvanced", payload={
        "workflow": row['workflow'], "subject_type": row['subject_type'],
        "subject_id": row['subject_id'], "from": row['current_state'],
        "to": payload.target_state, "note": payload.note, "manual": True
    }, created_by=user["id"], source='admin_ui')
    return {"ok": True}

def _backfill_processes():
    """기존 sites/workers 에 프로세스 인스턴스가 없으면 만들어줌. 멱등."""
    try:
        with conn() as c:
            n_proc = c.execute("SELECT COUNT(*) FROM process_instances").fetchone()[0]
            if n_proc > 0:
                return
            for s in c.execute("SELECT id, status FROM sites").fetchall():
                state = '준공' if s['status'] == 'closed' else '진행'
                c.execute("INSERT OR IGNORE INTO process_instances(workflow,subject_type,subject_id,current_state) "
                          "VALUES('sales','Place',?,?)", (s['id'], state))
                c.execute("INSERT OR IGNORE INTO process_instances(workflow,subject_type,subject_id,current_state) "
                          "VALUES('safety','Place',?,?)", (s['id'], '위험성평가'))
            for w in c.execute("SELECT id, daily_wage, note FROM workers").fetchall():
                state = '활성' if (w['daily_wage'] and w['daily_wage'] > 0) else '가입신청'
                c.execute("INSERT OR IGNORE INTO process_instances(workflow,subject_type,subject_id,current_state) "
                          "VALUES('hr_onboarding','Person',?,?)", (w['id'], state))
            print("[startup] processes backfill 완료")
    except Exception as e:
        print(f"[backfill_processes] {e}")

# ----- 나라장터 입찰 분석 (Phase 6) -----
class TenderReviewIn(BaseModel):
    review_status: str
    review_note: Optional[str] = None

class MyBidIn(BaseModel):
    tender_id: int
    company_id: Optional[int] = None
    bid_amount: int
    note: Optional[str] = None

class CompetitorIn(BaseModel):
    name: str
    business_no: Optional[str] = None
    note: Optional[str] = None

class VehicleIn(BaseModel):
    name: str
    plate_no: Optional[str] = None
    vehicle_type: Optional[str] = None
    capacity: Optional[str] = None
    company_id: Optional[int] = None
    status: Optional[str] = "available"
    purchased_at: Optional[str] = None
    note: Optional[str] = None

class VehicleAssignIn(BaseModel):
    vehicle_id: int
    driver_id: Optional[int] = None
    site_id: Optional[int] = None
    note: Optional[str] = None

class LicenseIn(BaseModel):
    company_id: int
    license_type: str
    license_no: Optional[str] = None
    issued_at: Optional[str] = None
    expires_at: Optional[str] = None
    capacity_amount: Optional[int] = 0
    status: Optional[str] = "active"
    note: Optional[str] = None

class LicenseWorkerIn(BaseModel):
    worker_id: int
    role: Optional[str] = None
    note: Optional[str] = None

# ---- 자격증 → 면허 매트릭스 (자동 매칭) ----
# cert_name (또는 cert_name + level) → 매칭 가능한 license_type 목록
# 매핑은 보수적으로 — 실제 시행령 제13조 기준 단순화
CERT_TO_LICENSE_MAP = [
    # 토목 자격
    ("토목", None,  ["토목공사업", "토목건축공사업", "지반조성·포장공사업", "상·하수도설비공사업",
                     "철도·궤도공사업", "수중·준설공사업", "철강구조물공사업"]),
    # 건축 자격
    ("건축", None,  ["건축공사업", "토목건축공사업", "실내건축공사업", "철근·콘크리트공사업",
                     "도장·습식·방수·석공사업", "금속창호·지붕건축물조립공사업"]),
    # 기계설비
    ("기계설비", None,  ["기계가스설비공사업", "가스난방공사업"]),
    ("건축설비", None,  ["기계가스설비공사업", "가스난방공사업"]),
    # 전기·정보통신·소방
    ("전기", None,      ["전기공사업"]),
    ("정보통신", None,  ["정보통신공사업"]),
    ("소방", None,      ["소방시설공사업"]),
    # 가스
    ("가스", None,      ["가스난방공사업", "기계가스설비공사업"]),
    # 안전
    ("건설안전", None,  ["구조물해체·비계공사업"]),  # 안전기사는 거의 모든 면허에 도움
    ("산업안전", None,  ["구조물해체·비계공사업"]),
    # 콘크리트
    ("콘크리트", None,  ["철근·콘크리트공사업"]),
    # 굴삭기·중장비
    ("굴삭기", None,    ["지반조성·포장공사업", "구조물해체·비계공사업"]),
    ("굴착기", None,    ["지반조성·포장공사업", "구조물해체·비계공사업"]),
    # 비계
    ("비계", None,      ["구조물해체·비계공사업"]),
    # 방수
    ("방수", None,      ["도장·습식·방수·석공사업"]),
    # 석면
    ("석면", None,      ["석면해체·제거업"]),
    # 정밀안전점검
    ("정밀안전", None,  ["정밀안전점검(시설물)"]),
    # 시설물
    ("시설물", None,    ["시설물유지관리업"]),
]

def find_matching_licenses(cert_name, cert_level=None):
    """자격증 이름에서 가능한 면허 종류 목록 반환."""
    if not cert_name: return []
    name = str(cert_name).strip()
    matches = []
    for key, level_req, lic_list in CERT_TO_LICENSE_MAP:
        if key in name:
            if level_req is None or (cert_level and level_req == cert_level):
                matches.extend(lic_list)
    return list(set(matches))

# ---- 표준 면허 종류 카탈로그 ----
LICENSE_TYPES_CATALOG = [
    # 종합
    {"name": "토목공사업", "category": "종합", "min_workers": 6, "min_capital": 500_000_000},
    {"name": "건축공사업", "category": "종합", "min_workers": 6, "min_capital": 500_000_000},
    {"name": "토목건축공사업", "category": "종합", "min_workers": 12, "min_capital": 1_200_000_000},
    # 전문 (2022 대업종 14)
    {"name": "지반조성·포장공사업", "category": "전문", "min_workers": 2, "min_capital": 150_000_000},
    {"name": "실내건축공사업", "category": "전문", "min_workers": 2, "min_capital": 150_000_000},
    {"name": "금속창호·지붕건축물조립공사업", "category": "전문", "min_workers": 2, "min_capital": 150_000_000},
    {"name": "도장·습식·방수·석공사업", "category": "전문", "min_workers": 2, "min_capital": 150_000_000},
    {"name": "조경식재·시설물공사업", "category": "전문", "min_workers": 2, "min_capital": 150_000_000},
    {"name": "철근·콘크리트공사업", "category": "전문", "min_workers": 2, "min_capital": 150_000_000},
    {"name": "구조물해체·비계공사업", "category": "전문", "min_workers": 2, "min_capital": 150_000_000},
    {"name": "상·하수도설비공사업", "category": "전문", "min_workers": 2, "min_capital": 150_000_000},
    {"name": "철도·궤도공사업", "category": "전문", "min_workers": 2, "min_capital": 200_000_000},
    {"name": "철강구조물공사업", "category": "전문", "min_workers": 2, "min_capital": 150_000_000},
    {"name": "수중·준설공사업", "category": "전문", "min_workers": 2, "min_capital": 200_000_000},
    {"name": "승강기·삭도공사업", "category": "전문", "min_workers": 2, "min_capital": 150_000_000},
    {"name": "기계가스설비공사업", "category": "전문", "min_workers": 2, "min_capital": 150_000_000},
    {"name": "가스난방공사업", "category": "전문", "min_workers": 1, "min_capital": 100_000_000},
    # 특수·환경
    {"name": "석면해체·제거업", "category": "특수", "min_workers": 2, "min_capital": 50_000_000},
    {"name": "정밀안전점검(시설물)", "category": "특수", "min_workers": 2, "min_capital": 30_000_000},
    {"name": "시설물유지관리업", "category": "특수", "min_workers": 4, "min_capital": 200_000_000},
    # 별도법
    {"name": "전기공사업", "category": "별도법", "min_workers": 3, "min_capital": 150_000_000},
    {"name": "정보통신공사업", "category": "별도법", "min_workers": 3, "min_capital": 150_000_000},
    {"name": "소방시설공사업", "category": "별도법", "min_workers": 2, "min_capital": 100_000_000},
    {"name": "문화재수리업", "category": "별도법", "min_workers": 2, "min_capital": 200_000_000},
]

@app.get("/api/procurement/tenders")
def list_tenders(
    review_status: Optional[str] = None,
    status: Optional[str] = None,
    q: Optional[str] = None,
    days_to_deadline: Optional[int] = None,
    _: dict = Depends(require_login),
):
    sql = "SELECT * FROM tenders WHERE 1=1"
    args = []
    if review_status:
        sql += " AND review_status=?"; args.append(review_status)
    if status:
        sql += " AND status=?"; args.append(status)
    if q:
        sql += " AND (title LIKE ? OR org_name LIKE ?)"
        args += [f"%{q}%", f"%{q}%"]
    if days_to_deadline is not None:
        sql += " AND date(deadline) BETWEEN date('now') AND date('now', '+' || ? || ' days')"
        args.append(days_to_deadline)
    sql += " ORDER BY deadline ASC NULLS LAST LIMIT 200"
    # SQLite는 NULLS LAST 미지원 — workaround
    sql = sql.replace("ASC NULLS LAST", "")
    sql = sql.replace("ORDER BY deadline",
                       "ORDER BY (CASE WHEN deadline IS NULL THEN 1 ELSE 0 END), deadline")
    with conn() as c:
        return rows(c.execute(sql, args).fetchall())

@app.get("/api/procurement/tender/{tid}")
def get_tender(tid: int, _: dict = Depends(require_login)):
    with conn() as c:
        t = c.execute("SELECT * FROM tenders WHERE id=?", (tid,)).fetchone()
        if not t:
            raise HTTPException(404, "tender not found")
        my_bids = rows(c.execute(
            "SELECT mb.*, comp.name AS company_name FROM my_bids mb "
            "LEFT JOIN companies comp ON mb.company_id=comp.id WHERE mb.tender_id=?",
            (tid,)).fetchall())
        comp_bids = rows(c.execute(
            "SELECT cb.*, c.name AS competitor_name FROM competitor_bids cb "
            "JOIN competitors c ON cb.competitor_id=c.id WHERE cb.tender_id=?",
            (tid,)).fetchall())
    return {"tender": dict(t), "my_bids": my_bids, "competitor_bids": comp_bids}

@app.post("/api/procurement/tender/{tid}/review")
def review_tender(tid: int, payload: TenderReviewIn, user: dict = Depends(require_login)):
    valid = ['new', 'interested', 'bidding', 'skipped', 'won', 'lost']
    if payload.review_status not in valid:
        raise HTTPException(400, "invalid review status")
    with conn() as c:
        t = c.execute("SELECT title, review_status FROM tenders WHERE id=?", (tid,)).fetchone()
        if not t:
            raise HTTPException(404, "tender not found")
        c.execute(
            "UPDATE tenders SET review_status=?, review_note=?, reviewed_by=?, reviewed_at=datetime('now') WHERE id=?",
            (payload.review_status, payload.review_note, user["id"], tid)
        )
    emit_event("TenderReviewed",
               actors={"user_id": user["id"]},
               payload={"tender_id": tid, "title": t["title"],
                        "from": t["review_status"], "to": payload.review_status,
                        "note": payload.review_note},
               created_by=user["id"], source="admin_ui")
    return {"ok": True}

@app.post("/api/procurement/my-bids")
def add_my_bid(payload: MyBidIn, user: dict = Depends(require_login)):
    with conn() as c:
        t = c.execute("SELECT title FROM tenders WHERE id=?", (payload.tender_id,)).fetchone()
        if not t: raise HTTPException(404, "tender not found")
        cur = c.execute(
            "INSERT INTO my_bids(tender_id,company_id,bid_amount,note,submitted_by) "
            "VALUES(?,?,?,?,?)",
            (payload.tender_id, payload.company_id, payload.bid_amount, payload.note, user["id"])
        )
        c.execute("UPDATE tenders SET review_status='bidding' WHERE id=? AND review_status IN ('new','interested')",
                  (payload.tender_id,))
    emit_event("BidSubmitted",
               actors={"user_id": user["id"], "company_id": payload.company_id},
               payload={"tender_id": payload.tender_id, "title": t["title"],
                        "bid_amount": payload.bid_amount},
               financial={"amount": payload.bid_amount, "account": f"입찰응찰/{t['title']}",
                          "kind": "bid"},
               created_by=user["id"], source="admin_ui")
    return {"id": cur.lastrowid}

@app.get("/api/procurement/my-bids")
def list_my_bids(_: dict = Depends(require_login)):
    with conn() as c:
        return rows(c.execute(
            """SELECT mb.*, t.title AS tender_title, t.org_name, t.deadline,
                      t.status AS tender_status, t.award_company, t.award_amount,
                      comp.name AS company_name
               FROM my_bids mb JOIN tenders t ON mb.tender_id=t.id
               LEFT JOIN companies comp ON mb.company_id=comp.id
               ORDER BY mb.bid_at DESC"""
        ).fetchall())

@app.get("/api/procurement/competitors")
def list_competitors(_: dict = Depends(require_login)):
    with conn() as c:
        comps = rows(c.execute("SELECT * FROM competitors ORDER BY name").fetchall())
        for c_ in comps:
            cnt = c.execute(
                "SELECT COUNT(*), MAX(detected_at) FROM competitor_bids WHERE competitor_id=?",
                (c_['id'],)).fetchone()
            c_['bid_count'] = cnt[0]
            c_['last_seen'] = cnt[1]
    return comps

@app.post("/api/procurement/competitors")
def add_competitor(payload: CompetitorIn, user: dict = Depends(require_login)):
    with conn() as c:
        try:
            cur = c.execute(
                "INSERT INTO competitors(name, business_no, note, added_by) VALUES(?,?,?,?)",
                (payload.name, payload.business_no, payload.note, user["id"])
            )
        except sqlite3.IntegrityError:
            raise HTTPException(409, "이미 등록된 사업자번호입니다")
    return {"id": cur.lastrowid}

@app.delete("/api/procurement/competitors/{cid}")
def delete_competitor(cid: int, _: dict = Depends(require_login)):
    with conn() as c:
        c.execute("DELETE FROM competitor_bids WHERE competitor_id=?", (cid,))
        c.execute("DELETE FROM competitors WHERE id=?", (cid,))
    return {"ok": True}

@app.get("/api/procurement/competitors/{cid}/activity")
def competitor_activity(cid: int, _: dict = Depends(require_login)):
    with conn() as c:
        comp = c.execute("SELECT * FROM competitors WHERE id=?", (cid,)).fetchone()
        if not comp: raise HTTPException(404, "competitor not found")
        bids = rows(c.execute(
            """SELECT cb.*, t.title, t.org_name, t.budget, t.deadline, t.status AS tender_status,
                      t.award_company, t.award_amount
               FROM competitor_bids cb JOIN tenders t ON cb.tender_id=t.id
               WHERE cb.competitor_id=? ORDER BY cb.detected_at DESC""",
            (cid,)).fetchall())
    return {"competitor": dict(comp), "bids": bids}

@app.get("/api/procurement/dashboard")
def procurement_dashboard(_: dict = Depends(require_login)):
    """대표용 종합 대시보드 — 놓친 게 없는지 한눈에."""
    today = date.today().isoformat()
    with conn() as c:
        # 카운터
        new_unreviewed = c.execute(
            "SELECT COUNT(*) FROM tenders WHERE review_status='new' AND status='open'").fetchone()[0]
        deadline_3d = c.execute(
            "SELECT COUNT(*) FROM tenders WHERE status='open' AND review_status IN ('new','interested') "
            "AND date(deadline) BETWEEN date('now') AND date('now','+3 days')").fetchone()[0]
        bidding_count = c.execute(
            "SELECT COUNT(*) FROM tenders WHERE review_status='bidding'").fetchone()[0]
        won_30d = c.execute(
            "SELECT COUNT(*) FROM tenders WHERE review_status='won' AND reviewed_at >= date('now','-30 days')"
        ).fetchone()[0]
        lost_30d = c.execute(
            "SELECT COUNT(*) FROM tenders WHERE review_status='lost' AND reviewed_at >= date('now','-30 days')"
        ).fetchone()[0]
        # 마감 임박 + 미검토
        urgent = rows(c.execute(
            "SELECT id, title, org_name, budget, deadline, review_status FROM tenders "
            "WHERE status='open' AND review_status='new' "
            "AND date(deadline) BETWEEN date('now') AND date('now','+7 days') "
            "ORDER BY deadline ASC LIMIT 20").fetchall())
        # 우리가 관심 표시한 진행 중 공고
        interested = rows(c.execute(
            "SELECT id, title, org_name, budget, deadline FROM tenders "
            "WHERE review_status='interested' AND status='open' ORDER BY deadline ASC LIMIT 20"
        ).fetchall())
        # 최근 경쟁사 활동
        competitor_recent = rows(c.execute(
            """SELECT cb.*, c.name AS competitor_name, t.title AS tender_title,
                      t.org_name, t.budget
               FROM competitor_bids cb JOIN competitors c ON cb.competitor_id=c.id
               JOIN tenders t ON cb.tender_id=t.id
               WHERE cb.detected_at >= datetime('now','-30 days')
               ORDER BY cb.detected_at DESC LIMIT 30"""
        ).fetchall())
        # 낙찰률 최근 90일
        my_bids_total = c.execute(
            "SELECT COUNT(*) FROM my_bids WHERE bid_at >= date('now','-90 days')").fetchone()[0]
        my_bids_won = c.execute(
            "SELECT COUNT(*) FROM my_bids WHERE result='won' AND bid_at >= date('now','-90 days')").fetchone()[0]
    return {
        "as_of": today,
        "kpi": {
            "new_unreviewed": new_unreviewed,
            "deadline_3d": deadline_3d,
            "bidding_count": bidding_count,
            "won_30d": won_30d, "lost_30d": lost_30d,
            "win_rate_90d": round(my_bids_won / my_bids_total * 100, 1) if my_bids_total else 0,
            "my_bids_total_90d": my_bids_total,
        },
        "urgent_unreviewed": urgent,
        "interested_pipeline": interested,
        "competitor_recent": competitor_recent,
    }

# ---- 나라장터 sync (실제 OpenAPI 또는 mock) ----
@app.post("/api/procurement/sync")
def sync_tenders(_: dict = Depends(require_login)):
    """나라장터 OpenAPI 에서 새 공고 가져오기.
    환경변수 G2B_API_KEY 가 있으면 실제 API, 없으면 mock 데이터로 시연."""
    api_key = os.environ.get("G2B_API_KEY", "").strip()
    if not api_key:
        return _seed_mock_tenders()
    # TODO: 실제 나라장터 OpenAPI 호출
    # https://www.data.go.kr/data/15129394/openapi.do (입찰공고정보)
    # 키 받으면 여기에 requests.get(...) 구현
    raise HTTPException(501, "G2B_API_KEY 환경변수 + 실제 API 호출 코드 필요. "
                              "지금은 /api/procurement/sync-mock 으로 샘플 공고를 추가할 수 있습니다.")

def _seed_mock_tenders():
    """API 키 없을 때 시연용 샘플 공고."""
    samples = [
        {"tender_no":"20260425-001","title":"OO시 도로 보수 공사","org_name":"OO시청",
         "category":"도로공사","license_required":"토목공사업",
         "budget":850000000,"region":"서울","site_address":"서울 OO구",
         "deadline":(date.today() + __import__('datetime').timedelta(days=2)).isoformat()+"T17:00",
         "bid_open_at":(date.today() + __import__('datetime').timedelta(days=3)).isoformat()+"T10:00",
         "raw_url":"https://www.g2b.go.kr/"},
        {"tender_no":"20260425-002","title":"공공도서관 신축 — 골조 공사","org_name":"교육청",
         "category":"건축공사","license_required":"건축공사업, 철근콘크리트공사업",
         "budget":3200000000,"region":"경기","site_address":"수원시",
         "deadline":(date.today() + __import__('datetime').timedelta(days=5)).isoformat()+"T17:00",
         "raw_url":"https://www.g2b.go.kr/"},
        {"tender_no":"20260425-003","title":"공항 격납고 설비 공사","org_name":"한국공항공사",
         "category":"기계설비","license_required":"기계설비공사업",
         "budget":1800000000,"region":"인천","site_address":"인천공항",
         "deadline":(date.today() + __import__('datetime').timedelta(days=10)).isoformat()+"T17:00",
         "raw_url":"https://www.g2b.go.kr/"},
        {"tender_no":"20260425-004","title":"OO대학교 체육관 리모델링","org_name":"OO대학교",
         "category":"건축공사","license_required":"건축공사업",
         "budget":1500000000,"region":"서울","site_address":"서울 OO구",
         "deadline":(date.today() + __import__('datetime').timedelta(days=1)).isoformat()+"T17:00",
         "raw_url":"https://www.g2b.go.kr/"},
        {"tender_no":"20260425-005","title":"하수처리장 증설 — 토목","org_name":"환경부",
         "category":"토목공사","license_required":"토목공사업",
         "budget":5500000000,"region":"부산","site_address":"부산 OO구",
         "deadline":(date.today() + __import__('datetime').timedelta(days=14)).isoformat()+"T17:00",
         "raw_url":"https://www.g2b.go.kr/"},
        {"tender_no":"20260420-007","title":"군부대 막사 신축 — 1차","org_name":"국방부",
         "category":"건축공사","license_required":"건축공사업",
         "budget":2100000000,"region":"경기","site_address":"포천시",
         "deadline":(date.today() - __import__('datetime').timedelta(days=2)).isoformat()+"T17:00",
         "status":"closed",
         "award_company":"(주)대형건설","award_amount":2050000000,
         "raw_url":"https://www.g2b.go.kr/"},
    ]
    inserted = 0
    pending_events = []  # conn 닫힌 후 발행 — nested conn lock 방지
    with conn() as c:
        for s in samples:
            try:
                c.execute(
                    """INSERT OR IGNORE INTO tenders(tender_no,title,org_name,category,license_required,
                       budget,region,site_address,posted_at,deadline,bid_open_at,
                       status,award_company,award_amount,source,raw_url)
                       VALUES(?,?,?,?,?,?,?,?,datetime('now'),?,?,?,?,?,?,?)""",
                    (s.get("tender_no"), s.get("title"), s.get("org_name"), s.get("category"),
                     s.get("license_required"), s.get("budget"), s.get("region"), s.get("site_address"),
                     s.get("deadline"), s.get("bid_open_at"),
                     s.get("status", "open"), s.get("award_company"), s.get("award_amount"),
                     "mock", s.get("raw_url"))
                )
                if c.execute("SELECT changes()").fetchone()[0]:
                    inserted += 1
                    pending_events.append({
                        "tender_no": s.get("tender_no"), "title": s.get("title"),
                        "org_name": s.get("org_name"), "budget": s.get("budget"),
                        "deadline": s.get("deadline")
                    })
            except Exception as e:
                print(f"[mock_seed] {e}")
    # 외부 conn 닫힌 후 이벤트 발행
    for ev in pending_events:
        emit_event("TenderDiscovered", payload=ev, source="system")
    return {"ok": True, "mode": "mock", "inserted": inserted, "total_now": _count_tenders()}

def _count_tenders():
    with conn() as c:
        return c.execute("SELECT COUNT(*) FROM tenders").fetchone()[0]

# ----- 차량 (Vehicles) -----
@app.get("/api/vehicles")
def list_vehicles(_: dict = Depends(require_login)):
    with conn() as c:
        vehicles = rows(c.execute("""
            SELECT v.*, c.name AS company_name FROM vehicles v
            LEFT JOIN companies c ON v.company_id=c.id
            ORDER BY v.status, v.name
        """).fetchall())
        # 활성 배정 정보 붙이기
        for v in vehicles:
            a = c.execute("""
                SELECT va.*, w.name AS driver_name, w.phone AS driver_phone,
                       s.name AS site_name
                FROM vehicle_assignments va
                LEFT JOIN workers w ON va.driver_id=w.id
                LEFT JOIN sites s ON va.site_id=s.id
                WHERE va.vehicle_id=? AND va.returned_at IS NULL
                ORDER BY va.assigned_at DESC LIMIT 1
            """, (v['id'],)).fetchone()
            v['active_assignment'] = dict(a) if a else None
    return vehicles

@app.post("/api/vehicles")
def create_vehicle(payload: VehicleIn, user: dict = Depends(require_login)):
    with conn() as c:
        cur = c.execute(
            """INSERT INTO vehicles(name,plate_no,vehicle_type,capacity,company_id,
                                    status,purchased_at,note)
               VALUES(?,?,?,?,?,?,?,?)""",
            (payload.name, payload.plate_no, payload.vehicle_type, payload.capacity,
             payload.company_id, payload.status or "available",
             payload.purchased_at, payload.note)
        )
        vid = cur.lastrowid
    emit_event("VehicleCreated",
               actors={"vehicle_id": vid, "company_id": payload.company_id},
               payload={"name": payload.name, "plate_no": payload.plate_no,
                        "vehicle_type": payload.vehicle_type},
               created_by=user["id"], source="admin_ui")
    if payload.company_id:
        add_relation('Resource', vid, 'owned_by', 'Org', payload.company_id)
    return {"id": vid}

@app.put("/api/vehicles/{vid}")
def update_vehicle(vid: int, payload: VehicleIn, user: dict = Depends(require_login)):
    with conn() as c:
        c.execute(
            """UPDATE vehicles SET name=?,plate_no=?,vehicle_type=?,capacity=?,company_id=?,
                                   status=?,purchased_at=?,note=? WHERE id=?""",
            (payload.name, payload.plate_no, payload.vehicle_type, payload.capacity,
             payload.company_id, payload.status or "available",
             payload.purchased_at, payload.note, vid)
        )
    emit_event("VehicleUpdated", actors={"vehicle_id": vid},
               payload={"name": payload.name, "status": payload.status},
               created_by=user["id"], source="admin_ui")
    return {"ok": True}

@app.delete("/api/vehicles/{vid}")
def delete_vehicle(vid: int, user: dict = Depends(require_login)):
    with conn() as c:
        c.execute("DELETE FROM vehicles WHERE id=?", (vid,))
    emit_event("VehicleDeleted", actors={"vehicle_id": vid},
               created_by=user["id"], source="admin_ui")
    remove_relations(subject_type='Resource', subject_id=vid)
    return {"ok": True}

@app.post("/api/vehicles/{vid}/assign")
def assign_vehicle(vid: int, payload: VehicleAssignIn, user: dict = Depends(require_login)):
    """차량을 기사·현장에 배정. 기존 활성 배정이 있으면 자동 반환 처리."""
    with conn() as c:
        v = c.execute("SELECT * FROM vehicles WHERE id=?", (vid,)).fetchone()
        if not v: raise HTTPException(404, "vehicle not found")
        # 기존 활성 배정 자동 반환
        c.execute("UPDATE vehicle_assignments SET returned_at=datetime('now') "
                  "WHERE vehicle_id=? AND returned_at IS NULL", (vid,))
        cur = c.execute(
            """INSERT INTO vehicle_assignments(vehicle_id,driver_id,site_id,note)
               VALUES(?,?,?,?)""",
            (vid, payload.driver_id, payload.site_id, payload.note)
        )
        c.execute("UPDATE vehicles SET status='in_use' WHERE id=?", (vid,))
        # 이름 가져오기
        driver_name = None
        if payload.driver_id:
            r = c.execute("SELECT name FROM workers WHERE id=?", (payload.driver_id,)).fetchone()
            driver_name = r[0] if r else None
        site_name = None
        if payload.site_id:
            r = c.execute("SELECT name FROM sites WHERE id=?", (payload.site_id,)).fetchone()
            site_name = r[0] if r else None
    emit_event("VehicleAssigned",
               actors={"vehicle_id": vid, "driver_id": payload.driver_id,
                       "driver_name": driver_name},
               place={"site_id": payload.site_id, "site_name": site_name},
               payload={"vehicle_name": v['name'], "plate_no": v['plate_no']},
               created_by=user["id"], source="admin_ui")
    if payload.driver_id:
        remove_relations(subject_type='Resource', subject_id=vid, predicate='operated_by')
        add_relation('Resource', vid, 'operated_by', 'Person', payload.driver_id)
    if payload.site_id:
        remove_relations(subject_type='Resource', subject_id=vid, predicate='deployed_at')
        add_relation('Resource', vid, 'deployed_at', 'Place', payload.site_id)
    return {"id": cur.lastrowid}

@app.post("/api/vehicles/{vid}/return")
def return_vehicle(vid: int, user: dict = Depends(require_login)):
    with conn() as c:
        c.execute("UPDATE vehicle_assignments SET returned_at=datetime('now') "
                  "WHERE vehicle_id=? AND returned_at IS NULL", (vid,))
        c.execute("UPDATE vehicles SET status='available' WHERE id=?", (vid,))
    emit_event("VehicleReturned", actors={"vehicle_id": vid},
               created_by=user["id"], source="admin_ui")
    remove_relations(subject_type='Resource', subject_id=vid, predicate='operated_by')
    remove_relations(subject_type='Resource', subject_id=vid, predicate='deployed_at')
    return {"ok": True}

@app.get("/api/fleet/site/{sid}")
def fleet_at_site(sid: int, _: dict = Depends(require_login)):
    """특정 현장에 배정된 차량 목록."""
    with conn() as c:
        return rows(c.execute("""
            SELECT v.*, va.assigned_at, va.driver_id,
                   w.name AS driver_name, w.phone AS driver_phone
            FROM vehicle_assignments va
            JOIN vehicles v ON va.vehicle_id=v.id
            LEFT JOIN workers w ON va.driver_id=w.id
            WHERE va.site_id=? AND va.returned_at IS NULL
            ORDER BY v.name
        """, (sid,)).fetchall())

# ----- 면허 (Licenses) -----
@app.get("/api/licenses")
def list_licenses(company_id: Optional[int] = None, _: dict = Depends(require_login)):
    sql = "SELECT l.*, c.name AS company_name FROM licenses l LEFT JOIN companies c ON l.company_id=c.id WHERE 1=1"
    args = []
    if company_id:
        sql += " AND l.company_id=?"; args.append(company_id)
    sql += " ORDER BY l.expires_at"
    with conn() as c:
        return rows(c.execute(sql, args).fetchall())

@app.post("/api/licenses")
def create_license(payload: LicenseIn, user: dict = Depends(require_login)):
    with conn() as c:
        cur = c.execute(
            """INSERT INTO licenses(company_id,license_type,license_no,issued_at,expires_at,
                                     capacity_amount,status,note)
               VALUES(?,?,?,?,?,?,?,?)""",
            (payload.company_id, payload.license_type, payload.license_no,
             payload.issued_at, payload.expires_at, payload.capacity_amount or 0,
             payload.status or "active", payload.note)
        )
        lid = cur.lastrowid
    emit_event("LicenseAdded",
               actors={"company_id": payload.company_id},
               payload={"license_id": lid, "license_type": payload.license_type,
                        "license_no": payload.license_no, "expires_at": payload.expires_at},
               created_by=user["id"], source="admin_ui")
    return {"id": lid}

@app.put("/api/licenses/{lid}")
def update_license(lid: int, payload: LicenseIn, user: dict = Depends(require_login)):
    with conn() as c:
        c.execute(
            """UPDATE licenses SET company_id=?,license_type=?,license_no=?,issued_at=?,
                                    expires_at=?,capacity_amount=?,status=?,note=? WHERE id=?""",
            (payload.company_id, payload.license_type, payload.license_no,
             payload.issued_at, payload.expires_at, payload.capacity_amount or 0,
             payload.status or "active", payload.note, lid)
        )
    emit_event("LicenseUpdated",
               payload={"license_id": lid, "license_type": payload.license_type},
               created_by=user["id"], source="admin_ui")
    return {"ok": True}

@app.delete("/api/licenses/{lid}")
def delete_license(lid: int, user: dict = Depends(require_login)):
    with conn() as c:
        c.execute("DELETE FROM licenses WHERE id=?", (lid,))
    emit_event("LicenseDeleted", payload={"license_id": lid},
               created_by=user["id"], source="admin_ui")
    return {"ok": True}

@app.get("/api/licenses/expiring")
def list_expiring_licenses(days: int = 90, _: dict = Depends(require_login)):
    with conn() as c:
        return rows(c.execute(
            """SELECT l.*, c.name AS company_name FROM licenses l
               JOIN companies c ON l.company_id=c.id
               WHERE l.status='active' AND l.expires_at IS NOT NULL
               AND date(l.expires_at) <= date('now', '+' || ? || ' days')
               ORDER BY l.expires_at""",
            (days,)).fetchall())

@app.get("/api/license-types-catalog")
def license_types_catalog(_: dict = Depends(require_login)):
    """표준 면허 종류 카탈로그 — 태그 입력용 자동완성 목록."""
    return LICENSE_TYPES_CATALOG

# ====== 참고용 카탈로그 — 전문건설업 면허 종류 + 자격증 ======
LICENSE_REFERENCE = [
    # 종합건설업 (3종)
    {"name":"토목공사업","category":"종합","min_workers":6,"min_capital":500_000_000,
     "description":"도로·교량·터널·하수도·항만 등 종합 토목 공사",
     "required_certs":["건설기술자(토목 분야) 6명 — 초급 이상","또는 토목·건축·기계 분야 국가기술자격"],
     "examples":"도로 보수공사, 하수처리장 신설, 교량 가설"},
    {"name":"건축공사업","category":"종합","min_workers":6,"min_capital":500_000_000,
     "description":"건축물의 종합 시공",
     "required_certs":["건설기술자(건축 분야) 6명 — 초급 이상"],
     "examples":"공동주택, 사무실, 학교 신축"},
    {"name":"토목건축공사업","category":"종합","min_workers":12,"min_capital":1_200_000_000,
     "description":"토목·건축을 통합한 대형 종합 시공 (위 두 면허 통합)",
     "required_certs":["건설기술자(토목+건축) 12명"],
     "examples":"대형 단지 개발, SOC 통합 사업"},

    # 전문건설업 — 대업종 14
    {"name":"지반조성·포장공사업","category":"전문","min_workers":2,"min_capital":150_000_000,
     "description":"부지 조성, 도로 포장, 터파기·되메우기",
     "required_certs":["건설기술자(토목) 2명 — 또는 굴삭기·기계 운전기능사"],
     "examples":"부지 평탄화, 아스팔트 포장, 보도블록"},
    {"name":"실내건축공사업","category":"전문","min_workers":2,"min_capital":150_000_000,
     "description":"건물 내부 인테리어 시공",
     "required_certs":["건설기술자(건축 또는 실내건축) 2명","또는 건축응용제도기능사·실내건축기능사"],
     "examples":"사무실 인테리어, 매장 리모델링"},
    {"name":"금속창호·지붕건축물조립공사업","category":"전문","min_workers":2,"min_capital":150_000_000,
     "description":"금속창호 설치 + 지붕 시공",
     "required_certs":["건설기술자(건축) 2명 — 또는 비계기능사·금속재창호기능사"],
     "examples":"커튼월, 알루미늄 창호, 샌드위치 패널 지붕"},
    {"name":"도장·습식·방수·석공사업","category":"전문","min_workers":2,"min_capital":150_000_000,
     "description":"도장·미장·방수·석재 시공 (구 도장방수·석공사업)",
     "required_certs":["건설기술자(건축) 2명 — 또는 방수·도배·미장·타일·도장 기능사"],
     "examples":"외벽 도장, 옥상 방수, 화장실 타일"},
    {"name":"조경식재·시설물공사업","category":"전문","min_workers":2,"min_capital":150_000_000,
     "description":"조경수 식재 + 조경 시설물",
     "required_certs":["건설기술자(조경) 2명 — 또는 조경기능사"],
     "examples":"공원 조경, 가로수 식재"},
    {"name":"철근·콘크리트공사업","category":"전문","min_workers":2,"min_capital":150_000_000,
     "description":"철근 배근 + 거푸집 + 콘크리트 타설",
     "required_certs":["건설기술자(건축 또는 토목) 2명","또는 콘크리트기능사·철근기능사"],
     "examples":"건축 골조, 옹벽, 슬라브"},
    {"name":"구조물해체·비계공사업","category":"전문","min_workers":2,"min_capital":150_000_000,
     "description":"건물 해체 + 비계 가설",
     "required_certs":["건설안전·산업안전기사 1명 + 건설기술자 1명","또는 비계기능사"],
     "examples":"노후 건물 해체, 시스템 비계"},
    {"name":"상·하수도설비공사업","category":"전문","min_workers":2,"min_capital":150_000_000,
     "description":"상수·하수 배관 매설",
     "required_certs":["건설기술자(토목) 2명","또는 배관기능사"],
     "examples":"상수도 본관, 우수·오수관"},
    {"name":"철도·궤도공사업","category":"전문","min_workers":2,"min_capital":200_000_000,
     "description":"철로 부설·교체",
     "required_certs":["건설기술자(토목·철도 분야) 2명"],
     "examples":"지하철·고속철도 궤도"},
    {"name":"철강구조물공사업","category":"전문","min_workers":2,"min_capital":150_000_000,
     "description":"철골조 제작·설치",
     "required_certs":["건설기술자(건축 또는 토목) 2명","또는 용접·철골기능사"],
     "examples":"공장 철골, 창고"},
    {"name":"수중·준설공사업","category":"전문","min_workers":2,"min_capital":200_000_000,
     "description":"수중 작업, 항만 준설",
     "required_certs":["건설기술자(토목·항만) + 잠수기능사"],
     "examples":"부두 보수, 해저 케이블"},
    {"name":"승강기·삭도공사업","category":"전문","min_workers":2,"min_capital":150_000_000,
     "description":"엘리베이터·에스컬레이터·곤돌라·삭도",
     "required_certs":["건설기술자(기계 또는 전기) 2명","또는 승강기기능사"],
     "examples":"신축 건물 EV, 케이블카"},
    {"name":"기계가스설비공사업","category":"전문","min_workers":2,"min_capital":150_000_000,
     "description":"기계설비 + 가스배관",
     "required_certs":["건설기술자(기계설비) 2명","또는 배관·가스기능사"],
     "examples":"공조·환기·소방설비, 가스배관"},
    {"name":"가스난방공사업","category":"전문","min_workers":1,"min_capital":100_000_000,
     "description":"가스 보일러·난방기 설치",
     "required_certs":["가스기능사 또는 보일러산업기사"],
     "examples":"주택 보일러, 상업용 가스기기"},

    # 특수·환경
    {"name":"석면해체·제거업","category":"특수","min_workers":2,"min_capital":50_000_000,
     "description":"건물 내 석면 함유물 안전 해체·제거 (산안법 별도)",
     "required_certs":["석면해체작업감독자 + 석면해체감리원 (고용노동부)","산업안전기사·건설안전기사"],
     "examples":"천정텍스, 슬레이트 지붕 제거"},
    {"name":"정밀안전점검(시설물)","category":"특수","min_workers":2,"min_capital":30_000_000,
     "description":"시특법상 시설물 정밀안전점검·진단 (별도등록)",
     "required_certs":["건설기술자(토목·건축) 특급 2명 + 정밀안전점검 교육이수"],
     "examples":"교량·터널·댐 정밀안전진단"},
    {"name":"시설물유지관리업","category":"특수","min_workers":4,"min_capital":200_000_000,
     "description":"시설물 유지보수 종합",
     "required_certs":["건설기술자(건축·토목 또는 기계·전기) 4명"],
     "examples":"학교·공공시설 유지관리"},

    # 별도법
    {"name":"전기공사업","category":"별도법","min_workers":3,"min_capital":150_000_000,
     "description":"전기공사업법 — 발전·송변전·배전·전기설비",
     "required_certs":["전기기술자 3명 — 전기·전자 분야 국가기술자격","또는 전기 분야 건설기술자 (기능사 이상)"],
     "examples":"건축물 전기, 가로등, 변전실"},
    {"name":"정보통신공사업","category":"별도법","min_workers":3,"min_capital":150_000_000,
     "description":"정보통신공사업법 — 통신선로·구내통신·방송",
     "required_certs":["정보통신기술자 3명 — 정보통신·전자 국가기술자격","또는 정보통신 분야 건설기술자"],
     "examples":"광케이블, 구내전화, CCTV"},
    {"name":"소방시설공사업","category":"별도법","min_workers":2,"min_capital":100_000_000,
     "description":"소방시설공사업법 — 소화·경보·피난설비",
     "required_certs":["소방시설관리사·소방기술자 2명","소방설비기사(전기/기계)"],
     "examples":"스프링클러, 화재경보, 비상조명"},
    {"name":"문화재수리업","category":"별도법","min_workers":2,"min_capital":200_000_000,
     "description":"문화재수리법 — 지정 문화재 수리·보수",
     "required_certs":["문화재수리기술자 2명 (문화재청)"],
     "examples":"사찰·궁궐 보수"},
]

CERT_CATALOG = [
    # ====== 국가기술자격 (한국산업인력공단) ======
    {"name":"토목기사","category":"국가기술자격","field":"토목","grade":"기사","issuer":"한국산업인력공단",
     "applicable":["토목공사업","지반조성·포장공사업","상·하수도설비공사업","철도·궤도공사업"]},
    {"name":"토목산업기사","category":"국가기술자격","field":"토목","grade":"산업기사","issuer":"한국산업인력공단",
     "applicable":["토목공사업","지반조성·포장공사업","상·하수도설비공사업"]},
    {"name":"건축기사","category":"국가기술자격","field":"건축","grade":"기사","issuer":"한국산업인력공단",
     "applicable":["건축공사업","실내건축공사업","철근·콘크리트공사업"]},
    {"name":"건축산업기사","category":"국가기술자격","field":"건축","grade":"산업기사","issuer":"한국산업인력공단",
     "applicable":["건축공사업","실내건축공사업"]},
    {"name":"건축일반시공산업기사","category":"국가기술자격","field":"건축","grade":"산업기사","issuer":"한국산업인력공단",
     "applicable":["건축공사업","실내건축공사업","도장·습식·방수·석공사업"]},
    {"name":"건설안전기사","category":"국가기술자격","field":"안전","grade":"기사","issuer":"한국산업인력공단",
     "applicable":["구조물해체·비계공사업","석면해체·제거업"], "note":"안전관리자 선임 시 필수"},
    {"name":"산업안전기사","category":"국가기술자격","field":"안전","grade":"기사","issuer":"한국산업인력공단",
     "applicable":["구조물해체·비계공사업","석면해체·제거업"]},
    {"name":"산업안전산업기사","category":"국가기술자격","field":"안전","grade":"산업기사","issuer":"한국산업인력공단",
     "applicable":["석면해체·제거업"]},
    {"name":"콘크리트기사","category":"국가기술자격","field":"건축","grade":"기사","issuer":"한국산업인력공단",
     "applicable":["철근·콘크리트공사업"]},
    {"name":"콘크리트산업기사","category":"국가기술자격","field":"건축","grade":"산업기사","issuer":"한국산업인력공단",
     "applicable":["철근·콘크리트공사업"]},
    {"name":"전기기사","category":"국가기술자격","field":"전기","grade":"기사","issuer":"한국산업인력공단",
     "applicable":["전기공사업"]},
    {"name":"전기산업기사","category":"국가기술자격","field":"전기","grade":"산업기사","issuer":"한국산업인력공단",
     "applicable":["전기공사업"]},
    {"name":"정보통신기사","category":"국가기술자격","field":"정보통신","grade":"기사","issuer":"한국산업인력공단",
     "applicable":["정보통신공사업"]},
    {"name":"정보통신산업기사","category":"국가기술자격","field":"정보통신","grade":"산업기사","issuer":"한국산업인력공단",
     "applicable":["정보통신공사업"]},
    {"name":"소방설비기사(전기)","category":"국가기술자격","field":"소방","grade":"기사","issuer":"한국산업인력공단",
     "applicable":["소방시설공사업"]},
    {"name":"소방설비기사(기계)","category":"국가기술자격","field":"소방","grade":"기사","issuer":"한국산업인력공단",
     "applicable":["소방시설공사업"]},
    {"name":"가스기사","category":"국가기술자격","field":"가스","grade":"기사","issuer":"한국산업인력공단",
     "applicable":["가스난방공사업","기계가스설비공사업"]},
    {"name":"건축설비기사","category":"국가기술자격","field":"기계설비","grade":"기사","issuer":"한국산업인력공단",
     "applicable":["기계가스설비공사업","가스난방공사업"]},
    {"name":"공조냉동기계기사","category":"국가기술자격","field":"기계설비","grade":"기사","issuer":"한국산업인력공단",
     "applicable":["기계가스설비공사업"]},
    {"name":"조경기사","category":"국가기술자격","field":"조경","grade":"기사","issuer":"한국산업인력공단",
     "applicable":["조경식재·시설물공사업"]},

    # ====== 기능사 (한국산업인력공단) ======
    {"name":"방수기능사","category":"기능사","field":"건축","grade":"기능사","issuer":"한국산업인력공단",
     "applicable":["도장·습식·방수·석공사업"]},
    {"name":"비계기능사","category":"기능사","field":"건축","grade":"기능사","issuer":"한국산업인력공단",
     "applicable":["구조물해체·비계공사업","금속창호·지붕건축물조립공사업"]},
    {"name":"도배기능사","category":"기능사","field":"건축","grade":"기능사","issuer":"한국산업인력공단",
     "applicable":["실내건축공사업","도장·습식·방수·석공사업"]},
    {"name":"타일기능사","category":"기능사","field":"건축","grade":"기능사","issuer":"한국산업인력공단",
     "applicable":["도장·습식·방수·석공사업","실내건축공사업"]},
    {"name":"미장기능사","category":"기능사","field":"건축","grade":"기능사","issuer":"한국산업인력공단",
     "applicable":["도장·습식·방수·석공사업"]},
    {"name":"도장기능사","category":"기능사","field":"건축","grade":"기능사","issuer":"한국산업인력공단",
     "applicable":["도장·습식·방수·석공사업"]},
    {"name":"건축응용제도기능사","category":"기능사","field":"건축","grade":"기능사","issuer":"한국산업인력공단",
     "applicable":["실내건축공사업"]},
    {"name":"실내건축기능사","category":"기능사","field":"건축","grade":"기능사","issuer":"한국산업인력공단",
     "applicable":["실내건축공사업"]},
    {"name":"콘크리트기능사","category":"기능사","field":"건축","grade":"기능사","issuer":"한국산업인력공단",
     "applicable":["철근·콘크리트공사업"]},
    {"name":"전기기능사","category":"기능사","field":"전기","grade":"기능사","issuer":"한국산업인력공단",
     "applicable":["전기공사업"]},
    {"name":"배관기능사","category":"기능사","field":"기계설비","grade":"기능사","issuer":"한국산업인력공단",
     "applicable":["기계가스설비공사업","상·하수도설비공사업","가스난방공사업"]},
    {"name":"가스기능사","category":"기능사","field":"가스","grade":"기능사","issuer":"한국산업인력공단",
     "applicable":["가스난방공사업","기계가스설비공사업"]},
    {"name":"굴삭기운전기능사","category":"기능사","field":"중장비","grade":"기능사","issuer":"한국산업인력공단",
     "applicable":["지반조성·포장공사업","구조물해체·비계공사업"]},
    {"name":"기중기운전기능사","category":"기능사","field":"중장비","grade":"기능사","issuer":"한국산업인력공단",
     "applicable":["지반조성·포장공사업","철강구조물공사업"]},
    {"name":"지게차운전기능사","category":"기능사","field":"중장비","grade":"기능사","issuer":"한국산업인력공단",
     "applicable":["도장·습식·방수·석공사업"]},
    {"name":"용접기능사","category":"기능사","field":"기계","grade":"기능사","issuer":"한국산업인력공단",
     "applicable":["철강구조물공사업","기계가스설비공사업"]},

    # ====== 건설기술자 (한국건설기술인협회) — 분야별 등급 ======
    {"name":"건설기술자(토목) 초급","category":"건설기술자","field":"토목","grade":"초급","issuer":"한국건설기술인협회",
     "applicable":["토목공사업","지반조성·포장공사업","상·하수도설비공사업"], "note":"국가기술자격 또는 학력+경력으로 신고"},
    {"name":"건설기술자(토목) 중급","category":"건설기술자","field":"토목","grade":"중급","issuer":"한국건설기술인협회",
     "applicable":["토목공사업","토목건축공사업"]},
    {"name":"건설기술자(토목) 고급","category":"건설기술자","field":"토목","grade":"고급","issuer":"한국건설기술인협회",
     "applicable":["토목공사업","토목건축공사업"]},
    {"name":"건설기술자(토목) 특급","category":"건설기술자","field":"토목","grade":"특급","issuer":"한국건설기술인협회",
     "applicable":["토목공사업","토목건축공사업","정밀안전점검(시설물)"]},
    {"name":"건설기술자(건축) 초급","category":"건설기술자","field":"건축","grade":"초급","issuer":"한국건설기술인협회",
     "applicable":["건축공사업","실내건축공사업","철근·콘크리트공사업"]},
    {"name":"건설기술자(건축) 중급","category":"건설기술자","field":"건축","grade":"중급","issuer":"한국건설기술인협회",
     "applicable":["건축공사업","토목건축공사업"]},
    {"name":"건설기술자(건축) 고급","category":"건설기술자","field":"건축","grade":"고급","issuer":"한국건설기술인협회",
     "applicable":["건축공사업"]},
    {"name":"건설기술자(건축) 특급","category":"건설기술자","field":"건축","grade":"특급","issuer":"한국건설기술인협회",
     "applicable":["건축공사업","정밀안전점검(시설물)"]},
    {"name":"건설기술자(기계설비)","category":"건설기술자","field":"기계설비","grade":"분야자격","issuer":"한국건설기술인협회",
     "applicable":["기계가스설비공사업","가스난방공사업"]},
    {"name":"건설기술자(전기)","category":"건설기술자","field":"전기","grade":"분야자격","issuer":"한국건설기술인협회",
     "applicable":["전기공사업"]},
    {"name":"건설기술자(정보통신)","category":"건설기술자","field":"정보통신","grade":"분야자격","issuer":"한국건설기술인협회",
     "applicable":["정보통신공사업"]},
    {"name":"건설기술자(조경)","category":"건설기술자","field":"조경","grade":"분야자격","issuer":"한국건설기술인협회",
     "applicable":["조경식재·시설물공사업"]},
    {"name":"건설기술자(안전)","category":"건설기술자","field":"안전","grade":"분야자격","issuer":"한국건설기술인협회",
     "applicable":["구조물해체·비계공사업"]},

    # ====== 환경·특수 ======
    {"name":"석면해체작업감독자","category":"특수교육","field":"안전·환경","grade":"교육이수","issuer":"고용노동부 (KOSHA 위탁)",
     "applicable":["석면해체·제거업"], "note":"산안법상 별도 교육·자격"},
    {"name":"석면해체감리원","category":"특수교육","field":"안전·환경","grade":"교육이수","issuer":"환경부",
     "applicable":["석면해체·제거업"]},
    {"name":"산업안전관리자","category":"법정선임","field":"안전","grade":"-","issuer":"고용노동부",
     "applicable":["구조물해체·비계공사업","석면해체·제거업"], "note":"50인 이상 사업장 필수"},
    {"name":"건설안전관리자","category":"법정선임","field":"안전","grade":"-","issuer":"고용노동부",
     "applicable":["토목공사업","건축공사업"], "note":"공사규모 별 의무 선임"},
    {"name":"소방시설관리사","category":"국가전문자격","field":"소방","grade":"-","issuer":"한국소방산업기술원",
     "applicable":["소방시설공사업"]},
    {"name":"문화재수리기술자","category":"국가전문자격","field":"문화재","grade":"-","issuer":"문화재청",
     "applicable":["문화재수리업"]},
]

@app.get("/api/reference/license-types")
def reference_license_types(_: dict = Depends(require_login)):
    """전문건설업 면허 종류 풀카탈로그 (참고용)."""
    return LICENSE_REFERENCE

@app.get("/api/reference/certifications")
def reference_certifications(_: dict = Depends(require_login)):
    """건설업 자격증 카탈로그 (참고용)."""
    return CERT_CATALOG

@app.get("/api/cert-license-map")
def cert_license_map(_: dict = Depends(require_login)):
    """자격증 키워드 → 매칭 가능한 면허 매트릭스. 프론트에서 자동 매칭에 사용."""
    return [{"cert_keyword": k, "level_required": lv, "license_types": lics}
            for k, lv, lics in CERT_TO_LICENSE_MAP]

@app.get("/api/workers/{wid}/full")
def worker_full(wid: int, _: dict = Depends(require_login)):
    """한 직원의 모든 관련 정보 — 자격증·매칭면허·등재면허·법인."""
    with conn() as c:
        w = c.execute(
            "SELECT w.*, c.name AS company_name FROM workers w "
            "LEFT JOIN companies c ON w.company_id=c.id WHERE w.id=?",
            (wid,)).fetchone()
        if not w: raise HTTPException(404, "worker not found")
        certs = rows(c.execute(
            "SELECT * FROM worker_certifications WHERE worker_id=? ORDER BY cert_name",
            (wid,)).fetchall())
        # 매칭 가능 면허 (자격증 → license_types_catalog 매칭)
        matching_set = set()
        for cert in certs:
            for lic_type in find_matching_licenses(cert['cert_name'], cert['cert_level']):
                matching_set.add(lic_type)
        matching_licenses = sorted(list(matching_set))
        # 현재 등재된 면허 (license_workers 기반)
        registered = rows(c.execute(
            """SELECT lw.*, l.license_type, l.license_no, l.expires_at, l.status,
                      co.id AS company_id, co.name AS company_name
               FROM license_workers lw
               JOIN licenses l ON lw.license_id=l.id
               JOIN companies co ON l.company_id=co.id
               WHERE lw.worker_id=? ORDER BY co.name, l.license_type""",
            (wid,)).fetchall())
    return {
        "worker": dict(w),
        "certifications": certs,
        "matching_license_types": matching_licenses,
        "registered_licenses": registered,
    }

# ----- 면허 등재 직원 (license_workers) -----
@app.get("/api/licenses/{lid}/available-workers")
def list_available_workers_for_license(lid: int, _: dict = Depends(require_login)):
    """이 면허에 등재 가능한 정규직 추천 — 같은 회사 + 자격증 매칭."""
    with conn() as c:
        lic = c.execute("SELECT * FROM licenses WHERE id=?", (lid,)).fetchone()
        if not lic: raise HTTPException(404, "license not found")
        license_type = lic['license_type']
        company_id = lic['company_id']
        # 이 회사 정규직 + 미등재
        workers_in_co = rows(c.execute(
            """SELECT w.id, w.name, w.phone, w.position, w.job_role,
                      EXISTS (SELECT 1 FROM license_workers WHERE license_id=? AND worker_id=w.id) AS is_registered
               FROM workers w
               WHERE w.company_id=? AND w.worker_type='office'
                 AND (w.resigned_at IS NULL OR w.resigned_at='')
               ORDER BY w.name""", (lid, company_id)).fetchall())
        # 각 직원의 자격증
        for w in workers_in_co:
            certs = rows(c.execute(
                "SELECT cert_name, cert_level FROM worker_certifications WHERE worker_id=?",
                (w['id'],)).fetchall())
            w['certifications'] = certs
            # 자격증 → 면허 매칭
            matched_certs = []
            for cert in certs:
                lics = find_matching_licenses(cert['cert_name'], cert['cert_level'])
                if license_type in lics:
                    matched_certs.append(cert)
            w['matching_certs'] = matched_certs
            w['is_qualified'] = len(matched_certs) > 0
        # 그룹사 (다른 회사) 자격자도 참고로
        other_co_workers = rows(c.execute(
            """SELECT w.id, w.name, w.position, c.name AS company_name, c.id AS company_id
               FROM workers w JOIN companies c ON w.company_id=c.id
               WHERE c.id != ? AND w.worker_type='office'
                 AND (w.resigned_at IS NULL OR w.resigned_at='')
               ORDER BY c.name, w.name""", (company_id,)).fetchall())
        for w in other_co_workers:
            certs = rows(c.execute(
                "SELECT cert_name, cert_level FROM worker_certifications WHERE worker_id=?",
                (w['id'],)).fetchall())
            matched = []
            for cert in certs:
                lics = find_matching_licenses(cert['cert_name'], cert['cert_level'])
                if license_type in lics:
                    matched.append(cert)
            w['certifications'] = certs
            w['matching_certs'] = matched
        other_qualified = [w for w in other_co_workers if w['matching_certs']]
    return {
        "license": dict(lic),
        "workers_in_company": workers_in_co,
        "other_company_qualified": other_qualified[:20],
    }

@app.get("/api/licenses/{lid}/workers")
def list_license_workers(lid: int, _: dict = Depends(require_login)):
    with conn() as c:
        rs = rows(c.execute(
            """SELECT lw.*, w.name AS worker_name, w.phone, w.job_role, w.worker_type, w.note AS worker_note
               FROM license_workers lw JOIN workers w ON lw.worker_id=w.id
               WHERE lw.license_id=? ORDER BY w.name""",
            (lid,)).fetchall())
    return rs

@app.post("/api/licenses/{lid}/workers")
def add_license_worker(lid: int, payload: LicenseWorkerIn, user: dict = Depends(require_login)):
    with conn() as c:
        lic = c.execute("SELECT license_type, company_id FROM licenses WHERE id=?", (lid,)).fetchone()
        if not lic: raise HTTPException(404, "license not found")
        w = c.execute("SELECT name, worker_type FROM workers WHERE id=?", (payload.worker_id,)).fetchone()
        if not w: raise HTTPException(404, "worker not found")
        if w['worker_type'] != 'office':
            raise HTTPException(400, "정규직(사무직)만 면허 기술자로 등재할 수 있습니다")
        try:
            c.execute(
                "INSERT INTO license_workers(license_id,worker_id,role,note) VALUES(?,?,?,?)",
                (lid, payload.worker_id, payload.role, payload.note)
            )
        except sqlite3.IntegrityError:
            raise HTTPException(409, "이미 등재된 직원입니다")
    emit_event("LicenseWorkerRegistered",
               actors={"worker_id": payload.worker_id, "worker_name": w['name'],
                       "company_id": lic['company_id']},
               payload={"license_id": lid, "license_type": lic['license_type'],
                        "role": payload.role},
               created_by=user["id"], source="admin_ui")
    add_relation('Person', payload.worker_id, 'registered_for', 'License', lid,
                 metadata={"role": payload.role})
    return {"ok": True}

@app.delete("/api/licenses/{lid}/workers/{wid}")
def remove_license_worker(lid: int, wid: int, user: dict = Depends(require_login)):
    with conn() as c:
        lic = c.execute("SELECT license_type FROM licenses WHERE id=?", (lid,)).fetchone()
        w = c.execute("SELECT name FROM workers WHERE id=?", (wid,)).fetchone()
        c.execute("DELETE FROM license_workers WHERE license_id=? AND worker_id=?", (lid, wid))
    emit_event("LicenseWorkerRemoved",
               actors={"worker_id": wid, "worker_name": w['name'] if w else None},
               payload={"license_id": lid, "license_type": lic['license_type'] if lic else None},
               created_by=user["id"], source="admin_ui")
    remove_relations(subject_type='Person', subject_id=wid,
                     predicate='registered_for', object_type='License', object_id=lid)
    return {"ok": True}

@app.get("/api/worker-certifications")
def list_all_worker_certs(_: dict = Depends(require_login)):
    """모든 직원의 자격증 — frontend에서 worker_id로 인덱싱."""
    with conn() as c:
        return rows(c.execute(
            "SELECT * FROM worker_certifications ORDER BY worker_id, cert_name").fetchall())

@app.post("/api/workers/{wid}/certifications")
def add_worker_cert(wid: int, payload: CertificationIn, user: dict = Depends(require_login)):
    with conn() as c:
        w = c.execute("SELECT name FROM workers WHERE id=?", (wid,)).fetchone()
        if not w: raise HTTPException(404, "worker not found")
        cur = c.execute(
            """INSERT INTO worker_certifications
               (worker_id, cert_name, cert_level, cert_no, acquired_at, expires_at, related_business, note)
               VALUES(?,?,?,?,?,?,?,?)""",
            (wid, payload.cert_name, payload.cert_level, payload.cert_no,
             payload.acquired_at, payload.expires_at, payload.related_business, payload.note))
        cid = cur.lastrowid
    emit_event("CertificationAdded",
               actors={"worker_id": wid, "worker_name": w['name']},
               payload={"cert_name": payload.cert_name, "cert_level": payload.cert_level},
               created_by=user["id"], source="admin_ui")
    return {"id": cid}

@app.delete("/api/workers/{wid}/certifications/{cert_id}")
def delete_worker_cert(wid: int, cert_id: int, user: dict = Depends(require_login)):
    with conn() as c:
        w = c.execute("SELECT name FROM workers WHERE id=?", (wid,)).fetchone()
        cert = c.execute("SELECT cert_name FROM worker_certifications WHERE id=? AND worker_id=?",
                         (cert_id, wid)).fetchone()
        if not cert: raise HTTPException(404, "cert not found")
        c.execute("DELETE FROM worker_certifications WHERE id=? AND worker_id=?", (cert_id, wid))
    emit_event("CertificationRemoved",
               actors={"worker_id": wid, "worker_name": w['name'] if w else None},
               payload={"cert_name": cert['cert_name']},
               created_by=user["id"], source="admin_ui")
    return {"ok": True}

@app.post("/api/companies/{cid}/shareholders")
def add_shareholder(cid: int, payload: ShareholderIn, user: dict = Depends(require_login)):
    with conn() as c:
        co = c.execute("SELECT name FROM companies WHERE id=?", (cid,)).fetchone()
        if not co: raise HTTPException(404, "company not found")
        cur = c.execute(
            """INSERT INTO shareholders
               (company_id, name, role, rrn, address, shares_pct, contribution,
                registered_at, worker_id, note)
               VALUES(?,?,?,?,?,?,?,?,?,?)""",
            (cid, payload.name, payload.role, payload.rrn, payload.address,
             payload.shares_pct, payload.contribution, payload.registered_at,
             payload.worker_id, payload.note))
        sid = cur.lastrowid
    emit_event("ShareholderAdded",
               actors={"company_id": cid},
               payload={"name": payload.name, "role": payload.role,
                        "shares_pct": payload.shares_pct},
               created_by=user["id"], source="admin_ui")
    return {"id": sid}

@app.put("/api/shareholders/{sid}")
def update_shareholder(sid: int, payload: ShareholderIn, user: dict = Depends(require_login)):
    with conn() as c:
        existing = c.execute("SELECT * FROM shareholders WHERE id=?", (sid,)).fetchone()
        if not existing: raise HTTPException(404, "shareholder not found")
        c.execute(
            """UPDATE shareholders SET name=?, role=?, rrn=?, address=?,
               shares_pct=?, contribution=?, registered_at=?, worker_id=?, note=?
               WHERE id=?""",
            (payload.name, payload.role, payload.rrn, payload.address,
             payload.shares_pct, payload.contribution, payload.registered_at,
             payload.worker_id, payload.note, sid))
    emit_event("ShareholderUpdated",
               actors={"company_id": existing['company_id']},
               payload={"name": payload.name},
               created_by=user["id"], source="admin_ui")
    return {"ok": True}

@app.delete("/api/shareholders/{sid}")
def delete_shareholder(sid: int, user: dict = Depends(require_login)):
    with conn() as c:
        existing = c.execute("SELECT * FROM shareholders WHERE id=?", (sid,)).fetchone()
        if not existing: raise HTTPException(404, "shareholder not found")
        c.execute("DELETE FROM shareholders WHERE id=?", (sid,))
    emit_event("ShareholderDeleted",
               actors={"company_id": existing['company_id']},
               payload={"name": existing['name']},
               created_by=user["id"], source="admin_ui")
    return {"ok": True}

# ====== 면허 종류별 보기 — '철근·콘크리트공사업' 면허를 가진 모든 회사 + 직원 ======
@app.get("/api/licenses/by-type")
def licenses_by_type(_: dict = Depends(require_login)):
    """면허 종류별로 그룹화 — 각 종류마다 보유 회사 리스트(만료일·등재 직원)."""
    with conn() as c:
        all_licenses = rows(c.execute(
            """SELECT l.*, co.name AS company_name, co.business_no, co.ceo
               FROM licenses l JOIN companies co ON l.company_id=co.id
               ORDER BY l.license_type, co.name""").fetchall())
        # 등재 직원 수 + 이름들
        for l in all_licenses:
            ws = rows(c.execute(
                """SELECT lw.worker_id, w.name, w.job_role, lw.role
                   FROM license_workers lw JOIN workers w ON lw.worker_id=w.id
                   WHERE lw.license_id=? ORDER BY w.name""", (l['id'],)).fetchall())
            l['registered_workers'] = ws
            l['registered_count'] = len(ws)
    # 카탈로그에서 min_workers 가져와 충족도 계산
    catalog_by_name = {x['name']: x for x in LICENSE_TYPES_CATALOG}
    grouped = {}
    for l in all_licenses:
        t = l['license_type']
        if t not in grouped:
            cat = catalog_by_name.get(t, {})
            grouped[t] = {
                "license_type": t,
                "category": cat.get('category', '기타'),
                "min_workers": cat.get('min_workers', 2),
                "min_capital": cat.get('min_capital', 0),
                "companies": [],
            }
        # 충족 여부
        satisfied = l['registered_count'] >= grouped[t]['min_workers']
        # 만료까지 일수
        days_to_expiry = None
        if l.get('expires_at'):
            try:
                from datetime import datetime as dt
                d = dt.fromisoformat(l['expires_at'][:10])
                days_to_expiry = (d - dt.utcnow()).days
            except: pass
        grouped[t]['companies'].append({
            "license_id": l['id'],
            "company_id": l['company_id'],
            "company_name": l['company_name'],
            "business_no": l['business_no'],
            "ceo": l['ceo'],
            "license_no": l['license_no'],
            "issued_at": l['issued_at'],
            "expires_at": l['expires_at'],
            "days_to_expiry": days_to_expiry,
            "capacity_amount": l['capacity_amount'],
            "status": l['status'],
            "registered_workers": l['registered_workers'],
            "registered_count": l['registered_count'],
            "is_satisfied": satisfied,
        })
    # 카탈로그에 있지만 누구도 안 가진 면허도 노출 (선택)
    return {
        "groups": list(grouped.values()),
        "catalog": LICENSE_TYPES_CATALOG,
    }

@app.get("/api/workers/{wid}/certifications")
def get_worker_certs(wid: int, _: dict = Depends(require_login)):
    with conn() as c:
        return rows(c.execute(
            "SELECT * FROM worker_certifications WHERE worker_id=? ORDER BY cert_name",
            (wid,)).fetchall())

@app.get("/api/companies/{cid}/shareholders")
def list_company_shareholders(cid: int, _: dict = Depends(require_login)):
    with conn() as c:
        return rows(c.execute(
            "SELECT sh.*, w.name AS worker_name FROM shareholders sh "
            "LEFT JOIN workers w ON sh.worker_id=w.id "
            "WHERE sh.company_id=? ORDER BY "
            "CASE sh.role WHEN '대표' THEN 1 WHEN '대표이사' THEN 1 WHEN '이사' THEN 2 "
            "WHEN '감사' THEN 3 ELSE 4 END, sh.id",
            (cid,)).fetchall())

@app.get("/api/workers/{wid}/licenses")
def list_worker_licenses(wid: int, _: dict = Depends(require_login)):
    """이 직원이 어느 법인의 어느 면허에 기술자로 등재되어 있는지."""
    with conn() as c:
        return rows(c.execute(
            """SELECT lw.*, l.license_type, l.license_no, l.expires_at, l.status AS license_status,
                      c.name AS company_name, c.id AS company_id
               FROM license_workers lw
               JOIN licenses l ON lw.license_id=l.id
               JOIN companies c ON l.company_id=c.id
               WHERE lw.worker_id=? ORDER BY c.name, l.license_type""",
            (wid,)).fetchall())

@app.get("/api/companies/{cid}/compliance")
def company_compliance(cid: int, _: dict = Depends(require_login)):
    """법인의 면허별 충족 현황 — 등재 직원 수 vs 요구 인원."""
    catalog = {x['name']: x for x in LICENSE_TYPES_CATALOG}
    with conn() as c:
        company = c.execute("SELECT * FROM companies WHERE id=?", (cid,)).fetchone()
        if not company: raise HTTPException(404, "company not found")
        licenses = rows(c.execute("SELECT * FROM licenses WHERE company_id=? ORDER BY license_type", (cid,)).fetchall())
        regular_workers = rows(c.execute(
            "SELECT id, name, phone, job_role, hired_date FROM workers "
            "WHERE company_id=? AND worker_type='office' ORDER BY name",
            (cid,)).fetchall())
        for lic in licenses:
            workers_in = rows(c.execute(
                """SELECT lw.id AS lw_id, w.id AS worker_id, w.name, w.phone, w.job_role, lw.role
                   FROM license_workers lw JOIN workers w ON lw.worker_id=w.id
                   WHERE lw.license_id=? ORDER BY w.name""",
                (lic['id'],)).fetchall())
            lic['registered_workers'] = workers_in
            lic['registered_count'] = len(workers_in)
            req = catalog.get(lic['license_type'])
            lic['required_workers'] = req['min_workers'] if req else 2
            lic['required_capital'] = req['min_capital'] if req else 150_000_000
            lic['category'] = req['category'] if req else '미분류'
            lic['is_satisfied'] = lic['registered_count'] >= lic['required_workers']
    return {"company": dict(company), "licenses": licenses, "regular_workers": regular_workers}

# ----- Notifications + Rules Engine (Phase 5) -----
def _upsert_notification(unique_key, rule_type, severity, title, message=None, link=None,
                         related_type=None, related_id=None):
    """이미 있으면 무시. 사용자가 read 하거나 resolved 처리한 것도 그대로 유지."""
    try:
        with conn() as c:
            c.execute(
                """INSERT OR IGNORE INTO notifications
                   (unique_key,rule_type,severity,title,message,link,related_entity_type,related_entity_id)
                   VALUES(?,?,?,?,?,?,?,?)""",
                (unique_key, rule_type, severity, title, message, link, related_type, related_id))
    except Exception as e:
        print(f"[upsert_notification] {e}")

def _evaluate_rules():
    """현 상태에서 알림 생성·해소. 멱등 + idempotent."""
    today_d = date.today()
    week_ago = (today_d - __import__('datetime').timedelta(days=7)).isoformat()
    month_ago = (today_d - __import__('datetime').timedelta(days=30)).isoformat()

    try:
        with conn() as c:
            # === RULE 1: 자가가입 7일 초과 검토 대기 ===
            for w in c.execute(
                "SELECT id, name FROM workers WHERE note LIKE '%검토 대기%' AND hired_date < ?",
                (week_ago,)).fetchall():
                _upsert_notification(
                    f"pending7_{w['id']}", 'pending_approval', 'warning',
                    f"{w['name']} 님 가입 7일째 검토 대기",
                    "본사가 일당·법인을 보강하지 않으면 배치할 수 없습니다.",
                    '#/lens', 'Person', w['id'])
            # 해결: 검토 완료된 워커는 resolved
            c.execute("UPDATE notifications SET resolved=1, resolved_at=datetime('now') "
                      "WHERE rule_type='pending_approval' AND resolved=0 AND related_entity_id IN ("
                      "  SELECT id FROM workers WHERE note NOT LIKE '%검토 대기%' OR daily_wage > 0)")

            # === RULE 2: 활성 현장 GPS 미설정 ===
            for s in c.execute(
                "SELECT id, name FROM sites WHERE status='active' "
                "AND (latitude IS NULL OR longitude IS NULL OR latitude=0)").fetchall():
                _upsert_notification(
                    f"no_gps_{s['id']}", 'no_gps', 'warning',
                    f"{s['name']} 현장 GPS 좌표 미설정",
                    "출퇴근 GPS 검증이 동작하지 않습니다.",
                    '#/sites', 'Place', s['id'])
            c.execute("UPDATE notifications SET resolved=1, resolved_at=datetime('now') "
                      "WHERE rule_type='no_gps' AND resolved=0 AND related_entity_id IN ("
                      "  SELECT id FROM sites WHERE latitude IS NOT NULL AND latitude!=0)")

            # === RULE 3: 일당 미설정 일용직 ===
            for w in c.execute(
                "SELECT id, name FROM workers "
                "WHERE worker_type='daily' AND (daily_wage IS NULL OR daily_wage=0)").fetchall():
                _upsert_notification(
                    f"no_wage_{w['id']}", 'no_wage', 'warning',
                    f"{w['name']} 님 일당 미설정",
                    "노무비 자동 집계가 작동하지 않습니다.",
                    '#/workers', 'Person', w['id'])
            c.execute("UPDATE notifications SET resolved=1, resolved_at=datetime('now') "
                      "WHERE rule_type='no_wage' AND resolved=0 AND related_entity_id IN ("
                      "  SELECT id FROM workers WHERE daily_wage > 0)")

            # === RULE 4: 법인 미배정 워커 ===
            for w in c.execute(
                "SELECT id, name FROM workers WHERE company_id IS NULL").fetchall():
                _upsert_notification(
                    f"no_company_{w['id']}", 'no_company', 'info',
                    f"{w['name']} 님 소속 법인 미배정",
                    "4대보험·세무 처리에 법인 배정이 필요합니다.",
                    '#/workers', 'Person', w['id'])
            c.execute("UPDATE notifications SET resolved=1, resolved_at=datetime('now') "
                      "WHERE rule_type='no_company' AND resolved=0 AND related_entity_id IN ("
                      "  SELECT id FROM workers WHERE company_id IS NOT NULL)")

            # === RULE 5: 30일 미배치 활성 워커 ===
            for w in c.execute("""
                SELECT w.id, w.name FROM workers w
                WHERE w.daily_wage > 0
                AND NOT EXISTS (SELECT 1 FROM events e
                    WHERE e.type IN ('ClockIn','Deploy')
                    AND json_extract(e.actors,'$.worker_id') = w.id
                    AND e.occurred_at >= ?)""", (month_ago,)).fetchall():
                _upsert_notification(
                    f"idle30_{w['id']}", 'idle', 'info',
                    f"{w['name']} 님 30일간 활동 없음",
                    "휴면·퇴사 검토가 필요할 수 있습니다.",
                    '#/workers', 'Person', w['id'])

            # === RULE 6: 기성 청구 미진행 (착공 후 60일+) ===
            for s in c.execute("""
                SELECT id, name, start_date FROM sites
                WHERE status='active' AND start_date IS NOT NULL
                AND date(start_date) <= date(?, '-60 days')
                AND id NOT IN (SELECT subject_id FROM process_instances
                               WHERE workflow='progress_billing' AND subject_type='Place')""",
                (today_d.isoformat(),)).fetchall():
                _upsert_notification(
                    f"billing_overdue_{s['id']}", 'billing_overdue', 'warning',
                    f"{s['name']} 현장 — 착공 60일 넘었는데 기성 청구 0회",
                    "기성 청구 사이클을 시작해야 자금 흐름이 정상화됩니다.",
                    '#/lens', 'Place', s['id'])

            # === RULE: 면허 등재 인원 미충족 ===
            catalog_by_name = {x['name']: x for x in LICENSE_TYPES_CATALOG}
            for r in c.execute(
                """SELECT l.id, l.license_type, comp.name AS cname,
                          (SELECT COUNT(*) FROM license_workers WHERE license_id=l.id) AS reg
                   FROM licenses l JOIN companies comp ON l.company_id=comp.id
                   WHERE l.status='active'""").fetchall():
                req = catalog_by_name.get(r['license_type'])
                required = req['min_workers'] if req else 2
                if r['reg'] < required:
                    lack = required - r['reg']
                    _upsert_notification(
                        f"lic_understaff_{r['id']}", 'license_understaffed',
                        'urgent' if lack >= 2 else 'warning',
                        f"⚠️ 면허 인원 미충족: {r['cname']} {r['license_type']}",
                        f"필요 {required}명 / 등재 {r['reg']}명 — {lack}명 추가 등재 필요",
                        '#/licenses', 'License', r['id'])
            c.execute("UPDATE notifications SET resolved=1, resolved_at=datetime('now') "
                      "WHERE rule_type='license_understaffed' AND resolved=0 AND related_entity_id IN ("
                      "  SELECT l.id FROM licenses l WHERE "
                      "    (SELECT COUNT(*) FROM license_workers WHERE license_id=l.id) >= 2)")

            # === RULE: 면허 만료 임박 (90일 이내) ===
            for l in c.execute(
                "SELECT l.id, l.license_type, l.expires_at, c.name AS cname "
                "FROM licenses l JOIN companies c ON l.company_id=c.id "
                "WHERE l.status='active' AND l.expires_at IS NOT NULL "
                "AND date(l.expires_at) BETWEEN date('now') AND date('now','+90 days')").fetchall():
                _upsert_notification(
                    f"license_exp_{l['id']}", 'license_expiring',
                    'urgent' if l['expires_at'] <= (today_d + __import__('datetime').timedelta(days=30)).isoformat() else 'warning',
                    f"⏰ 면허 만료 임박: {l['cname']} {l['license_type']}",
                    f"만료일 {l['expires_at']} — 갱신 신청 필요",
                    '#/licenses', 'License', l['id'])
            c.execute("UPDATE notifications SET resolved=1, resolved_at=datetime('now') "
                      "WHERE rule_type='license_expiring' AND resolved=0 AND related_entity_id IN ("
                      "  SELECT id FROM licenses WHERE date(expires_at) > date('now','+90 days') OR status != 'active')")

            # === RULE: 차량 정비 필요 (status='maintenance' 30일 초과) ===
            # placeholder for future

            # === RULE 8~10: 나라장터 (Phase 6) ===
            # 미검토 새 공고 (열린 것만)
            for t in c.execute(
                "SELECT id, title FROM tenders WHERE review_status='new' AND status='open' "
                "AND date(deadline) >= date('now')").fetchall():
                _upsert_notification(
                    f"tender_unreviewed_{t['id']}", 'tender_unreviewed', 'info',
                    f"새 공고 미검토: {t['title']}",
                    "검토 후 입찰 여부 결정해주세요.", '#/procurement', 'Tender', t['id'])
            c.execute("UPDATE notifications SET resolved=1, resolved_at=datetime('now') "
                      "WHERE rule_type='tender_unreviewed' AND resolved=0 AND related_entity_id IN ("
                      "  SELECT id FROM tenders WHERE review_status != 'new' OR status != 'open')")

            # 마감 3일 이내 + 관심·신규 미결정
            for t in c.execute(
                "SELECT id, title, deadline FROM tenders WHERE status='open' "
                "AND review_status IN ('new','interested') "
                "AND date(deadline) BETWEEN date('now') AND date('now','+3 days')").fetchall():
                _upsert_notification(
                    f"tender_deadline_{t['id']}", 'tender_deadline', 'urgent',
                    f"⏰ 입찰 마감 임박: {t['title']}",
                    f"마감일 {t['deadline']} — 빠르게 결정 필요",
                    '#/procurement', 'Tender', t['id'])
            c.execute("UPDATE notifications SET resolved=1, resolved_at=datetime('now') "
                      "WHERE rule_type='tender_deadline' AND resolved=0 AND related_entity_id IN ("
                      "  SELECT id FROM tenders WHERE status != 'open' OR review_status NOT IN ('new','interested') "
                      "  OR date(deadline) < date('now'))")

            # 경쟁사 활동 감지 (24시간 내)
            for r in c.execute(
                """SELECT cb.competitor_id AS cid, c.name AS cname, t.title AS ttitle, t.id AS tid
                   FROM competitor_bids cb JOIN competitors c ON cb.competitor_id=c.id
                   JOIN tenders t ON cb.tender_id=t.id
                   WHERE cb.detected_at >= datetime('now','-1 day')""").fetchall():
                _upsert_notification(
                    f"comp_bid_{r['cid']}_{r['tid']}", 'competitor_activity', 'info',
                    f"👁 경쟁사 활동: {r['cname']} → {r['ttitle']}",
                    "경쟁사가 입찰에 응찰했습니다.", '#/competitors', 'Tender', r['tid'])

            # === RULE 7: 자가가입 신규 (24시간 안) — 알림용 ===
            yesterday = (today_d - __import__('datetime').timedelta(days=1)).isoformat()
            for r in c.execute(
                "SELECT json_extract(actors,'$.worker_id') AS wid, "
                "json_extract(payload,'$.name') AS name "
                "FROM events WHERE type='WorkerSelfRegistered' AND occurred_at >= ? "
                "ORDER BY occurred_at DESC LIMIT 50",
                (yesterday,)).fetchall():
                if r['wid'] and r['name']:
                    _upsert_notification(
                        f"new_signup_{r['wid']}", 'new_signup', 'info',
                        f"{r['name']} 님 신규 가입 — 정보 보강 필요",
                        "일당·법인 등을 보강하면 배치할 수 있습니다.",
                        '#/workers', 'Person', int(r['wid']))
    except Exception as e:
        print(f"[evaluate_rules] {e}")

@app.get("/api/notifications")
def list_notifications(unread_only: bool = False, _: dict = Depends(require_login)):
    sql = "SELECT * FROM notifications WHERE resolved=0"
    if unread_only:
        sql += " AND is_read=0"
    sql += " ORDER BY CASE severity WHEN 'urgent' THEN 1 WHEN 'warning' THEN 2 ELSE 3 END, created_at DESC LIMIT 200"
    with conn() as c:
        return rows(c.execute(sql).fetchall())

@app.get("/api/notifications/count")
def count_notifications(_: dict = Depends(require_login)):
    with conn() as c:
        unread = c.execute("SELECT COUNT(*) FROM notifications WHERE is_read=0 AND resolved=0").fetchone()[0]
        urgent = c.execute("SELECT COUNT(*) FROM notifications WHERE severity='urgent' AND resolved=0").fetchone()[0]
    return {"unread": unread, "urgent": urgent}

@app.post("/api/notifications/{nid}/read")
def mark_read(nid: int, _: dict = Depends(require_login)):
    with conn() as c:
        c.execute("UPDATE notifications SET is_read=1, read_at=datetime('now') WHERE id=?", (nid,))
    return {"ok": True}

@app.post("/api/notifications/read-all")
def mark_all_read(_: dict = Depends(require_login)):
    with conn() as c:
        c.execute("UPDATE notifications SET is_read=1, read_at=datetime('now') WHERE is_read=0")
    return {"ok": True}

@app.post("/api/notifications/refresh")
def refresh_rules(_: dict = Depends(require_login)):
    """룰 재평가 — 새 알림 생성 + 해결된 것 자동 닫기."""
    _evaluate_rules()
    return {"ok": True}

# ----- 일일 요약 (경영진용) -----
@app.get("/api/morning")
def morning_summary(_: dict = Depends(require_login)):
    """아침에 한 번 보면 끝나는 한 페이지."""
    today_d = date.today()
    today_str = today_d.isoformat()
    yesterday = (today_d - __import__('datetime').timedelta(days=1)).isoformat()
    week_ago = (today_d - __import__('datetime').timedelta(days=7)).isoformat()

    _evaluate_rules()  # 매번 룰 갱신

    with conn() as c:
        sites_active = c.execute("SELECT COUNT(*) FROM sites WHERE status='active'").fetchone()[0]
        workers_total = c.execute("SELECT COUNT(*) FROM workers").fetchone()[0]
        clocked_today = c.execute(
            "SELECT COUNT(DISTINCT json_extract(actors,'$.worker_id')) "
            "FROM events WHERE type='ClockIn' AND occurred_at >= ?", (today_str,)
        ).fetchone()[0]
        clocked_yesterday = c.execute(
            "SELECT COUNT(DISTINCT json_extract(actors,'$.worker_id')) "
            "FROM events WHERE type='ClockIn' AND occurred_at >= ? AND occurred_at < ?",
            (yesterday, today_str)).fetchone()[0]
        new_signups_week = c.execute(
            "SELECT COUNT(*) FROM events WHERE type='WorkerSelfRegistered' AND occurred_at >= ?",
            (week_ago,)).fetchone()[0]

        # 알림 통계
        notif_urgent = c.execute(
            "SELECT COUNT(*) FROM notifications WHERE severity='urgent' AND resolved=0").fetchone()[0]
        notif_warning = c.execute(
            "SELECT COUNT(*) FROM notifications WHERE severity='warning' AND resolved=0").fetchone()[0]
        notif_info = c.execute(
            "SELECT COUNT(*) FROM notifications WHERE severity='info' AND resolved=0").fetchone()[0]

        # 진행 중 프로세스 통계
        proc_by_state = rows(c.execute(
            "SELECT workflow, current_state, COUNT(*) AS cnt FROM process_instances "
            "WHERE completed_at IS NULL GROUP BY workflow, current_state ORDER BY workflow, current_state"
        ).fetchall())

        # 오늘의 액션 추천 — top warning notifications
        top_actions = rows(c.execute(
            "SELECT id, title, severity, link FROM notifications "
            "WHERE resolved=0 ORDER BY CASE severity WHEN 'urgent' THEN 1 WHEN 'warning' THEN 2 ELSE 3 END, created_at DESC LIMIT 10"
        ).fetchall())

        # 재무 한 줄
        contract_total = c.execute("SELECT IFNULL(SUM(contract_amount),0) FROM sites WHERE status='active'").fetchone()[0]
        paid_total = c.execute("SELECT IFNULL(SUM(paid_amount),0) FROM sites WHERE status='active'").fetchone()[0]

    return {
        "as_of": today_str,
        "kpi": {
            "active_sites": sites_active,
            "workers_total": workers_total,
            "clocked_today": clocked_today,
            "clocked_yesterday": clocked_yesterday,
            "new_signups_week": new_signups_week,
            "contract_total": contract_total,
            "paid_total": paid_total,
            "remaining": contract_total - paid_total,
        },
        "notifications": {
            "urgent": notif_urgent, "warning": notif_warning, "info": notif_info,
        },
        "processes_by_state": proc_by_state,
        "top_actions": top_actions,
    }

@app.get("/api/graph/stats")
def graph_stats(_: dict = Depends(require_login)):
    """엔티티 타입별 카운트 + 관계 총수."""
    with conn() as c:
        out = {
            "Person": c.execute("SELECT COUNT(*) FROM workers").fetchone()[0],
            "Place":  c.execute("SELECT COUNT(*) FROM sites").fetchone()[0],
            "Org":    c.execute("SELECT COUNT(*) FROM companies").fetchone()[0],
            "User":   c.execute("SELECT COUNT(*) FROM users").fetchone()[0],
            "Relations": c.execute("SELECT COUNT(*) FROM relations").fetchone()[0],
            "Events":    c.execute("SELECT COUNT(*) FROM events").fetchone()[0],
        }
    return out

# ----- Projects (현장별 일정·인력·비용·손익 종합) -----
@app.get("/api/projects")
def projects_overview(include_closed: bool = False, _: dict = Depends(require_login)):
    """현장을 큰 단위로 보기 위한 통합 데이터.
    누적 투입 인일, 누적 노무비, 일정 진행률, 예상 잔여 노무비, 예상 손익까지."""
    today_d = date.today()
    with conn() as c:
        sql = """SELECT s.*, c.name AS company_name FROM sites s
                 LEFT JOIN companies c ON s.company_id=c.id"""
        if not include_closed:
            sql += " WHERE s.status='active'"
        sql += " ORDER BY s.start_date IS NULL, s.start_date, s.id"
        sites = c.execute(sql).fetchall()
        out = []
        for s in sites:
            sd = dict(s)
            stats = c.execute(
                """SELECT COUNT(*) AS person_days,
                          IFNULL(SUM(w.daily_wage),0) AS labor_cost,
                          COUNT(DISTINCT d.worker_id) AS unique_workers,
                          MIN(d.date) AS first_day,
                          MAX(d.date) AS last_day
                   FROM deployments d JOIN workers w ON d.worker_id=w.id
                   WHERE d.site_id=? AND d.kind='actual'""", (sd['id'],)
            ).fetchone()
            sd['person_days']    = stats['person_days'] or 0
            sd['labor_cost']     = stats['labor_cost'] or 0
            sd['unique_workers'] = stats['unique_workers'] or 0
            sd['first_actual']   = stats['first_day']
            sd['last_actual']    = stats['last_day']

            # 오늘 인원 (실적)
            today_count = c.execute(
                "SELECT COUNT(DISTINCT worker_id) FROM deployments WHERE site_id=? AND date=? AND kind='actual'",
                (sd['id'], today_d.isoformat())
            ).fetchone()[0]
            sd['today_count'] = today_count

            # 일정 진행률
            try:
                start = datetime.fromisoformat(sd['start_date']).date() if sd['start_date'] else None
            except Exception: start = None
            try:
                end = datetime.fromisoformat(sd['end_date']).date() if sd['end_date'] else None
            except Exception: end = None
            total_days = (end - start).days if (start and end) else 0
            if start and today_d < start:
                elapsed = 0; sched = 'upcoming'
            elif end and today_d > end:
                elapsed = total_days; sched = 'overdue'
            elif start:
                elapsed = (today_d - start).days; sched = 'in_progress'
            else:
                elapsed = 0; sched = 'unknown'
            progress_pct = round(elapsed / total_days * 100) if total_days else 0
            remaining_days = max(0, total_days - elapsed)

            # 예상 총 노무비 (선형 외삽)
            if elapsed > 0 and total_days > 0:
                projected_labor = round(sd['labor_cost'] * total_days / elapsed)
            else:
                projected_labor = sd['labor_cost']
            projected_remaining_labor = max(0, projected_labor - sd['labor_cost'])
            estimated_profit = sd['contract_amount'] - projected_labor

            sd['days_total']       = total_days
            sd['days_elapsed']     = elapsed
            sd['days_remaining']   = remaining_days
            sd['progress_pct']     = progress_pct
            sd['schedule_status']  = sched
            sd['projected_labor']  = projected_labor
            sd['projected_remaining_labor'] = projected_remaining_labor
            sd['estimated_profit'] = estimated_profit
            sd['contract_remaining'] = sd['contract_amount'] - sd['paid_amount']
            out.append(sd)
        return out

# ----- Dashboard -----
@app.get("/api/dashboard")
def dashboard(_: dict = Depends(require_login)):
    today = date.today().isoformat()
    with conn() as c:
        sites_active = c.execute("SELECT COUNT(*) FROM sites WHERE status='active'").fetchone()[0]
        workers_total = c.execute("SELECT COUNT(*) FROM workers").fetchone()[0]
        deployed_today = c.execute(
            "SELECT COUNT(DISTINCT worker_id) FROM deployments WHERE date=? AND kind='actual'", (today,)
        ).fetchone()[0]
        planned_today = c.execute(
            "SELECT COUNT(DISTINCT worker_id) FROM deployments WHERE date=? AND kind='plan'", (today,)
        ).fetchone()[0]
        contract_total = c.execute("SELECT IFNULL(SUM(contract_amount),0) FROM sites WHERE status='active'").fetchone()[0]
        paid_total = c.execute("SELECT IFNULL(SUM(paid_amount),0) FROM sites WHERE status='active'").fetchone()[0]
        site_summary = rows(c.execute("""
            SELECT s.id, s.name, s.contract_amount, s.paid_amount,
                   (SELECT COUNT(DISTINCT worker_id) FROM deployments WHERE site_id=s.id AND date=? AND kind='actual') AS today_actual,
                   (SELECT COUNT(DISTINCT worker_id) FROM deployments WHERE site_id=s.id AND date=? AND kind='plan')   AS today_plan
            FROM sites s WHERE s.status='active' ORDER BY s.id DESC
        """, (today, today)).fetchall())
        company_breakdown = rows(c.execute("""
            SELECT c.id, c.name,
                   (SELECT COUNT(*) FROM sites WHERE company_id=c.id AND status='active') AS sites,
                   (SELECT COUNT(*) FROM workers WHERE company_id=c.id) AS workers,
                   (SELECT IFNULL(SUM(contract_amount),0) FROM sites WHERE company_id=c.id AND status='active') AS contract,
                   (SELECT IFNULL(SUM(paid_amount),0)     FROM sites WHERE company_id=c.id AND status='active') AS paid
            FROM companies c ORDER BY c.id
        """).fetchall())
    return {
        "today": today,
        "sites_active": sites_active,
        "workers_total": workers_total,
        "deployed_today": deployed_today,
        "planned_today": planned_today,
        "contract_total": contract_total,
        "paid_total": paid_total,
        "remaining": contract_total - paid_total,
        "site_summary": site_summary,
        "companies": company_breakdown,
    }

# ----- Static -----
@app.get("/")
def index():
    return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))

@app.get("/m")
def mobile_page():
    return FileResponse(os.path.join(FRONTEND_DIR, "mobile.html"))

@app.get("/register")
def register_page():
    return FileResponse(os.path.join(FRONTEND_DIR, "register.html"))

@app.get("/api/workers/excel-template")
def download_workers_template(_: dict = Depends(require_login)):
    """직원 일괄 등록용 표준 엑셀 양식. /workers 페이지에서 다운로드해서 채워서 다시 업로드.
    구조: 직원현황 시트 + 회사별 주주 시트 6개 (건우/아이엔/인우/새암/다우/유신).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
    except ImportError:
        raise HTTPException(500, "openpyxl 미설치")
    wb = openpyxl.Workbook()
    # 첫 시트 = 직원현황
    ws = wb.active
    ws.title = "직원현황"
    headers = ["회사명", "이   름", "주민번호", "입사일", "퇴사일",
               "자격증종류", "관련업종", "석면", "비  고"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2D4A8A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # 컬럼 폭
    widths = [16, 12, 16, 12, 12, 22, 16, 10, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w
    # 샘플 행 (안내·예시)
    samples = [
        ["[건우건설]", "1.안 경 이",   "650228-2226831", "11.05.20", "",
         "건축,토목 초급", "실내건축", "석면인력", "대표이사"],
        ["",            "2.김 성 헌",   "650616-1386114", "23.02.07", "",
         "토목특급,건축초급", "토목", "석면인력", ""],
        ["",            "3.박 인 식",   "570228-1387118", "17.12.01", "",
         "토목특급",     "토목", "", ""],
        ["",            "임 상 철",     "580417-1386719", "24.10.07", "25.03.31",
         "토목중급",     "", "", "퇴사 후 아이엔 입사"],
        ["[아이엔]",    "1.이 건 우",   "601116-1386715", "16.04.25", "",
         "건축.토목초급", "토공", "석면인력", "대표이사"],
        ["",            "2.김강석",     "790310-1386129", "24.10.21", "",
         "토목초급",     "포장", "", ""],
    ]
    for row in samples:
        ws.append(row)
    # 안내 행
    ws.append([])
    info_row = ws.max_row + 1
    ws.cell(info_row, 1, "📌 작성 규칙:").font = Font(bold=True, color="C0392B")
    rules = [
        "1) 회사명은 [건우건설]/[아이엔]/[인우건설]/[새암건설]/[다우건설] 처럼 대괄호로 — 그 아래 직원들이 그 회사 소속",
        "2) 이름 앞에 '1.', '2.' 가 붙으면 정규직 (현재 등록된 기술인). 번호 없으면 일반 직원/퇴사자",
        "3) 입사일/퇴사일은 'YY.MM.DD' (예: 23.07.01) 또는 'YYYY-MM-DD' 둘 다 OK",
        "4) 자격증종류는 콤마(,) 또는 점(.) 으로 여러 개. 마지막 등급(초급/중급/특급/기능사)은 앞 자격증에도 자동 적용",
        "5) 석면 컬럼에 '석면인력' 입력하면 석면 자격자 표시",
        "6) 같은 사람이 여러 회사에 입사/퇴사 반복은 그냥 회사별 블록에 따로 행 추가",
    ]
    for i, r in enumerate(rules):
        ws.cell(info_row + 1 + i, 1, r).alignment = Alignment(wrap_text=True)

    # 주주명부 시트들 — 각 회사별
    for sn in ["건우", "아이엔", "인우", "새암", "다우", "유신"]:
        sh = wb.create_sheet(sn)
        sh["B3"] = "주   주   명   부"
        sh["B3"].font = Font(bold=True, size=14)
        sh.merge_cells("B3:G3")
        sh["B3"].alignment = Alignment(horizontal="center")
        # 헤더 R7
        sh_headers = ["", "직  위", "성명", "주민등록번호", "주      소", "주식수", "비  율"]
        for ci, h in enumerate(sh_headers, 1):
            cell = sh.cell(7, ci, h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2D4A8A")
            cell.alignment = Alignment(horizontal="center")
        # 컬럼 폭
        for i, w in enumerate([3, 10, 12, 16, 30, 12, 8], 1):
            sh.column_dimensions[chr(64+i)].width = w
        # 예시 (건우만)
        if sn == "건우":
            sh.cell(8, 2, "대표이사"); sh.cell(8, 3, "안경이"); sh.cell(8, 4, "650228-2226831")
            sh.cell(8, 5, "충북영동군양강면국촌1길6-6"); sh.cell(8, 6, 41250); sh.cell(8, 7, 0.5)
            sh.cell(9, 2, "이    사"); sh.cell(9, 3, "이건우"); sh.cell(9, 4, "601116-1386715")
            sh.cell(9, 5, "충북영동군양강면국촌1길6-6"); sh.cell(9, 6, 41250); sh.cell(9, 7, 0.5)
            sh.cell(11, 2, "계"); sh.cell(11, 3, "2명"); sh.cell(11, 6, 82500); sh.cell(11, 7, 1)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import Response
    fname = "직원_주주명부_표준양식.xlsx"
    # RFC 5987 인코딩 (한글 파일명)
    import urllib.parse
    fname_enc = urllib.parse.quote(fname)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname_enc}"}
    )

# ====================================================================
# 급여·노무 모듈 (Phase 7)
# ====================================================================

# ----- 공제율 (tax_rates) -----
@app.get("/api/payroll/tax-rates")
def list_tax_rates(year: Optional[int] = None, _: dict = Depends(require_login)):
    sql = "SELECT * FROM tax_rates"
    args = []
    if year: sql += " WHERE year=?"; args.append(year)
    sql += " ORDER BY year DESC, rate_type"
    with conn() as c:
        return rows(c.execute(sql, args).fetchall())

@app.put("/api/payroll/tax-rates/{rid}")
def update_tax_rate(rid: int, payload: TaxRateIn, user: dict = Depends(require_login)):
    with conn() as c:
        c.execute("UPDATE tax_rates SET year=?, rate_type=?, rate=?, is_amount=?, note=? WHERE id=?",
                  (payload.year, payload.rate_type, payload.rate, payload.is_amount or 0, payload.note, rid))
    return {"ok": True}

@app.post("/api/payroll/tax-rates")
def create_tax_rate(payload: TaxRateIn, user: dict = Depends(require_login)):
    with conn() as c:
        cur = c.execute("INSERT OR REPLACE INTO tax_rates(year, rate_type, rate, is_amount, note) VALUES(?,?,?,?,?)",
                        (payload.year, payload.rate_type, payload.rate, payload.is_amount or 0, payload.note))
    return {"id": cur.lastrowid}

# ----- 직종별 일당 (wage_rates) -----
@app.get("/api/payroll/wage-rates")
def list_wage_rates(_: dict = Depends(require_login)):
    with conn() as c:
        return rows(c.execute("SELECT * FROM wage_rates ORDER BY job_role").fetchall())

@app.post("/api/payroll/wage-rates")
def create_wage_rate(payload: WageRateIn, user: dict = Depends(require_login)):
    with conn() as c:
        cur = c.execute(
            "INSERT OR REPLACE INTO wage_rates(job_role, daily_wage, effective_from, effective_to, note) VALUES(?,?,?,?,?)",
            (payload.job_role, payload.daily_wage, payload.effective_from, payload.effective_to, payload.note))
    return {"id": cur.lastrowid}

@app.put("/api/payroll/wage-rates/{rid}")
def update_wage_rate(rid: int, payload: WageRateIn, user: dict = Depends(require_login)):
    with conn() as c:
        c.execute("UPDATE wage_rates SET job_role=?, daily_wage=?, effective_from=?, effective_to=?, note=? WHERE id=?",
                  (payload.job_role, payload.daily_wage, payload.effective_from, payload.effective_to, payload.note, rid))
    return {"ok": True}

@app.delete("/api/payroll/wage-rates/{rid}")
def delete_wage_rate(rid: int, user: dict = Depends(require_login)):
    with conn() as c: c.execute("DELETE FROM wage_rates WHERE id=?", (rid,))
    return {"ok": True}

# ----- 협력사 (subcontractors) -----
@app.get("/api/payroll/subcontractors")
def list_subcontractors(_: dict = Depends(require_login)):
    with conn() as c:
        return rows(c.execute("SELECT * FROM subcontractors ORDER BY name").fetchall())

@app.post("/api/payroll/subcontractors")
def create_subcontractor(payload: SubcontractorIn, user: dict = Depends(require_login)):
    with conn() as c:
        cur = c.execute(
            """INSERT INTO subcontractors(name, business_no, work_type, bank_name, account_holder,
                                          account_no, leader_name, phone, address, note)
               VALUES(?,?,?,?,?,?,?,?,?,?)""",
            (payload.name, payload.business_no, payload.work_type, payload.bank_name, payload.account_holder,
             payload.account_no, payload.leader_name, payload.phone, payload.address, payload.note))
    return {"id": cur.lastrowid}

@app.put("/api/payroll/subcontractors/{sid}")
def update_subcontractor(sid: int, payload: SubcontractorIn, user: dict = Depends(require_login)):
    with conn() as c:
        c.execute(
            """UPDATE subcontractors SET name=?, business_no=?, work_type=?, bank_name=?, account_holder=?,
               account_no=?, leader_name=?, phone=?, address=?, note=? WHERE id=?""",
            (payload.name, payload.business_no, payload.work_type, payload.bank_name, payload.account_holder,
             payload.account_no, payload.leader_name, payload.phone, payload.address, payload.note, sid))
    return {"ok": True}

@app.delete("/api/payroll/subcontractors/{sid}")
def delete_subcontractor(sid: int, user: dict = Depends(require_login)):
    with conn() as c: c.execute("DELETE FROM subcontractors WHERE id=?", (sid,))
    return {"ok": True}

# ----- 특고직 (equipment_operators) -----
@app.get("/api/payroll/equipment-operators")
def list_equipment_operators(_: dict = Depends(require_login)):
    with conn() as c:
        return rows(c.execute("SELECT * FROM equipment_operators ORDER BY name").fetchall())

@app.post("/api/payroll/equipment-operators")
def create_eq_operator(payload: EquipmentOperatorIn, user: dict = Depends(require_login)):
    with conn() as c:
        cur = c.execute(
            """INSERT INTO equipment_operators(name, rrn, business_no, equipment_type, vendor_name,
                                               daily_rate, phone, bank_name, account_holder, account_no, note)
               VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
            (payload.name, payload.rrn, payload.business_no, payload.equipment_type, payload.vendor_name,
             payload.daily_rate or 0, payload.phone, payload.bank_name, payload.account_holder,
             payload.account_no, payload.note))
    return {"id": cur.lastrowid}

@app.put("/api/payroll/equipment-operators/{oid}")
def update_eq_operator(oid: int, payload: EquipmentOperatorIn, user: dict = Depends(require_login)):
    with conn() as c:
        c.execute(
            """UPDATE equipment_operators SET name=?, rrn=?, business_no=?, equipment_type=?, vendor_name=?,
               daily_rate=?, phone=?, bank_name=?, account_holder=?, account_no=?, note=? WHERE id=?""",
            (payload.name, payload.rrn, payload.business_no, payload.equipment_type, payload.vendor_name,
             payload.daily_rate or 0, payload.phone, payload.bank_name, payload.account_holder,
             payload.account_no, payload.note, oid))
    return {"ok": True}

@app.delete("/api/payroll/equipment-operators/{oid}")
def delete_eq_operator(oid: int, user: dict = Depends(require_login)):
    with conn() as c: c.execute("DELETE FROM equipment_operators WHERE id=?", (oid,))
    return {"ok": True}

# ----- 출역 그리드 — 메인 입력 -----
def _payroll_grid_data(year_month: str, company_id: Optional[int] = None):
    """순수 함수: 그리드 데이터 계산. 다른 엔드포인트(monthly 등)에서도 재사용."""
    if not re.match(r'^\d{4}-\d{2}$', year_month):
        raise HTTPException(400, "year_month 형식: YYYY-MM")
    year, month = year_month.split('-')
    from datetime import timedelta as _td
    first = date(int(year), int(month), 1)
    if int(month) == 12: last = date(int(year)+1, 1, 1) - _td(days=1)
    else: last = date(int(year), int(month)+1, 1) - _td(days=1)
    days_in_month = last.day

    with conn() as c:
        # 일용직 — w.* 로 가져와 컬럼 존재 여부 무관하게 안전
        sql = """SELECT w.*, co.name AS company_name
                 FROM workers w LEFT JOIN companies co ON w.company_id=co.id
                 WHERE w.worker_type='daily' AND (w.resigned_at IS NULL OR w.resigned_at='')"""
        args = []
        if company_id:
            sql += " AND w.company_id=?"; args.append(company_id)
        sql += " ORDER BY w.job_role, w.name"
        workers_rows = rows(c.execute(sql, args).fetchall())

        deps = c.execute(
            """SELECT worker_id, site_id, date, kind FROM deployments
               WHERE date BETWEEN ? AND ? AND kind='actual'""",
            (first.isoformat(), last.isoformat())).fetchall()
        dep_map = {}
        for d in deps:
            try: day = int(d['date'][8:10])
            except: continue
            dep_map.setdefault(d['worker_id'], set()).add(day)

        results = []
        for w in workers_rows:
            attended = sorted(list(dep_map.get(w.get('id'), set())))
            days = len(attended)
            wage = w.get('daily_wage') or 0
            exempt = bool(w.get('exempt_employment_ins') or 0)
            calc = calculate_payroll(wage, days, exempt)
            results.append({
                **w,
                "attended_days": attended,
                "days_worked": days,
                **calc,
            })

    return {
        "year_month": year_month,
        "days_in_month": days_in_month,
        "workers": results,
    }

@app.get("/api/payroll/grid")
def payroll_grid(year_month: str, company_id: Optional[int] = None,
                 _: dict = Depends(require_login)):
    return _payroll_grid_data(year_month, company_id)

class PayrollDayIn(BaseModel):
    worker_id: int
    date: str       # YYYY-MM-DD
    site_id: Optional[int] = None
    attended: bool

@app.post("/api/payroll/grid/toggle")
def payroll_toggle(payload: PayrollDayIn, user: dict = Depends(require_login)):
    """출역 그리드의 한 셀 토글 — 출근/결근 변경."""
    with conn() as c:
        if payload.attended:
            # INSERT or REPLACE — actual 배치
            c.execute("DELETE FROM deployments WHERE worker_id=? AND date=? AND kind='actual'",
                      (payload.worker_id, payload.date))
            c.execute(
                """INSERT INTO deployments(worker_id, site_id, date, kind, note)
                   VALUES(?,?,?,'actual','출역그리드')""",
                (payload.worker_id, payload.site_id or None, payload.date))
        else:
            c.execute("DELETE FROM deployments WHERE worker_id=? AND date=? AND kind='actual'",
                      (payload.worker_id, payload.date))
    return {"ok": True}

# ----- 월별 명세서 + 분류 -----
@app.get("/api/payroll/monthly")
def payroll_monthly(year_month: str, company_id: Optional[int] = None,
                    _: dict = Depends(require_login)):
    """월별 명세서 — payroll/grid 와 같은 데이터를 명세서 형식으로.
    추가로 외주 협력사·특고직 입금 내역도 포함."""
    g = _payroll_grid_data(year_month, company_id)
    workers = g['workers']
    # 분류
    subject_4ins = [w for w in workers if w['is_subject_4ins']]
    not_subject = [w for w in workers if not w['is_subject_4ins'] and w['days_worked'] > 0]
    no_attendance = [w for w in workers if w['days_worked'] == 0]

    # 합계
    totals = {
        "gross_pay": sum(w['gross_pay'] for w in workers),
        "income_tax": sum(w['income_tax'] for w in workers),
        "local_tax": sum(w['local_tax'] for w in workers),
        "employment_insurance": sum(w['employment_insurance'] for w in workers),
        "national_pension": sum(w['national_pension'] for w in workers),
        "health_insurance": sum(w['health_insurance'] for w in workers),
        "ltc_insurance": sum(w['ltc_insurance'] for w in workers),
        "retirement_fund": sum(w['retirement_fund'] for w in workers),
        "total_deductions": sum(w['total_deductions'] for w in workers),
        "net_pay": sum(w['net_pay'] for w in workers),
        "total_days": sum(w['days_worked'] for w in workers),
        "subject_4ins_count": len(subject_4ins),
        "not_subject_count": len(not_subject),
        "active_workers": len(subject_4ins) + len(not_subject),
    }
    return {
        "year_month": year_month,
        "days_in_month": g['days_in_month'],
        "workers_all": workers,
        "subject_4ins": subject_4ins,
        "not_subject": not_subject,
        "no_attendance": no_attendance,
        "totals": totals,
    }

# ----- 외주 협력사 / 특고직 입금 -----
class SubPaymentIn(BaseModel):
    subcontractor_id: int
    site_id: Optional[int] = None
    work_period: Optional[str] = None
    amount: int
    paid_at: Optional[str] = None
    invoice_no: Optional[str] = None
    note: Optional[str] = None
    period_id: Optional[int] = None

@app.get("/api/payroll/subcontractor-payments")
def list_sub_payments(year_month: Optional[str] = None, _: dict = Depends(require_login)):
    sql = """SELECT sp.*, sc.name AS subcontractor_name, sc.work_type, sc.account_no, sc.bank_name,
                    s.name AS site_name
             FROM subcontractor_payments sp
             JOIN subcontractors sc ON sp.subcontractor_id=sc.id
             LEFT JOIN sites s ON sp.site_id=s.id
             WHERE 1=1"""
    args = []
    if year_month:
        sql += " AND sp.work_period LIKE ?"
        args.append(f"%{year_month}%")
    sql += " ORDER BY sp.paid_at DESC"
    with conn() as c:
        return rows(c.execute(sql, args).fetchall())

@app.post("/api/payroll/subcontractor-payments")
def create_sub_payment(payload: SubPaymentIn, user: dict = Depends(require_login)):
    with conn() as c:
        cur = c.execute(
            """INSERT INTO subcontractor_payments(period_id, subcontractor_id, site_id, work_period,
                                                  amount, paid_at, invoice_no, note)
               VALUES(?,?,?,?,?,?,?,?)""",
            (payload.period_id, payload.subcontractor_id, payload.site_id, payload.work_period,
             payload.amount, payload.paid_at, payload.invoice_no, payload.note))
    return {"id": cur.lastrowid}

@app.delete("/api/payroll/subcontractor-payments/{pid}")
def delete_sub_payment(pid: int, _: dict = Depends(require_login)):
    with conn() as c: c.execute("DELETE FROM subcontractor_payments WHERE id=?", (pid,))
    return {"ok": True}

class EqPaymentIn(BaseModel):
    operator_id: int
    site_id: Optional[int] = None
    days_worked: Optional[int] = 0
    daily_rate: Optional[int] = 0
    amount: int
    paid_at: Optional[str] = None
    note: Optional[str] = None
    period_id: Optional[int] = None

@app.get("/api/payroll/equipment-payments")
def list_eq_payments(_: dict = Depends(require_login)):
    with conn() as c:
        return rows(c.execute(
            """SELECT ep.*, eo.name AS operator_name, eo.equipment_type, eo.vendor_name,
                      eo.business_no, s.name AS site_name
               FROM equipment_payments ep
               JOIN equipment_operators eo ON ep.operator_id=eo.id
               LEFT JOIN sites s ON ep.site_id=s.id
               ORDER BY ep.paid_at DESC""").fetchall())

@app.post("/api/payroll/equipment-payments")
def create_eq_payment(payload: EqPaymentIn, user: dict = Depends(require_login)):
    with conn() as c:
        cur = c.execute(
            """INSERT INTO equipment_payments(period_id, operator_id, site_id, days_worked, daily_rate,
                                              amount, paid_at, note)
               VALUES(?,?,?,?,?,?,?,?)""",
            (payload.period_id, payload.operator_id, payload.site_id, payload.days_worked or 0,
             payload.daily_rate or 0, payload.amount, payload.paid_at, payload.note))
    return {"id": cur.lastrowid}

@app.delete("/api/payroll/equipment-payments/{pid}")
def delete_eq_payment(pid: int, _: dict = Depends(require_login)):
    with conn() as c: c.execute("DELETE FROM equipment_payments WHERE id=?", (pid,))
    return {"ok": True}

# ----- 월별 노무비 대시보드 -----
@app.get("/api/payroll/dashboard")
def payroll_dashboard(year: Optional[int] = None, _: dict = Depends(require_login)):
    """연간 월별 노무비 추이 — 직영/외주/특고직 별도."""
    if not year: year = date.today().year
    out = []
    for m in range(1, 13):
        ym = f"{year}-{m:02d}"
        first = date(year, m, 1)
        from datetime import timedelta as _td
        if m == 12: last = date(year+1, 1, 1) - _td(days=1)
        else: last = date(year, m+1, 1) - _td(days=1)
        with conn() as c:
            # 직영 일용직 — deployments 기반 임금
            direct = c.execute("""
                SELECT IFNULL(SUM(w.daily_wage),0) AS total, COUNT(*) AS person_days,
                       COUNT(DISTINCT d.worker_id) AS unique_workers
                FROM deployments d JOIN workers w ON d.worker_id=w.id
                WHERE d.kind='actual' AND d.date BETWEEN ? AND ?
                  AND w.worker_type='daily'""", (first.isoformat(), last.isoformat())).fetchone()
            # 외주 협력사 입금
            sub = c.execute("""
                SELECT IFNULL(SUM(amount),0) AS total, COUNT(*) AS payments
                FROM subcontractor_payments WHERE work_period LIKE ?""",
                (f"%{ym}%",)).fetchone()
            # 특고직
            eq = c.execute("""
                SELECT IFNULL(SUM(amount),0) AS total, COUNT(*) AS payments
                FROM equipment_payments WHERE paid_at LIKE ?""",
                (f"{ym}%",)).fetchone()
        out.append({
            "year_month": ym,
            "direct_labor": direct['total'] or 0,
            "direct_person_days": direct['person_days'] or 0,
            "direct_unique_workers": direct['unique_workers'] or 0,
            "subcontractor_total": sub['total'] or 0,
            "subcontractor_count": sub['payments'] or 0,
            "equipment_total": eq['total'] or 0,
            "equipment_count": eq['payments'] or 0,
            "grand_total": (direct['total'] or 0) + (sub['total'] or 0) + (eq['total'] or 0),
        })
    return {"year": year, "months": out}

@app.post("/api/admin/full-reset")
def admin_full_reset(user: dict = Depends(require_login)):
    """전체 데이터 초기화 — 회사·면허·직원·자격증·주주·현장·배치·차량 모두 삭제 후
    회사 5개 (이름만) 만 재생성. 사용자 계정·이벤트 로그·알림은 보존."""
    if user.get("role") != "admin":
        raise HTTPException(403, "관리자 전용입니다")
    deleted = {}
    with conn() as c:
        # 순서 — FK 제약 안 걸리게
        for tbl in [
            "license_workers", "worker_certifications", "shareholders",
            "deployments", "clock_records", "vehicle_assignments",
            "vehicles", "licenses", "sites", "workers",
            "process_instances", "competitor_bids", "my_bids",
            "tenders", "competitors", "relations",
        ]:
            try:
                cur = c.execute(f"DELETE FROM {tbl}")
                deleted[tbl] = cur.rowcount
            except Exception as e:
                deleted[tbl] = f"err: {e}"
        # 회사도 모두 삭제 후 5개 이름만 재시드
        c.execute("DELETE FROM companies")
        seed_companies = [
            ("건우건설주식회사",),
            ("(주)아이엔건설환경",),
            ("인우건설(주)",),
            ("새암건설(주)",),
            ("다우건설(주)",),
        ]
        c.executemany("INSERT INTO companies(name) VALUES(?)", seed_companies)
        deleted["companies_reseeded"] = 5
    emit_event("FullReset", payload=deleted, created_by=user["id"], source="admin_ui")
    return {"ok": True, "deleted": deleted}

@app.post("/api/sites/wipe")
def wipe_all_sites(user: dict = Depends(require_login)):
    """모든 현장 + 배치 + 출퇴근 + 차량 배정 삭제. 회사·면허·직원은 유지."""
    if user.get("role") != "admin":
        raise HTTPException(403, "관리자 전용입니다")
    deleted = {}
    with conn() as c:
        for tbl in ["clock_records", "deployments", "vehicle_assignments"]:
            cur = c.execute(f"DELETE FROM {tbl}")
            deleted[tbl] = cur.rowcount
        cur = c.execute("DELETE FROM sites")
        deleted["sites"] = cur.rowcount
    emit_event("SitesWiped", payload=deleted, created_by=user["id"], source="admin_ui")
    return {"ok": True, "deleted": deleted}

@app.post("/api/workers/wipe")
def wipe_all_workers(user: dict = Depends(require_login)):
    """모든 직원·자격증·면허등재·주주 삭제 (회사·면허는 유지). 관리자 전용."""
    if user.get("role") != "admin":
        raise HTTPException(403, "관리자 전용입니다")
    with conn() as c:
        deleted = {}
        for tbl in ["license_workers", "worker_certifications", "shareholders"]:
            cur = c.execute(f"DELETE FROM {tbl}")
            deleted[tbl] = cur.rowcount
        cur = c.execute("DELETE FROM workers")
        deleted["workers"] = cur.rowcount
    emit_event("WorkersWiped", payload=deleted, created_by=user["id"], source="admin_ui")
    return {"ok": True, "deleted": deleted}

@app.post("/api/workers/upload-excel")
async def upload_workers_excel(
    file: UploadFile = File(...),
    replace: bool = Form(False),
    user: dict = Depends(require_login),
):
    """직원,주주명부 엑셀 업로드 → backend/initial_data 에 저장 후 임포트.
    replace=true 면 기존 직원/자격증/주주 모두 와이프 후 재구축.
    """
    if user.get("role") != "admin":
        raise HTTPException(403, "관리자 전용입니다")
    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        raise HTTPException(400, "엑셀 파일(.xlsx)만 업로드 가능합니다")
    # 저장
    save_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "initial_data")
    os.makedirs(save_dir, exist_ok=True)
    save_path = os.path.join(save_dir, "직원,주주명부-5개회사 (자동 저장됨).xlsx")
    content = await file.read()
    with open(save_path, "wb") as f:
        f.write(content)
    # 임포트
    try:
        import import_excel, importlib
        importlib.reload(import_excel)
        c = sqlite3.connect(DB_PATH)
        c.row_factory = sqlite3.Row
        c.execute("PRAGMA foreign_keys = ON")
        emp_stats = import_excel.import_employee_directory(c, Path(save_path), replace=replace)
        # 자동 면허 등재
        auto_n = import_excel.auto_register_workers_to_licenses(c)
        c.commit()
        final = {
            "workers": c.execute("SELECT COUNT(*) FROM workers").fetchone()[0],
            "office":  c.execute("SELECT COUNT(*) FROM workers WHERE worker_type='office'").fetchone()[0],
            "active":  c.execute("SELECT COUNT(*) FROM workers WHERE resigned_at IS NULL OR resigned_at=''").fetchone()[0],
            "certs":   c.execute("SELECT COUNT(*) FROM worker_certifications").fetchone()[0],
            "shareholders": c.execute("SELECT COUNT(*) FROM shareholders").fetchone()[0],
            "license_workers": c.execute("SELECT COUNT(*) FROM license_workers").fetchone()[0],
        }
        c.close()
    except Exception as e:
        import traceback
        return {"ok": False, "error": str(e), "trace": traceback.format_exc()[:2000]}
    emit_event("WorkersExcelUploaded",
               actors={"user_id": user["id"]},
               payload={"filename": file.filename, "replace": replace, "final": final, "stats": emp_stats},
               created_by=user["id"], source="admin_ui")
    return {"ok": True, "stats": emp_stats, "auto_registered": auto_n, "final": final}

class PasteDailyIn(BaseModel):
    text: str
    replace: Optional[bool] = False

@app.post("/api/payroll/paste-daily-workers")
def paste_daily_workers(payload: PasteDailyIn, user: dict = Depends(require_login)):
    """엑셀 행 복사 → paste 로 일용직 일괄 등록.
    한 줄 = 한 사람 (탭으로 구분):
      이름 / 공종 / 주민번호 / 주소 / 은행 / 예금주 / 계좌 / 일당 / 연락처 / 비고
    Excel 에서 행 선택 → Ctrl+C → 우리 textarea 에 Ctrl+V 하면 자동 탭 형식으로 들어옴.
    """
    text = (payload.text or "").strip()
    if not text:
        raise HTTPException(400, "텍스트가 비어있습니다")
    lines = [ln for ln in text.split('\n') if ln.strip()]
    added = 0
    updated = 0
    skipped_header = 0
    skipped_dup = 0
    skipped_invalid = 0
    error_lines = []

    with conn() as c:
        if payload.replace:
            c.execute("DELETE FROM workers WHERE worker_type='daily'")

        for line_no, ln in enumerate(lines, start=1):
            # 탭 또는 다중공백 구분
            cells = [x.strip() for x in re.split(r'\t', ln.rstrip('\r'))]
            # 단일 공백으로 구분된 경우 (탭이 없는 경우) — 보통 안 그렇지만
            if len(cells) < 2 and '  ' in ln:
                cells = [x.strip() for x in re.split(r' {2,}', ln) if x.strip()]
            if not cells: continue
            name_raw = cells[0]
            if not name_raw: continue

            # 헤더 라인 스킵
            if name_raw in {'이름', '성명', '이 름', '직 영', '#'}:
                skipped_header += 1
                continue
            # 한글 이름 검증 (공백 제거 후)
            name = name_raw.replace(' ', '').strip()
            if not re.match(r'^[가-힣]{2,4}$', name):
                skipped_invalid += 1
                error_lines.append(f"L{line_no}: 이름 형식 이상 → {name_raw[:30]}")
                continue

            job_role = cells[1] if len(cells) > 1 and cells[1] else None
            rrn = None
            if len(cells) > 2:
                rrn_raw = re.sub(r'\s+', '', cells[2])
                if re.match(r'^\d{6}-\d{7}$', rrn_raw):
                    rrn = rrn_raw
            address = cells[3] if len(cells) > 3 else None
            bank_name = cells[4] if len(cells) > 4 else None
            account_holder = cells[5] if len(cells) > 5 else None
            account_no = cells[6] if len(cells) > 6 else None
            daily_wage = 0
            if len(cells) > 7 and cells[7]:
                try: daily_wage = int(re.sub(r'[^\d]', '', cells[7]) or '0')
                except: daily_wage = 0
            phone = cells[8] if len(cells) > 8 else None
            note = cells[9] if len(cells) > 9 else None

            # 1952년 이전 출생자 = 고용보험 제외 (rrn 7번째 자리로 1900s/2000s 판정)
            exempt = 0; birth_date = None
            if rrn and len(rrn) >= 8:
                try:
                    yy = int(rrn[:2])
                    third = rrn[7]
                    full_year = 1900 + yy if third in '12569' else 2000 + yy
                    if full_year < 1953: exempt = 1
                    birth_date = f"{full_year}-{rrn[2:4]}-{rrn[4:6]}"
                except: pass

            # 이미 같은 RRN 또는 이름+전화 직원 있나
            existing = None
            if rrn:
                existing = c.execute("SELECT id FROM workers WHERE rrn=?", (rrn,)).fetchone()
            if not existing and phone:
                existing = c.execute(
                    "SELECT id FROM workers WHERE name=? AND phone=?", (name, phone)).fetchone()

            try:
                if existing:
                    c.execute(
                        """UPDATE workers SET worker_type='daily',
                            address=COALESCE(NULLIF(address,''),?),
                            bank_name=COALESCE(NULLIF(bank_name,''),?),
                            account_holder=COALESCE(NULLIF(account_holder,''),?),
                            bank_account=COALESCE(NULLIF(bank_account,''),?),
                            daily_wage=CASE WHEN ?>0 THEN ? ELSE daily_wage END,
                            phone=COALESCE(NULLIF(phone,''),?),
                            job_role=COALESCE(NULLIF(job_role,''),?),
                            birth_date=COALESCE(birth_date,?),
                            exempt_employment_ins=?,
                            note=COALESCE(NULLIF(note,''),?)
                           WHERE id=?""",
                        (address, bank_name, account_holder, account_no,
                         daily_wage, daily_wage, phone, job_role, birth_date,
                         exempt, note, existing[0]))
                    updated += 1
                else:
                    c.execute(
                        """INSERT INTO workers(name, rrn, worker_type, address, bank_name, account_holder,
                                                bank_account, daily_wage, phone, job_role,
                                                birth_date, exempt_employment_ins, note)
                           VALUES(?,?,'daily',?,?,?,?,?,?,?,?,?,?)""",
                        (name, rrn, address, bank_name, account_holder, account_no,
                         daily_wage, phone, job_role, birth_date, exempt, note))
                    added += 1
            except Exception as e:
                error_lines.append(f"L{line_no}: {str(e)[:100]} ({name})")

    return {
        "ok": True,
        "total_lines": len(lines),
        "added": added,
        "updated": updated,
        "skipped_header": skipped_header,
        "skipped_invalid": skipped_invalid,
        "skipped_dup": skipped_dup,
        "errors": error_lines[:20],   # 처음 20개만
    }

@app.post("/api/payroll/upload-daily-workers")
async def upload_daily_workers_excel(
    file: UploadFile = File(...),
    replace: bool = Form(False),
    user: dict = Depends(require_login),
):
    """일용직 + 특고직 + 협력사 엑셀 업로드 (2025년10월-일용-총괄.xlsx 형식)."""
    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        raise HTTPException(400, "엑셀 파일만")
    save_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "initial_data")
    os.makedirs(save_dir, exist_ok=True)
    save_path = os.path.join(save_dir, "2025년10월-일용-총괄.xlsx")
    content = await file.read()
    with open(save_path, "wb") as f: f.write(content)
    try:
        import import_excel, importlib
        importlib.reload(import_excel)
        c = sqlite3.connect(DB_PATH, timeout=30.0)
        c.row_factory = sqlite3.Row
        stats = import_excel.import_daily_workers(c, Path(save_path), replace=replace)
        c.commit(); c.close()
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[upload-daily-workers ERR] {e}\n{tb}")
        return {"ok": False, "error": str(e), "trace": tb[:2500]}
    return {"ok": True, "stats": stats}

@app.post("/api/companies/upload-outline")
async def upload_company_outline_excel(
    file: UploadFile = File(...),
    user: dict = Depends(require_login),
):
    """회사별 아웃라인 엑셀 업로드 → 회사 정보 + 기술자 보강."""
    if user.get("role") != "admin":
        raise HTTPException(403, "관리자 전용입니다")
    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        raise HTTPException(400, "엑셀 파일만 업로드 가능합니다")
    save_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "initial_data")
    os.makedirs(save_dir, exist_ok=True)
    save_path = os.path.join(save_dir, "회사별아웃라인-(기술인력보유현황포함).xlsx")
    content = await file.read()
    with open(save_path, "wb") as f: f.write(content)
    try:
        import import_excel, importlib
        importlib.reload(import_excel)
        c = sqlite3.connect(DB_PATH)
        c.row_factory = sqlite3.Row
        stats = import_excel.import_company_outline(c, Path(save_path))
        c.commit(); c.close()
    except Exception as e:
        import traceback
        return {"ok": False, "error": str(e), "trace": traceback.format_exc()[:2000]}
    return {"ok": True, "stats": stats}

@app.post("/api/admin/import-excel")
def admin_import_excel(replace: bool = False, user: dict = Depends(require_login)):
    """엑셀 4종 일괄 임포트 (admin 전용). backend/initial_data/ 의 파일 사용.
    replace=true 이면 기존 직원/자격증/주주/면허등재 모두 와이프 후 재구축.
    """
    if user.get("role") != "admin":
        raise HTTPException(403, "관리자 전용입니다")
    try:
        import import_excel
        # 모듈 캐시 갱신 (코드 업데이트 반영)
        import importlib
        importlib.reload(import_excel)
        report = import_excel.import_all(db_path=DB_PATH, replace=replace)
    except Exception as e:
        import traceback
        return {"ok": False, "error": str(e), "trace": traceback.format_exc()[:2000]}
    emit_event("ExcelImportRun",
               actors={"user_id": user["id"]},
               payload={"replace": replace,
                        "final": report.get("final", {}),
                        "employees": report.get("employees", {}),
                        "outline": report.get("outline", {}),
                        "payroll": report.get("payroll", {}),
                        "auto_register_total": report.get("auto_register_total", 0)},
               created_by=user["id"], source="admin_ui")
    # 임포트 후 룰 재평가
    _evaluate_rules()
    return {"ok": True, "report": report}

@app.get("/login")
def login_page():
    return FileResponse(os.path.join(FRONTEND_DIR, "login.html"))

@app.get("/signup")
def signup_page():
    return FileResponse(os.path.join(FRONTEND_DIR, "signup.html"))

app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")

if __name__ == "__main__":
    import uvicorn
    init_db()
    port = int(os.environ.get("PORT", 8765))
    uvicorn.run(app, host="0.0.0.0", port=port)
