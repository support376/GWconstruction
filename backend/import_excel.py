"""
엑셀 4종 → 시스템 일괄 임포트
- 직원,주주명부-5개회사.xlsx → workers + shareholders + worker_certifications
- 회사별아웃라인-(기술인력보유현황포함).xlsx → companies 보강 + 기술자 보강
- 건우건설급여대장.xlsx → 직원명부 보강 (계좌·예금주·전화)
- 2025년10월-일용-총괄.xlsx → (다음 라운드)

멱등성: 이미 있는 행은 갱신, 없으면 추가. 한 번 더 돌려도 안전.
"""
import os, re, sqlite3, json
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

DATA_DIR = Path(__file__).parent / "initial_data"

# ====== 회사 시트명 → 표준 법인명 (직원,주주명부.xlsx) ======
SHEET_TO_COMPANY = {
    "건우": "건우건설주식회사",
    "아이엔": "(주)아이엔건설환경",
    "인우": "인우건설(주)",
    "새암": "새암건설(주)",
    "다우": "다우건설(주)",
    "유신": "유신건설(주)",
    "건우이전분": "건우건설주식회사",
    "아이엔이전분": "(주)아이엔건설환경",
}

# 주주시트 사이드의 면허 종목 약식명 → 표준 카탈로그명
LICENSE_SHORT_TO_FULL = {
    "토목공사업": "토목공사업",
    "건축공사업": "건축공사업",
    "토목건축공사업": "토목건축공사업",
    "시설물(토목)": "토목공사업",
    "시설물": "토목공사업",
    "지반조성": "지반조성·포장공사업",
    "지반": "지반조성·포장공사업",
    "지반공사업": "지반조성·포장공사업",
    "지반포장": "지반조성·포장공사업",
    "포장": "지반조성·포장공사업",
    "철콘": "철근·콘크리트공사업",
    "철근콘크리트": "철근·콘크리트공사업",
    "콘크리트": "철근·콘크리트공사업",
    "비계": "구조물해체·비계공사업",
    "비계구조물": "구조물해체·비계공사업",
    "구조물": "구조물해체·비계공사업",
    "구조물해체": "구조물해체·비계공사업",
    "해체비계": "구조물해체·비계공사업",
    "실내건축": "실내건축공사업",
    "도장습식": "도장·습식·방수·석공사업",
    "도장습식석공": "도장·습식·방수·석공사업",
    "도장방수": "도장·습식·방수·석공사업",
    "방수": "도장·습식·방수·석공사업",
    "도장": "도장·습식·방수·석공사업",
    "금속창호지붕": "금속창호·지붕건축물조립공사업",
    "금속창호": "금속창호·지붕건축물조립공사업",
    "상하수도": "상·하수도설비공사업",
    "상하": "상·하수도설비공사업",
    "기계가스": "기계가스설비공사업",
    "기계설비": "기계가스설비공사업",
    "가스난방": "가스난방공사업",
    "전기": "전기공사업",
    "전기공사": "전기공사업",
    "정보통신": "정보통신공사업",
    "소방": "소방시설공사업",
    "소방시설": "소방시설공사업",
    "조경": "조경식재·시설물공사업",
    "조경식재": "조경식재·시설물공사업",
    "철도": "철도·궤도공사업",
    "철도궤도": "철도·궤도공사업",
    "수중준설": "수중·준설공사업",
    "철강구조물": "철강구조물공사업",
    "승강기": "승강기·삭도공사업",
    "정밀안전점검기관": "정밀안전점검(시설물)",
    "정밀안전점검": "정밀안전점검(시설물)",
    "정밀안전": "정밀안전점검(시설물)",
    "석면해체제거": "석면해체·제거업",
    "석면해체": "석면해체·제거업",
    "석면": "석면해체·제거업",
    "시설물유지관리": "시설물유지관리업",
    "문화재수리": "문화재수리업",
    "문화재": "문화재수리업",
}

def map_license_short(short_name):
    if not short_name: return None
    s = str(short_name).strip().replace(' ', '')
    # 정확 매칭 우선
    if s in LICENSE_SHORT_TO_FULL: return LICENSE_SHORT_TO_FULL[s]
    # 부분 매칭
    for k, v in LICENSE_SHORT_TO_FULL.items():
        if k in s: return v
    return None

# ====== 회사명 → DB 매칭 (시드와 일치하는 표준명) ======
COMPANY_NAME_MAP = {
    "건우건설": "건우건설주식회사",
    "건우건설(주)": "건우건설주식회사",
    "건우건설㈜": "건우건설주식회사",
    "건우건설주식회사": "건우건설주식회사",
    "아이엔": "(주)아이엔건설환경",
    "(주)아이엔": "(주)아이엔건설환경",
    "(주)아이엔건설환경": "(주)아이엔건설환경",
    "아이엔건설환경": "(주)아이엔건설환경",
    "인우": "인우건설(주)",
    "인우건설": "인우건설(주)",
    "인우건설(주)": "인우건설(주)",
    "인우건설㈜": "인우건설(주)",
    "새암": "새암건설(주)",
    "새암건설": "새암건설(주)",
    "새암건설(주)": "새암건설(주)",
    "새암건설㈜": "새암건설(주)",
    "다우": "다우건설(주)",
    "다우건설": "다우건설(주)",
    "다우건설(주)": "다우건설(주)",
    "유신건설": "유신건설(주)",  # 시드에 없음 → 자동 추가됨
    "유신건설(주)": "유신건설(주)",
    "유신건설㈜": "유신건설(주)",
    "유신": "유신건설(주)",
}

# ====== 자격증명 정규화 ======
def parse_certifications(cert_str):
    """
    '건축,토목 초급'  → [{name:'건축', level:'초급'}, {name:'토목', level:'초급'}]
    '토목특급,건축초급' → [{name:'토목', level:'특급'}, {name:'건축', level:'초급'}]
    '굴삭기' → [{name:'굴삭기'}]
    """
    if not cert_str: return []
    s = str(cert_str).strip()
    if not s: return []
    levels = ['특급', '고급', '중급', '초급', '기능사', '기사', '기술사', '산업기사']
    out = []
    # 콤마 분리
    parts = re.split(r'[,/、·]+', s)
    last_level = None
    # 두 번 패스: 마지막에 명시된 레벨이 앞에도 적용되는 경우 ('건축,토목 초급')
    parsed_parts = []
    for p in reversed(parts):
        p = p.strip()
        if not p: continue
        level = None
        name = p
        for lv in levels:
            if lv in p:
                level = lv
                name = p.replace(lv, '').strip()
                break
        if level is None and last_level:
            level = last_level
        if level: last_level = level
        parsed_parts.append({"name": name or p, "level": level})
    return list(reversed(parsed_parts))

def normalize_date(d):
    """엑셀 날짜를 ISO로. '23.07.01', datetime, '24.10.31' 등."""
    if d is None: return None
    if isinstance(d, datetime):
        return d.date().isoformat()
    if isinstance(d, date):
        return d.isoformat()
    s = str(d).strip()
    if not s: return None
    # 첫 줄만 (multi-line인 경우)
    s = s.split('\n')[0].strip()
    # YY.MM.DD → 20YY-MM-DD
    m = re.match(r'^(\d{2})\.(\d{1,2})\.(\d{1,2})$', s)
    if m:
        yy, mm, dd = m.groups()
        year = 2000 + int(yy) if int(yy) < 50 else 1900 + int(yy)
        return f"{year:04d}-{int(mm):02d}-{int(dd):02d}"
    # YYYY-MM-DD or YYYY/MM/DD
    m = re.match(r'^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})', s)
    if m:
        y, mm, dd = m.groups()
        return f"{int(y):04d}-{int(mm):02d}-{int(dd):02d}"
    return None

def pick_last_date(d):
    """multi-line 날짜에서 가장 마지막(최신) 날짜를 ISO로.
    '23.08.01\\n24.06.16' → '2024-06-16'
    '24.02.29 08.17 24' → '2024-08-17' (공백 분리도 처리)"""
    if d is None: return None
    if isinstance(d, (datetime, date)):
        return normalize_date(d)
    s = str(d).strip()
    if not s: return None
    # 줄/공백/콤마 모두 구분자
    parts = re.split(r'[\n\r,]+', s)
    candidates = []
    for part in parts:
        # 한 줄 안에 여러 'YY.MM.DD' 가 공백으로 구분된 경우도
        for tok in re.findall(r'\d{2,4}[\.\-/]\d{1,2}[\.\-/]\d{1,2}', part):
            iso = normalize_date(tok)
            if iso: candidates.append(iso)
    if not candidates:
        # 마지막 한 줄에 정규화 시도
        return normalize_date(parts[-1].strip())
    return max(candidates)  # ISO 문자열 비교 = 날짜순 정렬

def clean_name(n):
    if not n: return None
    s = str(n).strip()
    s = re.sub(r'^\d+\.\s*', '', s)  # "1.안 경 이" → "안 경 이"
    s = re.sub(r'\s+', '', s)         # 공백 제거 → "안경이"
    return s or None

def get_or_create_company(c, name, strict=True):
    """COMPANY_NAME_MAP 으로 표준화 후 companies 테이블에서 찾기/생성.
    strict=True: 매핑에 없는 이름이면 None 반환 (잘못된 회사 자동 생성 방지)."""
    if not name: return None
    raw = str(name).strip()
    # bracketed: [건우건설]
    m = re.search(r'\[(.+?)\]', raw)
    if m:
        key = m.group(1).strip()
    else:
        key = raw.split('\n')[0].strip()
    # 정규화 후 매핑
    standard = COMPANY_NAME_MAP.get(key)
    if not standard:
        # 매핑에 정확히 없으면 키워드 기반 휴리스틱 (포함 검사)
        for k, v in COMPANY_NAME_MAP.items():
            if k in key or key in k:
                standard = v
                break
    if not standard:
        if strict: return None
        standard = key
    row = c.execute("SELECT id FROM companies WHERE name=?", (standard,)).fetchone()
    if row: return row[0]
    cur = c.execute("INSERT INTO companies(name) VALUES(?)", (standard,))
    return cur.lastrowid

# 주주명부에서 제외할 라벨·헤더
SHAREHOLDER_EXCLUDE = {
    '직위', '이사', '감사', '주주', '대표', '대표이사', '사내이사', '사외이사',
    '관계', '성명', '주민번호', '주소', '주식수', '지분율', '보유종목', '비고',
    '소계', '합계', '계', '회사', '회사별', '구분', '종목', '직책',
    '철콘', '상하수도', '비계', '구조물', '실내건축', '도장습식', '도장방수',
    '석면해체', '지반조성', '토목', '건축', '전기', '정보통신', '소방',
    '주주명부', '주식의총수', '주식의', '총수', '회 사 별', '지위',
}

# ====== 1) 직원,주주명부 임포트 ======
def wipe_employee_data(c):
    """직원/자격증/면허등재/주주를 모두 삭제.
    회사·면허·일용직 출퇴근/배치 등은 유지. (worker_type='daily' 도 같이 와이프 — 직원 전체 재구축)
    """
    deleted = {}
    for tbl in ["license_workers", "worker_certifications", "shareholders"]:
        cur = c.execute(f"DELETE FROM {tbl}")
        deleted[tbl] = cur.rowcount
    # workers 는 ON DELETE CASCADE 로 deployments/attendances 등도 삭제됨
    cur = c.execute("DELETE FROM workers")
    deleted["workers"] = cur.rowcount
    return deleted

def _split_dates(s):
    """'23.08.01/24.06.16' 또는 '24.02.29 08.17 24.10.14' 같은 멀티-날짜를 ISO 리스트로."""
    if s is None: return []
    text = str(s).strip()
    if not text: return []
    iso_dates = []
    # YY.MM.DD 또는 YYYY-MM-DD 모두 정규식으로 추출
    for tok in re.findall(r'\d{2,4}[\.\-/]\d{1,2}[\.\-/]\d{1,2}', text):
        iso = normalize_date(tok)
        if iso: iso_dates.append(iso)
    return iso_dates

def import_employee_directory(c, path, replace=False):
    """직원,주주명부 엑셀 → workers + worker_certifications + shareholders + companies 메타 + licenses.

    이 함수는 9개 시트(직원현황 + 5 회사 + 2 이전분 + 유신) 의 모든 정보를 처리한다:
      1) 직원현황: 회사 블록 단위로 직원·자격증·재입사·퇴사·메모·계좌·명절보너스 모두 임포트
      2) 회사 시트: 주주 + 회사 메타(사업자번호·상호·주소·전화) + 사이드 면허 카탈로그(등록번호·자본금)
      3) 이전분 시트: 과거 주주 정보 (note 에 보존) — 주주 등록은 현재 시트가 우선
    """
    if not path.exists():
        print(f"  [skip] {path.name} 없음")
        return {"workers_added": 0, "workers_updated": 0, "certs_added": 0,
                "shareholders": 0, "wiped": None}
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    stats = {"workers_added": 0, "workers_updated": 0, "certs_added": 0,
             "shareholders": 0, "company_meta_updated": 0, "licenses_added": 0,
             "licenses_updated": 0, "wiped": None}
    if replace:
        stats["wiped"] = wipe_employee_data(c)
        # 면허도 함께 와이프 (사이드 면허 카탈로그 새로 채워넣기 위해)
        cur = c.execute("DELETE FROM licenses")
        stats["wiped"]["licenses"] = cur.rowcount
        print(f"  🗑  와이프: {stats['wiped']}")

    # ===== 직원현황 시트 — 모든 회사 직원 통합 (정밀 모드) =====
    if "직원현황" in wb.sheetnames:
        ws = wb["직원현황"]
        current_company_id = None
        current_company_name = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0: continue  # 헤더
            cells = [str(x or '') for x in row]
            # 18컬럼까지 보존 (실제로는 0~10이 데이터, 11+ 는 거의 비어있음)
            col0 = cells[0] if len(cells) > 0 else ''
            # [건우건설] / [아이엔] 등 — 회사 블록 마커 ([] 있으면 무조건 회사 정보)
            if col0 and '[' in col0 and ']' in col0:
                m = re.search(r'\[([^\]]+)\]', col0)
                if m:
                    co_name_raw = m.group(1).strip()
                    cur_co_id = get_or_create_company(c, co_name_raw, strict=False)
                    if cur_co_id:
                        current_company_id = cur_co_id
                        current_company_name = co_name_raw
            if not current_company_id: continue

            name_raw = cells[1].strip() if len(cells) > 1 else ''
            if not name_raw: continue
            # 라벨 제거 후 "1.안 경 이" → 정규직 #1
            is_regular = bool(re.match(r'^\s*\*?\s*\d+\.', name_raw))
            regular_no = None
            m = re.match(r'^\s*\*?\s*(\d+)\.', name_raw)
            if m: regular_no = int(m.group(1))
            # "*김정길" — 별표 prefix 제거
            name_clean_raw = re.sub(r'^\s*\*\s*', '', name_raw)
            name = clean_name(name_clean_raw)
            if not name or len(name) < 2: continue
            if name in {'이름', '성명', '회사명', '직위'}: continue

            rrn_raw = cells[2].strip() if len(cells) > 2 else ''
            rrn = re.sub(r'\s+', '', rrn_raw)
            # "700920-1" 처럼 잘려있는 경우도 그대로 보존 (None 처리하지 말고 raw 저장)
            if rrn and not re.match(r'^\d{6}-\d{7}$', rrn):
                # 부분 RRN — 그대로 저장하되 표준 검증 통과 안 함 표시
                rrn_partial = rrn
                rrn = None
            else:
                rrn_partial = None

            # 입사일/퇴사일 멀티-라인 → 모든 날짜 추출, 마지막이 현재
            hired_dates = _split_dates(row[3] if len(row) > 3 else None)
            resigned_dates = _split_dates(row[4] if len(row) > 4 else None)
            hired_iso = hired_dates[-1] if hired_dates else None
            resigned_iso = resigned_dates[-1] if resigned_dates else None
            # 재입사 이력
            rehire_history = None
            if len(hired_dates) > 1 or len(resigned_dates) > 1:
                rehire_history = f"입사이력: {' → '.join(hired_dates)}"
                if resigned_dates:
                    rehire_history += f" / 퇴사이력: {' → '.join(resigned_dates)}"

            cert_main = cells[5].strip() if len(cells) > 5 else ''
            related = cells[6].strip() if len(cells) > 6 else ''
            asbestos_cell = cells[7].strip() if len(cells) > 7 else ''
            note_main = cells[8].strip() if len(cells) > 8 else ''
            note_extra1 = cells[9].strip() if len(cells) > 9 else ''
            note_extra2 = cells[10].strip() if len(cells) > 10 else ''
            note_extra3 = cells[11].strip() if len(cells) > 11 else ''

            # 8번 컬럼 분석 — 석면인력 / 사망 / 잡정보
            asbestos_flag = 1 if '석면' in asbestos_cell and '인력' in asbestos_cell else 0
            is_deceased = '사망' in asbestos_cell or '사망' in note_main
            asbestos_extra_note = ''
            if asbestos_cell and not asbestos_flag and '사망' not in asbestos_cell:
                # "대민충주구치소" / "새암입사" 같은 잡정보
                asbestos_extra_note = asbestos_cell

            # note — 사용자가 보일만한 핵심 메모만. 부속정보(메모2/메모3/계좌)는 제외
            note_parts = []
            if note_main: note_parts.append(note_main)
            if asbestos_extra_note: note_parts.append(asbestos_extra_note)
            if is_deceased: note_parts.append('★ 사망')
            if rrn_partial: note_parts.append(f"부분주민: {rrn_partial}")
            note_combined = ' / '.join(note_parts) if note_parts else None

            job_role = ('정규직' if is_regular else
                        ('퇴사자' if resigned_iso else '재직'))
            position = f"#{regular_no}" if regular_no else None

            # 같은 회사 + 같은 사람 (rrn 우선, 없으면 name)
            existing = None
            if rrn:
                existing = c.execute(
                    "SELECT id FROM workers WHERE rrn=? AND company_id=?",
                    (rrn, current_company_id)).fetchone()
            if not existing:
                existing = c.execute(
                    "SELECT id FROM workers WHERE name=? AND company_id=?",
                    (name, current_company_id)).fetchone()

            if existing:
                wid = existing[0]
                c.execute(
                    """UPDATE workers SET company_id=?, rrn=COALESCE(?,rrn),
                       hired_date=COALESCE(?,hired_date), resigned_at=?,
                       asbestos_certified=?, note=?, worker_type='office',
                       job_role=?, position=COALESCE(?,position)
                       WHERE id=?""",
                    (current_company_id, rrn, hired_iso, resigned_iso,
                     asbestos_flag, note_combined, job_role, position, wid))
                stats["workers_updated"] += 1
            else:
                cur = c.execute(
                    """INSERT INTO workers(company_id, name, rrn, worker_type, hired_date,
                       resigned_at, asbestos_certified, note, job_role, position,
                       job_specialty)
                       VALUES(?,?,?,'office',?,?,?,?,?,?,?)""",
                    (current_company_id, name, rrn, hired_iso, resigned_iso,
                     asbestos_flag, note_combined, job_role, position, related or None))
                wid = cur.lastrowid
                stats["workers_added"] += 1

            # 자격증 — 메인 컬럼
            if cert_main:
                for cert in parse_cert_field(cert_main):
                    if not cert.get('name'): continue
                    dup = c.execute(
                        "SELECT id FROM worker_certifications WHERE worker_id=? AND cert_name=? AND IFNULL(cert_level,'')=IFNULL(?,'')",
                        (wid, cert["name"], cert.get("level"))).fetchone()
                    if not dup:
                        c.execute(
                            """INSERT INTO worker_certifications(worker_id, cert_name, cert_level, related_business)
                               VALUES(?,?,?,?)""",
                            (wid, cert["name"], cert.get("level"),
                             related or None))
                        stats["certs_added"] += 1
            # 비고에서 추가 자격증 (예: "건축중급" "굴삭기,기중기운전기능사 (G00xxx)")
            for src in [note_main, note_extra1, note_extra2, note_extra3]:
                if not src: continue
                for nc in parse_note_certs(src):
                    if not nc.get('name'): continue
                    dup = c.execute(
                        "SELECT id FROM worker_certifications WHERE worker_id=? AND cert_name=?",
                        (wid, nc["name"])).fetchone()
                    if not dup:
                        c.execute(
                            """INSERT INTO worker_certifications(worker_id, cert_name, cert_level, cert_no, related_business, note)
                               VALUES(?,?,?,?,?,?)""",
                            (wid, nc["name"], nc.get("level"), nc.get("cert_no"),
                             related or None, '비고에서 추출'))
                        stats["certs_added"] += 1

    # ===== 주주명부 시트 (회사별) — 고정 컬럼 헤더 파서 =====
    # 9개 시트 (5 + 2 이전분 + 유신 + 직원현황) 중 회사 시트만 — SHEET_TO_COMPANY 사용
    for sn in wb.sheetnames:
        if sn not in SHEET_TO_COMPANY: continue
        ws = wb[sn]
        is_previous = sn.endswith("이전분")
        co_id = get_or_create_company(c, SHEET_TO_COMPANY[sn], strict=False)
        if not co_id: continue
        # R7 헤더 — 고정 컬럼 위치 파악
        col_idx = {}
        rows_data = list(ws.iter_rows(values_only=True))
        if len(rows_data) < 8: continue
        header = rows_data[6]  # R7 (0-indexed = 6)
        for ci, cell in enumerate(header or []):
            key = (str(cell or '')).replace(' ', '').strip()
            if key == '직위': col_idx['role'] = ci
            elif key == '성명': col_idx['name'] = ci
            elif '주민' in key: col_idx['rrn'] = ci
            elif '주소' in key or key == '주      주소': col_idx['address'] = ci
            elif '주식수' in key: col_idx['shares'] = ci
            elif key in {'비율', '비  율'} or '비' in key and '율' in key:
                col_idx['ratio'] = ci
        if 'name' not in col_idx: continue
        # R8~ 데이터 — '계' 행 만나면 종료
        for ri in range(7, len(rows_data)):
            row = rows_data[ri]
            if not row: continue
            cells = [str(x or '').strip() for x in row]
            # 종료 조건: '계' '합계' '소계' 셀 발견
            first_nonempty = next((c_val for c_val in cells if c_val), '')
            if first_nonempty in {'계', '합계', '소계'}: break
            name_raw = cells[col_idx['name']] if col_idx['name'] < len(cells) else ''
            name = name_raw.replace(' ', '').strip()
            if not name or len(name) < 2: continue
            if not re.match(r'^[가-힣]{2,4}$', name): continue
            if name in SHAREHOLDER_EXCLUDE: continue
            # 주민번호
            rrn = None
            if 'rrn' in col_idx and col_idx['rrn'] < len(cells):
                rrn_raw = cells[col_idx['rrn']]
                rrn = re.sub(r'\s+', '', rrn_raw)
                if not re.match(r'^\d{6}-\d{7}$', rrn or ''): rrn = None
            role = cells[col_idx['role']] if 'role' in col_idx and col_idx['role'] < len(cells) else None
            if role:
                role = role.strip()
                if role in {'직위', ''}: role = None
            address = cells[col_idx['address']] if 'address' in col_idx and col_idx['address'] < len(cells) else None
            if address: address = re.sub(r'\s+', ' ', address.replace('\n', ' ')).strip() or None
            # 지분율 — 0~1 (소수) 또는 0~100 둘다 허용
            ratio = None
            if 'ratio' in col_idx and col_idx['ratio'] < len(cells):
                rraw = cells[col_idx['ratio']].replace('%', '').strip()
                try:
                    v = float(rraw)
                    ratio = v * 100.0 if 0 < v <= 1 else v
                except: pass
            shares_count = None
            if 'shares' in col_idx and col_idx['shares'] < len(cells):
                try: shares_count = int(float(cells[col_idx['shares']]))
                except: pass
            # 직원과 매칭 시도 (rrn 우선, 그 다음 name+회사)
            w_row = None
            if rrn:
                w_row = c.execute("SELECT id FROM workers WHERE rrn=?", (rrn,)).fetchone()
            if not w_row:
                w_row = c.execute("SELECT id FROM workers WHERE name=? AND company_id=?",
                                  (name, co_id)).fetchone()
            # 중복 방지 (replace=False 일 때)
            dup = c.execute(
                "SELECT id FROM shareholders WHERE company_id=? AND name=? AND IFNULL(rrn,'')=IFNULL(?,'')",
                (co_id, name, rrn or '')).fetchone()
            if dup: continue
            c.execute(
                """INSERT INTO shareholders(company_id, name, role, rrn, address, shares_pct, contribution, worker_id)
                   VALUES(?,?,?,?,?,?,?,?)""",
                (co_id, name, role, rrn, address, ratio, shares_count,
                 w_row[0] if w_row else None))
            stats["shareholders"] += 1

        # ----- 추가 1) 회사 메타 (R14~R17 영역: 사업자번호 / 주소) -----
        for ri in range(13, min(18, len(rows_data))):
            row = rows_data[ri]
            if not row: continue
            full_text = ' '.join(str(x or '').strip() for x in row if x)
            if '사업자' in full_text and '등록' in full_text:
                m = re.search(r'(\d{3})\s*-\s*(\d{2})\s*-\s*(\d{5})', full_text)
                if m:
                    biz = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
                    c.execute(
                        "UPDATE companies SET business_no=COALESCE(NULLIF(business_no,''),?) WHERE id=?",
                        (biz, co_id))
                    stats["company_meta_updated"] += 1
            elif ('주' in full_text and '소' in full_text and
                  '주민' not in full_text and '주식' not in full_text):
                m = re.search(r'소\s*[:：]\s*(.+)$', full_text)
                if m:
                    addr = m.group(1).strip()
                    if addr and len(addr) > 5 and addr not in {'-', '없음'}:
                        c.execute(
                            "UPDATE companies SET address=COALESCE(NULLIF(address,''),?) WHERE id=?",
                            (addr, co_id))

        # ----- 추가 2) 사이드 면허 카탈로그 (이전분 시트는 면허 양식 다름 → 스킵) -----
        if not is_previous:
            lic_hdr_row = None
            lic_col_idx = {}
            for ri in range(15, min(20, len(rows_data))):
                row = rows_data[ri]
                if not row: continue
                cells_clean = [str(x or '').strip().replace(' ', '') for x in row]
                if any(k in cells_clean for k in ['종목', '보유종목']):
                    lic_hdr_row = ri
                    for ci, k in enumerate(cells_clean):
                        if k in ['종목', '보유종목']: lic_col_idx['type'] = ci
                        elif k in ['면허번호', '등록번호']: lic_col_idx['no'] = ci
                        elif '자본금' in k: lic_col_idx['cap'] = ci
                        elif '기술자' in k: lic_col_idx['tech'] = ci
                    break
            if lic_hdr_row is not None and 'type' in lic_col_idx:
                for ri in range(lic_hdr_row + 1, len(rows_data)):
                    row = rows_data[ri]
                    if not row: continue
                    cells = [str(x or '').strip() for x in row]
                    type_raw = cells[lic_col_idx['type']] if lic_col_idx['type'] < len(cells) else ''
                    type_clean = type_raw.replace(' ', '')
                    if not type_clean or type_clean in {'합계', '합', '소계', '계'}: continue
                    license_type = map_license_short(type_clean) or type_clean
                    license_no = None
                    if 'no' in lic_col_idx and lic_col_idx['no'] < len(cells):
                        no_raw = cells[lic_col_idx['no']].strip()
                        if no_raw and no_raw not in {'-', '없음'}:
                            license_no = no_raw
                    capacity_amount = 0
                    if 'cap' in lic_col_idx and lic_col_idx['cap'] < len(cells):
                        cap_raw = cells[lic_col_idx['cap']].replace('억','').replace(',','').strip()
                        try:
                            v = float(cap_raw)
                            capacity_amount = int(v * 100_000_000)  # 억 → 원
                        except: pass
                    existing_lic = c.execute(
                        "SELECT id FROM licenses WHERE company_id=? AND license_type=?",
                        (co_id, license_type)).fetchone()
                    if existing_lic:
                        c.execute(
                            """UPDATE licenses SET
                                license_no=COALESCE(NULLIF(license_no,''),?),
                                capacity_amount=CASE WHEN capacity_amount=0 THEN ? ELSE capacity_amount END
                                WHERE id=?""",
                            (license_no, capacity_amount, existing_lic[0]))
                        stats["licenses_updated"] += 1
                    else:
                        c.execute(
                            """INSERT INTO licenses(company_id, license_type, license_no,
                                                     capacity_amount, status)
                               VALUES(?,?,?,?,'active')""",
                            (co_id, license_type, license_no, capacity_amount))
                        stats["licenses_added"] += 1
    wb.close()
    return stats

# ====== 자격증 ↔ 면허 자동 매칭 (전체 직원 대상) ======
CERT_KEYWORD_TO_LICENSE = [
    ("토목",      ["토목공사업","토목건축공사업","지반조성·포장공사업","상·하수도설비공사업"]),
    ("건축",      ["건축공사업","토목건축공사업","실내건축공사업","철근·콘크리트공사업","도장·습식·방수·석공사업","금속창호·지붕건축물조립공사업"]),
    ("실내건축",  ["실내건축공사업"]),
    ("기계설비",  ["기계가스설비공사업","가스난방공사업"]),
    ("전기",      ["전기공사업"]),
    ("정보통신",  ["정보통신공사업"]),
    ("소방",      ["소방시설공사업"]),
    ("가스",      ["가스난방공사업","기계가스설비공사업"]),
    ("건설안전",  ["구조물해체·비계공사업"]),
    ("산업안전",  ["구조물해체·비계공사업"]),
    ("콘크리트",  ["철근·콘크리트공사업"]),
    ("굴삭기",    ["지반조성·포장공사업","구조물해체·비계공사업"]),
    ("굴착기",    ["지반조성·포장공사업","구조물해체·비계공사업"]),
    ("비계",      ["구조물해체·비계공사업"]),
    ("방수",      ["도장·습식·방수·석공사업"]),
    ("도장",      ["도장·습식·방수·석공사업"]),
    ("습식",      ["도장·습식·방수·석공사업"]),
    ("배관",      ["기계가스설비공사업","가스난방공사업","상·하수도설비공사업"]),
    ("석면",      ["석면해체·제거업"]),
]

def auto_register_workers_to_licenses(c):
    """모든 재직 직원의 자격증 ↔ 회사 면허 매칭 → license_workers 자동 등록."""
    rows = c.execute("""
        SELECT w.id, w.company_id, wc.cert_name FROM workers w
        JOIN worker_certifications wc ON wc.worker_id = w.id
        WHERE (w.resigned_at IS NULL OR w.resigned_at = '')
          AND w.company_id IS NOT NULL
    """).fetchall()
    auto_count = 0
    # 회사별 면허 캐시
    lic_cache = {}
    for wid, co_id, cert_name in rows:
        if not cert_name: continue
        if co_id not in lic_cache:
            lic_cache[co_id] = c.execute(
                "SELECT id, license_type FROM licenses WHERE company_id=? AND status='active'",
                (co_id,)).fetchall()
        for lic_id, lic_type in lic_cache[co_id]:
            for kw, lic_list in CERT_KEYWORD_TO_LICENSE:
                if kw in cert_name and lic_type in lic_list:
                    cur = c.execute(
                        """INSERT OR IGNORE INTO license_workers(license_id, worker_id, role, note)
                           VALUES(?,?,?,?)""",
                        (lic_id, wid, '기술인 (자동매칭)',
                         f'{cert_name} → {lic_type}'))
                    if cur.rowcount: auto_count += 1
                    break
    return auto_count

# ====== 자격증 파싱 (메인 + 비고) ======
def parse_cert_field(cert_str):
    """ 자격증 컬럼 → [{name, level}] 리스트.
    지원 케이스:
      - 단일:           '토목특급' → [{토목, 특급}]
      - 콤마/점/플러스:  '건축,토목 초급' / '건축.토목초급' / '건축+토목 초급' → 둘 다 초급
      - 멀티라인:        '건축일반산업기사\\n타일기능사' → 두 자격증
      - 마지막 레벨 전파: '건축, 토목 초급' → 건축도 초급
      - 단일 인정 패턴:  '토목기사(중급인정)' → {토목, 중급}
    """
    if not cert_str: return []
    s = str(cert_str).strip()
    if not s: return []
    # 길이순 — 산업기사 > 기술사 > 기능사 > 기사 (짧은 것이 긴 것을 잡아먹지 않게)
    levels = ['산업기사', '기술사', '기능사', '특급', '고급', '중급', '초급', '기사']
    # 단일 패턴 — "토목기사(중급인정)"
    m = re.match(r'^(\S+?)기사\(([특고중초]급)인정\)$', s)
    if m:
        return [{"name": m.group(1), "level": m.group(2)}]
    # 구분자: \n, ,, ., +, /, ·, 、
    parts = re.split(r'[\n\r,./+·、]+', s)
    parts = [p.strip() for p in parts if p.strip()]
    parsed = []
    for p in parts:
        level = None; name = p
        for lv in levels:
            if lv in p:
                level = lv
                name = p.replace(lv, '').strip()
                break
        if not name: name = p
        parsed.append({"name": name, "level": level})
    # 마지막 레벨 → 앞쪽 None entry 에 전파 (명시된 레벨은 보존)
    last_level = None
    for entry in reversed(parsed):
        if entry["level"]:
            last_level = entry["level"]
        elif last_level:
            entry["level"] = last_level
    return parsed

def parse_note_certs(note):
    """비고 컬럼에서 추가 자격증 추출."""
    if not note: return []
    s = str(note).strip()
    if not s: return []
    out = []
    for m in re.finditer(r'([가-힣\w]+?)(기능사|산업기사|기사|기술사|특급|고급|중급|초급)\s*\(([A-Z0-9]+)\)', s):
        out.append({"name": m.group(1).strip(), "level": m.group(2).strip(), "cert_no": m.group(3).strip()})
    if not out and '운전기능사' in s:
        for piece in re.split(r'[,、/]', s):
            p = piece.strip()
            m2 = re.match(r'^([가-힣]+)\s*운전기능사', p)
            if m2:
                out.append({"name": m2.group(1) + '운전', "level": "기능사", "cert_no": None})
    return out

def parse_license_no_pair(no_str):
    """'G00629876 (#0630375)' → ('G00629876', '0630375')"""
    if not no_str: return (None, None)
    s = str(no_str).strip()
    main = None; sub = None
    m = re.match(r'^([A-Z0-9-]+)', s)
    if m: main = m.group(1)
    m2 = re.search(r'\(#?([A-Z0-9]+)\)', s)
    if m2: sub = m2.group(1)
    if not main:
        m3 = re.search(r'[A-Z0-9]{8,}', s)
        if m3: main = m3.group(0)
    return (main, sub)

# ====== 2) 회사별 아웃라인 — 회사 정보 보강 + 기술자 정밀 임포트 ======
def import_company_outline(c, path):
    if not path.exists():
        print(f"  [skip] {path.name} 없음")
        return {"companies_updated": 0, "tech_added": 0, "auto_registered": 0}
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    stats = {"companies_updated": 0, "tech_added": 0, "auto_registered": 0}

    # ===== 아웃라인(2) — 사업자번호·국민연금·결산일 =====
    if "아웃라인(2)" in wb.sheetnames:
        ws = wb["아웃라인(2)"]
        # R2: 회사명들, R3: 국민연금사업장기호 (사업자번호와 비슷)
        rows_data = list(ws.iter_rows(values_only=True))
        if len(rows_data) >= 3:
            company_cells = rows_data[1] if len(rows_data) > 1 else []
            for ci, name_cell in enumerate(company_cells):
                if not name_cell: continue
                co_id = get_or_create_company(c, name_cell)
                if not co_id: continue
                # 다른 행에서 정보 찾기
                for ri, row in enumerate(rows_data):
                    if ri < 2: continue
                    if not row or len(row) <= ci: continue
                    label = str(row[0] or '').strip() if row else ''
                    val = row[ci] if ci < len(row) else None
                    if not val: continue
                    val_s = str(val).strip()
                    if '사업자' in label or '국민연금' in label:
                        # 사업자번호 형식 추출 (xxx-xx-xxxxx)
                        m = re.search(r'\d{3}-\d{2}-\d{5}', val_s)
                        if m:
                            biz = m.group(0)
                            c.execute(
                                "UPDATE companies SET business_no=COALESCE(business_no, ?) WHERE id=? AND (business_no IS NULL OR business_no='')",
                                (biz, co_id))
                    elif '결산' in label or '결산일' in label:
                        m = re.search(r'(\d{1,2})월', val_s)
                        if m:
                            c.execute("UPDATE companies SET fiscal_year_end=COALESCE(fiscal_year_end,?) WHERE id=?",
                                      (m.group(1), co_id))
                    elif '소재' in label or '주소' in label:
                        c.execute("UPDATE companies SET address=COALESCE(address,?) WHERE id=?", (val_s, co_id))
                stats["companies_updated"] += 1

    # ===== 기술자 시트 — 정밀 임포트 (메인+비고 자격증 + 자동 면허 등재) =====
    if "기술자" in wb.sheetnames:
        ws = wb["기술자"]
        current_co_id = None
        in_resignation_block = False  # "퇴사현황" 블록 여부
        rows_data = list(ws.iter_rows(values_only=True))
        # 컬럼 인덱스 — 시트 R5 헤더 기반 (고정)
        # 0:성명 1:직책 2:주민번호 3:주소 4:자격종목 5:등록번호 6:취득일 7:입사일 8:퇴사일 9:급여 10:연락처 11:비고
        col = {'name': 0, 'position': 1, 'rrn': 2, 'address': 3,
               'cert_main': 4, 'cert_no': 5, 'acquired': 6,
               'hired': 7, 'resigned': 8, 'salary': 9, 'phone': 10, 'note': 11}

        for ri, row in enumerate(rows_data):
            if not row: continue
            cells = [str(x or '').strip() for x in row]
            first = cells[0] if cells else ''
            # "퇴사현황" 블록 진입
            if '퇴사' in first and '현황' in first:
                in_resignation_block = True
                continue
            # 회사명 행 감지 — 건설·아이엔이 cell[0] 에 있고 다른 셀들은 비어있음
            if first and ('건설' in first or '아이엔' in first or '유신' in first or '인우' in first or '새암' in first):
                # 헤더가 아닌 첫 셀에 회사명만 있는 행
                if all(not c for c in cells[1:8]):
                    cid = get_or_create_company(c, first, strict=False)
                    if cid:
                        current_co_id = cid
                        in_resignation_block = False
                    continue
            # 헤더 행 ('성명') 스킵
            if first == '성명': continue
            # 데이터 행 — current_co_id 필요
            if not current_co_id: continue
            if not first: continue
            name = clean_name(first)
            if not name: continue

            rrn = re.sub(r'\s+', '', cells[col['rrn']]) if col['rrn'] < len(cells) else None
            position = cells[col['position']] if col['position'] < len(cells) else None
            address = cells[col['address']] if col['address'] < len(cells) else None
            cert_main = cells[col['cert_main']] if col['cert_main'] < len(cells) else None
            cert_no_raw = cells[col['cert_no']] if col['cert_no'] < len(cells) else None
            acquired = normalize_date(row[col['acquired']]) if col['acquired'] < len(row) else None
            hired = normalize_date(row[col['hired']]) if col['hired'] < len(row) else None
            resigned = normalize_date(row[col['resigned']]) if col['resigned'] < len(row) else None
            salary_raw = row[col['salary']] if col['salary'] < len(row) else None
            phone = cells[col['phone']] if col['phone'] < len(cells) else None
            note = cells[col['note']] if col['note'] < len(cells) else None

            # 직원 찾기/생성 (정규직 사무직)
            existing = None
            if rrn:
                existing = c.execute("SELECT id FROM workers WHERE name=? AND rrn=?", (name, rrn)).fetchone()
            if not existing:
                existing = c.execute("SELECT id FROM workers WHERE name=? AND company_id=?",
                                     (name, current_co_id)).fetchone()
            if existing:
                wid = existing[0]
                c.execute("""UPDATE workers SET
                    company_id=COALESCE(?,company_id),
                    rrn=COALESCE(NULLIF(rrn,''),?),
                    address=COALESCE(NULLIF(address,''),?),
                    position=COALESCE(NULLIF(position,''),?),
                    phone=COALESCE(NULLIF(phone,''),?),
                    hired_date=COALESCE(hired_date,?),
                    resigned_at=COALESCE(?,resigned_at),
                    worker_type='office'
                    WHERE id=?""",
                    (current_co_id, rrn, address, position, phone, hired,
                     resigned if in_resignation_block else None, wid))
            else:
                cur = c.execute(
                    """INSERT INTO workers(company_id,name,rrn,address,position,phone,hired_date,
                                            resigned_at,worker_type)
                       VALUES(?,?,?,?,?,?,?,?,'office')""",
                    (current_co_id, name, rrn, address, position, phone, hired,
                     resigned if in_resignation_block else None))
                wid = cur.lastrowid

            # 메인 자격증 파싱 + 등록번호 분리
            (cert_no_main, cert_no_sub) = parse_license_no_pair(cert_no_raw)
            main_certs = parse_cert_field(cert_main)
            for cert in main_certs:
                # 중복 방지 (같은 사람·자격명·레벨)
                dup = c.execute(
                    "SELECT id FROM worker_certifications WHERE worker_id=? AND cert_name=? AND IFNULL(cert_level,'')=IFNULL(?,'')",
                    (wid, cert["name"], cert.get("level"))).fetchone()
                if not dup:
                    c.execute(
                        """INSERT INTO worker_certifications(worker_id,cert_name,cert_level,cert_no,acquired_at,note)
                           VALUES(?,?,?,?,?,?)""",
                        (wid, cert["name"], cert.get("level"), cert_no_main, acquired,
                         f"sub: {cert_no_sub}" if cert_no_sub else None))
                    stats["tech_added"] += 1
                else:
                    # 등록번호 보강
                    c.execute(
                        "UPDATE worker_certifications SET cert_no=COALESCE(cert_no,?), acquired_at=COALESCE(acquired_at,?) WHERE id=?",
                        (cert_no_main, acquired, dup[0]))

            # 비고 칼럼에서 추가 자격증 추출
            for nc in parse_note_certs(note):
                dup = c.execute(
                    "SELECT id FROM worker_certifications WHERE worker_id=? AND cert_name=?",
                    (wid, nc["name"])).fetchone()
                if not dup:
                    c.execute(
                        """INSERT INTO worker_certifications(worker_id,cert_name,cert_level,cert_no,note)
                           VALUES(?,?,?,?,?)""",
                        (wid, nc["name"], nc.get("level"), nc.get("cert_no"), '비고에서 추출'))
                    stats["tech_added"] += 1

            # 석면 인력 표시 (cell 12, 13 등에서 '석면' 검출)
            asbestos = any('석면' in (cells[k] or '') for k in range(12, min(15, len(cells))))
            if asbestos:
                c.execute("UPDATE workers SET asbestos_certified=1 WHERE id=?", (wid,))

            # 자동 면허 등재 — 이 직원이 이 회사의 어느 면허 요건에 매칭되는지
            # 이 회사의 모든 면허에 대해, 직원 자격증과 매칭되면 license_workers에 추가
            if not in_resignation_block:
                co_licenses = c.execute(
                    "SELECT id, license_type FROM licenses WHERE company_id=? AND status='active'",
                    (current_co_id,)).fetchall()
                worker_certs = c.execute(
                    "SELECT cert_name, cert_level FROM worker_certifications WHERE worker_id=?",
                    (wid,)).fetchall()
                for lic_id, lic_type in co_licenses:
                    matched = False
                    for cert_name, cert_level in worker_certs:
                        # 키워드 매칭 — CERT_TO_LICENSE_MAP 단순 적용 (포함 검사)
                        for key, lvl_req, lic_list in [
                            ("토목", None, ["토목공사업","토목건축공사업","지반조성·포장공사업","상·하수도설비공사업"]),
                            ("건축", None, ["건축공사업","토목건축공사업","실내건축공사업","철근·콘크리트공사업","도장·습식·방수·석공사업","금속창호·지붕건축물조립공사업"]),
                            ("기계설비", None, ["기계가스설비공사업","가스난방공사업"]),
                            ("전기", None, ["전기공사업"]),
                            ("정보통신", None, ["정보통신공사업"]),
                            ("소방", None, ["소방시설공사업"]),
                            ("가스", None, ["가스난방공사업","기계가스설비공사업"]),
                            ("건설안전", None, ["구조물해체·비계공사업"]),
                            ("산업안전", None, ["구조물해체·비계공사업"]),
                            ("콘크리트", None, ["철근·콘크리트공사업"]),
                            ("굴삭기", None, ["지반조성·포장공사업","구조물해체·비계공사업"]),
                            ("굴착기", None, ["지반조성·포장공사업","구조물해체·비계공사업"]),
                            ("비계", None, ["구조물해체·비계공사업"]),
                            ("방수", None, ["도장·습식·방수·석공사업"]),
                            ("석면", None, ["석면해체·제거업"]),
                        ]:
                            if key in (cert_name or '') and lic_type in lic_list:
                                matched = True; break
                        if matched: break
                    if matched:
                        try:
                            c.execute(
                                """INSERT OR IGNORE INTO license_workers(license_id,worker_id,role,note)
                                   VALUES(?,?,?,?)""",
                                (lic_id, wid, '기술인 (엑셀 자동매칭)', '엑셀 임포트 시 자격 매칭으로 자동 등재'))
                            if c.execute("SELECT changes()").fetchone()[0]:
                                stats["auto_registered"] += 1
                        except Exception: pass
    wb.close()
    return stats

# ====== 3) 건우건설 급여대장 — 직원명부 보강 ======
def import_payroll_directory(c, path):
    if not path.exists():
        return {"workers_updated": 0}
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    stats = {"workers_updated": 0}
    if "직원명부" in wb.sheetnames:
        ws = wb["직원명부"]
        co_id = get_or_create_company(c, "건우건설")
        # 헤더: 번호, 성명, 직책, 주민등록번호, 주소, 연락번호, e-mail, 예금주, 은행명, 계좌, 입사일, 퇴사일
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 1: continue  # R1 = 헤더
            cells = [str(x).strip() if x is not None else '' for x in row]
            if len(cells) < 5 or not cells[1]: continue
            name = clean_name(cells[1])
            position = cells[2] if len(cells) > 2 else ''
            rrn = re.sub(r'\s+', '', cells[3]) if len(cells) > 3 else ''
            address = cells[4] if len(cells) > 4 else ''
            phone = cells[5] if len(cells) > 5 else ''
            account_holder = cells[7] if len(cells) > 7 else ''
            bank_name = cells[8] if len(cells) > 8 else ''
            bank_account = cells[9] if len(cells) > 9 else ''
            hired = normalize_date(cells[10]) if len(cells) > 10 else None
            resigned = normalize_date(cells[11]) if len(cells) > 11 else None
            if not name: continue
            w = c.execute("SELECT id FROM workers WHERE name=? AND company_id=?", (name, co_id)).fetchone()
            if not w:
                cur = c.execute(
                    """INSERT INTO workers(company_id, name, rrn, phone, address, position,
                       bank_name, account_holder, bank_account, hired_date, resigned_at, worker_type)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,'office')""",
                    (co_id, name, rrn, phone, address, position, bank_name,
                     account_holder, bank_account, hired, resigned))
                stats["workers_updated"] += 1
            else:
                c.execute(
                    """UPDATE workers SET
                        rrn=COALESCE(NULLIF(rrn,''),?), phone=COALESCE(NULLIF(phone,''),?),
                        address=COALESCE(NULLIF(address,''),?), position=COALESCE(NULLIF(position,''),?),
                        bank_name=COALESCE(NULLIF(bank_name,''),?), account_holder=COALESCE(NULLIF(account_holder,''),?),
                        bank_account=COALESCE(NULLIF(bank_account,''),?),
                        hired_date=COALESCE(hired_date,?), resigned_at=COALESCE(resigned_at,?)
                        WHERE id=?""",
                    (rrn, phone, address, position, bank_name, account_holder, bank_account,
                     hired, resigned, w[0]))
                stats["workers_updated"] += 1
    wb.close()
    return stats

# ====== 메인 ======
def import_all(db_path=None, replace=False):
    """엑셀 일괄 임포트.
    replace=True 면 기존 직원/자격증/주주/면허등재를 모두 와이프하고 다시 구축.
    """
    if not openpyxl:
        return {"error": "openpyxl 미설치"}
    if db_path is None:
        from app import DB_PATH
        db_path = DB_PATH
    c = sqlite3.connect(db_path)
    c.row_factory = sqlite3.Row
    c.execute("PRAGMA foreign_keys = ON")

    report = {"replace_mode": replace}
    print("=" * 60)
    print(f"📥 1) 직원,주주명부 임포트  (replace={replace})")
    report["employees"] = import_employee_directory(
        c, DATA_DIR / "직원,주주명부-5개회사 (자동 저장됨).xlsx",
        replace=replace)
    print(f"  → {report['employees']}")

    print("\n📥 2) 회사별 아웃라인 임포트 (회사정보 + 기술자)")
    report["outline"] = import_company_outline(
        c, DATA_DIR / "회사별아웃라인-(기술인력보유현황포함).xlsx")
    print(f"  → {report['outline']}")

    print("\n📥 3) 건우건설 급여대장 — 직원명부 보강")
    report["payroll"] = import_payroll_directory(
        c, DATA_DIR / "건우건설급여대장.xlsx")
    print(f"  → {report['payroll']}")

    print("\n🔗 4) 자격증 ↔ 면허 자동 매칭 (전체 직원 대상)")
    auto_n = auto_register_workers_to_licenses(c)
    report["auto_register_total"] = auto_n
    print(f"  → 자동 등재 추가: {auto_n}건")

    c.commit()
    # 최종 카운트
    final = {
        "companies":        c.execute("SELECT COUNT(*) FROM companies").fetchone()[0],
        "workers_total":    c.execute("SELECT COUNT(*) FROM workers").fetchone()[0],
        "workers_office":   c.execute("SELECT COUNT(*) FROM workers WHERE worker_type='office'").fetchone()[0],
        "workers_active":   c.execute("SELECT COUNT(*) FROM workers WHERE resigned_at IS NULL OR resigned_at=''").fetchone()[0],
        "workers_resigned": c.execute("SELECT COUNT(*) FROM workers WHERE resigned_at IS NOT NULL AND resigned_at!=''").fetchone()[0],
        "certifications":   c.execute("SELECT COUNT(*) FROM worker_certifications").fetchone()[0],
        "shareholders":     c.execute("SELECT COUNT(*) FROM shareholders").fetchone()[0],
        "license_workers":  c.execute("SELECT COUNT(*) FROM license_workers").fetchone()[0],
        "asbestos_certified": c.execute("SELECT COUNT(*) FROM workers WHERE asbestos_certified=1").fetchone()[0],
    }
    report["final"] = final
    c.close()
    print("\n" + "=" * 60)
    print("✅ 임포트 완료")
    for k, v in final.items():
        print(f"  {k}: {v}")
    return report

if __name__ == "__main__":
    import sys
    replace = '--replace' in sys.argv
    r = import_all(replace=replace)
    print("\n[REPORT]", json.dumps(r, ensure_ascii=False, indent=2, default=str))
